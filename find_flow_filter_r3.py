
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

file_path = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_Dashboard_v25_direct_let.xlsx'
wb = load_workbook(file_path, data_only=True)
ws = wb['📋 T_RawData']

print("=== Scanning Row 3 for Headers ===")
for cell in ws[3]:
    if cell.value:
        print(f"{cell.value} -> {get_column_letter(cell.column)}")
