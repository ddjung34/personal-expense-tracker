import openpyxl
from openpyxl.utils import get_column_letter

FILE_PATH = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-07~2025-12-07_v2.xlsx"

SAMPLE_DATA = [
    ["2024-12-08", "12:00", "지출", "식비", "점심", 10000, "신한카드", "김밥천국", "테스트 데이터"],
    ["2024-12-09", "18:00", "지출", "쇼핑", "의류", 50000, "현대카드", "유니클로", "겨울옷 구매"],
    ["2024-12-10", "10:00", "수입", "급여", "정기", 3000000, "은행", "회사", "12월 월급"]
]

def process_file():
    print(f"Loading {FILE_PATH}...")
    wb = openpyxl.load_workbook(FILE_PATH)
    
    # 1. Add Sample Data to DB_Raw
    ws_data = wb["DB_Raw"]
    print("Adding sample data/transactions...")
    
    # Find first empty row (start after header)
    # If sheet is new, max_row might be 1.
    start_row = ws_data.max_row + 1
    
    for i, data in enumerate(SAMPLE_DATA):
        for col, value in enumerate(data, 1):
            ws_data.cell(row=start_row + i, column=col, value=value)
            
    # 2. Add Formulas to Report Sheet ('뱅샐현황')
    if "뱅샐현황" in wb.sheetnames:
        ws_report = wb["뱅샐현황"]
        print("Applying formulas to '뱅샐현황'...")
        
        # Based on previous analysis:
        # Header Row: 11
        # Category Column: B (2)
        # December 2024 Column: E (5)
        
        CATEGORY_COL_IDX = 2
        DEC_COL_IDX = 5
        HEADER_ROW = 11
        
        # Iterate rows starting from 12
        for row in range(HEADER_ROW + 1, ws_report.max_row + 1):
            category_cell = ws_report.cell(row=row, column=CATEGORY_COL_IDX)
            target_cell = ws_report.cell(row=row, column=DEC_COL_IDX)
            
            category_name = category_cell.value
            
            # Only apply if there is a category name and it's a string (skip empty or sums for now if ambiguous)
            # Also avoiding "총계" (Total) or subheaders if possible, but for now apply to all valid text
            if category_name and isinstance(category_name, str):
                # Valid categories usually don't have large merged titles in B, assuming standard list
                
                # Formula: =SUMIFS(DB_Raw!F:F, DB_Raw!D:D, "CategoryName", DB_Raw!A:A, ">=2024-12-01", DB_Raw!A:A, "<=2024-12-31")
                # Note: Adjust column letters for DB_Raw if they changed. 
                # F=Amount, D=MainCategory, A=Date
                
                # Using cell references for category is better: D:D, B12
                cat_ref = f"{get_column_letter(CATEGORY_COL_IDX)}{row}" # e.g. B12
                
                formula = f'=SUMIFS(DB_Raw!$F:$F, DB_Raw!$D:$D, {cat_ref}, DB_Raw!$A:$A, ">=2024-12-01", DB_Raw!$A:$A, "<=2024-12-31")'
                
                # Check if cell is part of a merged range
                from openpyxl.cell.cell import Cell, MergedCell
                
                if isinstance(target_cell, MergedCell):
                    print(f"Skipping merged cell at {get_column_letter(DEC_COL_IDX)}{row} (Not top-left)")
                    continue
                
                # Double check if it is the top-left of a merge (openpyxl usually returns Cell for top-left, MergedCell for others)
                # But sometimes it's trickier. Let's just try/except.
                try:
                    target_cell.value = formula
                    print(f"Applied formula at {get_column_letter(DEC_COL_IDX)}{row} for '{category_name}'")
                except AttributeError:
                    print(f"Could not write to {get_column_letter(DEC_COL_IDX)}{row} - likely merged.")
                
    else:
        print("Warning: Sheet '뱅샐현황' not found!")

    print("Saving...")
    wb.save(FILE_PATH)
    print("Done.")

if __name__ == "__main__":
    process_file()
