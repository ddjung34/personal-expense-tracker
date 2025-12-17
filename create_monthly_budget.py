import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict

INPUT_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-10~2025-12-10.xlsx"
OUTPUT_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\{date}_월별상세가계부.xlsx".format(
    date=datetime.now().strftime("%Y%m%d")
)

def create_monthly_detailed_budget():
    print(f"📊 Loading data from: {INPUT_FILE}")
    
    # Read data
    df = pd.read_excel(INPUT_FILE, sheet_name='Rawdata', engine='openpyxl')
    
    # Data cleaning
    df['date'] = pd.to_datetime(df['날짜'], errors='coerce')
    df['type'] = df['타입']
    df['main_category'] = df['대분류'].fillna('기타')
    df['sub_category'] = df['소분류'].fillna('기타')
    df['amount'] = pd.to_numeric(df['금액'], errors='coerce')
    
    df = df.dropna(subset=['date', 'amount', 'type'])
    
    print(f"✅ Loaded {len(df)} transactions")
    
    # Get date range
    start_date = df['date'].min()
    end_date = df['date'].max()
    
    # Group data by month
    df['year_month'] = df['date'].dt.to_period('M')
    months = sorted(df['year_month'].unique())
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Color scheme matching the image
    HEADER_COLOR = "4472C4"  # Blue
    SUB_HEADER_COLOR = "D9E1F2"  # Light blue
    INCOME_COLOR = "70AD47"  # Green
    EXPENSE_COLOR = "ED7D31"  # Orange
    TOTAL_COLOR = "FFC000"  # Yellow
    
    # ==================================================
    # Create Monthly Detail Sheets
    # ==================================================
    
    for month_period in months:
        month_str = str(month_period)
        month_df = df[df['year_month'] == month_period]
        
        # Create sheet
        ws = wb.create_sheet(f"{month_str}", len(wb.sheetnames))
        ws.sheet_view.showGridLines = False
        
        # Title
        ws['B2'] = f"{month_str} 가계부"
        ws['B2'].font = Font(size=18, bold=True, color=HEADER_COLOR, name="맑은 고딕")
        ws.merge_cells('B2:F2')
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 30
        
        # Income Section
        row = 4
        ws[f'B{row}'] = "수입"
        ws[f'B{row}'].fill = PatternFill(start_color=INCOME_COLOR, end_color=INCOME_COLOR, fill_type="solid")
        ws[f'B{row}'].font = Font(bold=True, color="FFFFFF", size=12, name="맑은 고딕")
        ws.merge_cells(f'B{row}:F{row}')
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
        
        row += 1
        # Headers
        headers = ['대분류', '소분류', '금액', '비고']
        for col_idx, header in enumerate(headers, 2):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = PatternFill(start_color=SUB_HEADER_COLOR, end_color=SUB_HEADER_COLOR, fill_type="solid")
            cell.font = Font(bold=True, size=10, name="맑은 고딕")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        # Income data
        income_df = month_df[month_df['type'] == '수입']
        income_start_row = row + 1
        
        if not income_df.empty:
            for idx, transaction in enumerate(income_df.itertuples(), row + 1):
                ws[f'B{idx}'] = transaction.main_category
                ws[f'C{idx}'] = transaction.sub_category
                ws[f'D{idx}'] = transaction.amount
                ws[f'D{idx}'].number_format = '₩#,##0'
                ws[f'E{idx}'] = ''  # 비고
                
                for col in ['B', 'C', 'D', 'E']:
                    ws[f'{col}{idx}'].font = Font(name="맑은 고딕", size=10)
                    ws[f'{col}{idx}'].border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    ws[f'{col}{idx}'].alignment = Alignment(horizontal='center' if col != 'E' else 'left', vertical='center')
            
            row = income_start_row + len(income_df)
        else:
            row += 1
        
        # Income Total
        ws[f'B{row}'] = "수입 합계"
        ws[f'B{row}'].font = Font(bold=True, size=11, name="맑은 고딕")
        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{row}'].fill = PatternFill(start_color=TOTAL_COLOR, end_color=TOTAL_COLOR, fill_type="solid")
        
        # Formula for income total
        if not income_df.empty:
            ws[f'D{row}'] = f"=SUM(D{income_start_row}:D{row-1})"
        else:
            ws[f'D{row}'] = 0
        ws[f'D{row}'].number_format = '₩#,##0'
        ws[f'D{row}'].font = Font(bold=True, size=11, name="맑은 고딕")
        ws[f'D{row}'].fill = PatternFill(start_color=TOTAL_COLOR, end_color=TOTAL_COLOR, fill_type="solid")
        ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        for col in ['B', 'C', 'D']:
            ws[f'{col}{row}'].border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='medium'), bottom=Side(style='medium')
            )
        
        income_total_row = row
        
        # Expense Section
        row += 2
        ws[f'B{row}'] = "지출"
        ws[f'B{row}'].fill = PatternFill(start_color=EXPENSE_COLOR, end_color=EXPENSE_COLOR, fill_type="solid")
        ws[f'B{row}'].font = Font(bold=True, color="FFFFFF", size=12, name="맑은 고딕")
        ws.merge_cells(f'B{row}:F{row}')
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
        
        row += 1
        # Headers
        for col_idx, header in enumerate(headers, 2):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = PatternFill(start_color=SUB_HEADER_COLOR, end_color=SUB_HEADER_COLOR, fill_type="solid")
            cell.font = Font(bold=True, size=10, name="맑은 고딕")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        # Expense data
        expense_df = month_df[month_df['type'] == '지출']
        expense_start_row = row + 1
        
        if not expense_df.empty:
            for idx, transaction in enumerate(expense_df.itertuples(), row + 1):
                ws[f'B{idx}'] = transaction.main_category
                ws[f'C{idx}'] = transaction.sub_category
                ws[f'D{idx}'] = transaction.amount
                ws[f'D{idx}'].number_format = '₩#,##0'
                ws[f'E{idx}'] = ''  # 비고
                
                for col in ['B', 'C', 'D', 'E']:
                    ws[f'{col}{idx}'].font = Font(name="맑은 고딕", size=10)
                    ws[f'{col}{idx}'].border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    ws[f'{col}{idx}'].alignment = Alignment(horizontal='center' if col != 'E' else 'left', vertical='center')
            
            row = expense_start_row + len(expense_df)
        else:
            row += 1
        
        # Expense Total
        ws[f'B{row}'] = "지출 합계"
        ws[f'B{row}'].font = Font(bold=True, size=11, name="맑은 고딕")
        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{row}'].fill = PatternFill(start_color=TOTAL_COLOR, end_color=TOTAL_COLOR, fill_type="solid")
        
        # Formula for expense total
        if not expense_df.empty:
            ws[f'D{row}'] = f"=SUM(D{expense_start_row}:D{row-1})"
        else:
            ws[f'D{row}'] = 0
        ws[f'D{row}'].number_format = '₩#,##0'
        ws[f'D{row}'].font = Font(bold=True, size=11, name="맑은 고딕")
        ws[f'D{row}'].fill = PatternFill(start_color=TOTAL_COLOR, end_color=TOTAL_COLOR, fill_type="solid")
        ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        for col in ['B', 'C', 'D']:
            ws[f'{col}{row}'].border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='medium'), bottom=Side(style='medium')
            )
        
        expense_total_row = row
        
        # Monthly Net Income
        row += 2
        ws[f'B{row}'] = "월 순수익"
        ws[f'B{row}'].font = Font(bold=True, size=12, name="맑은 고딕", color="FFFFFF")
        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        ws[f'D{row}'] = f"=D{income_total_row}-D{expense_total_row}"
        ws[f'D{row}'].number_format = '₩#,##0'
        ws[f'D{row}'].font = Font(bold=True, size=12, name="맑은 고딕", color="FFFFFF")
        ws[f'D{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 30
        
        for col in ['B', 'C', 'D']:
            ws[f'{col}{row}'].border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='medium'), bottom=Side(style='medium')
            )
        
        # Auto-fit columns
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 30
    
    # ==================================================
    # Create Annual Summary Sheet
    # ==================================================
    ws_summary = wb.create_sheet("연간 요약", 0)
    ws_summary.sheet_view.showGridLines = False
    
    # Title
    ws_summary['B2'] = f"{start_date.year}년 연간 가계부 요약"
    ws_summary['B2'].font = Font(size=20, bold=True, color=HEADER_COLOR, name="맑은 고딕")
    ws_summary.merge_cells('B2:N2')
    ws_summary['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.row_dimensions[2].height = 35
    
    # Monthly summary table
    row = 5
    ws_summary[f'B{row}'] = "월"
    ws_summary[f'C{row}'] = "수입"
    ws_summary[f'D{row}'] = "지출"
    ws_summary[f'E{row}'] = "순수익"
    
    for col in ['B', 'C', 'D', 'E']:
        ws_summary[f'{col}{row}'].fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
        ws_summary[f'{col}{row}'].font = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")
        ws_summary[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws_summary[f'{col}{row}'].border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )
    ws_summary.row_dimensions[row].height = 25
    
    # Monthly data with formulas
    for idx, month_period in enumerate(months, row + 1):
        month_str = str(month_period)
        ws_summary[f'B{idx}'] = month_str
        
        # Formulas referencing monthly sheets
        ws_summary[f'C{idx}'] = f"='{month_str}'!D{income_total_row}"  # Will update with actual row
        ws_summary[f'D{idx}'] = f"='{month_str}'!D{expense_total_row}"
        ws_summary[f'E{idx}'] = f"=C{idx}-D{idx}"
        
        for col in ['B', 'C', 'D', 'E']:
            ws_summary[f'{col}{idx}'].font = Font(name="맑은 고딕", size=10)
            ws_summary[f'{col}{idx}'].alignment = Alignment(horizontal='center', vertical='center')
            ws_summary[f'{col}{idx}'].border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            if col in ['C', 'D', 'E']:
                ws_summary[f'{col}{idx}'].number_format = '₩#,##0'
    
    # Annual totals
    total_row = row + len(months) + 1
    ws_summary[f'B{total_row}'] = "연간 합계"
    ws_summary[f'B{total_row}'].fill = PatternFill(start_color=TOTAL_COLOR, end_color=TOTAL_COLOR, fill_type="solid")
    ws_summary[f'B{total_row}'].font = Font(bold=True, size=12, name="맑은 고딕")
    ws_summary[f'B{total_row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    for col, col_letter in zip(['C', 'D', 'E'], ['C', 'D', 'E']):
        ws_summary[f'{col}{total_row}'] = f"=SUM({col}{row+1}:{col}{row+len(months)})"
        ws_summary[f'{col}{total_row}'].number_format = '₩#,##0'
        ws_summary[f'{col}{total_row}'].fill = PatternFill(start_color=TOTAL_COLOR, end_color=TOTAL_COLOR, fill_type="solid")
        ws_summary[f'{col}{total_row}'].font = Font(bold=True, size=12, name="맑은 고딕")
        ws_summary[f'{col}{total_row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    for col in ['B', 'C', 'D', 'E']:
        ws_summary[f'{col}{total_row}'].border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )
    ws_summary.row_dimensions[total_row].height = 30
    
    # Auto-fit columns
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 20
    ws_summary.column_dimensions['D'].width = 20
    ws_summary.column_dimensions['E'].width = 20
    
    # Save
    print(f"💾 Saving to: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    print("✅ Monthly Detailed Budget created successfully!")
    print(f"\n📊 Summary:")
    print(f"   - 연간 요약 시트: 1개")
    print(f"   - 월별 상세 시트: {len(months)}개")
    print(f"   - 모든 금액: 수식 기반 (₩#,##0 형식)")
    print(f"\n✅ File saved: {OUTPUT_FILE}")

if __name__ == "__main__":
    create_monthly_detailed_budget()
