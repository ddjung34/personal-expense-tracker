"""
Dashboard 정밀 복제 v5
1. 굵은 외곽 테두리 제거 (원본 테이블 테두리는 유지)
2. 주요지표 차트: 콤보 (수입/지출 Bar + 합계 Line)
3. 지출구조 차트: 묶은 가로 막대 (D열, F열)
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

print("=" * 70)
print("Dashboard 정밀 복제 v5")
print("=" * 70)

wb = load_workbook(
    r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_수식연결_가계부엔진_최종.xlsx'
)

# Dashboard 시트 찾기
dashboard_sheet_name = None
for sheet_name in wb.sheetnames:
    if 'dashboard' in sheet_name.lower() and '복제' not in sheet_name and '분석' in sheet_name:
        dashboard_sheet_name = sheet_name
        break
if not dashboard_sheet_name:
    dashboard_sheet_name = wb.sheetnames[1]

print(f"✅ 원본 시트: '{dashboard_sheet_name}'")
ws_original = wb[dashboard_sheet_name]

# [1] 데이터 수집 (셀 스타일 원본 그대로 복사)
cell_data = []
min_row, max_row = 1, 60
min_col, max_col = 1, 20

for row in range(min_row, max_row + 1):
    for col in range(min_col, max_col + 1):
        cell = ws_original.cell(row=row, column=col)
        
        # 테두리 정보 수집
        b_data = {}
        if cell.border:
            for s in ['left', 'right', 'top', 'bottom']:
                side = getattr(cell.border, s)
                if side:
                    # 색상 안전 추출
                    c_val = None
                    if side.color:
                        if hasattr(side.color, 'rgb') and isinstance(side.color.rgb, str) and len(side.color.rgb) <= 8:
                            c_val = side.color.rgb
                        elif hasattr(side.color, 'theme'):
                            pass # Theme color ignored for safety
                    b_data[s] = {'style': side.style, 'color': c_val}

        # 폰트 색상
        f_color = None
        if cell.font.color and hasattr(cell.font.color, 'rgb') and isinstance(cell.font.color.rgb, str):
            f_color = cell.font.color.rgb

        # 배경 색상
        bg_color = None
        if cell.fill and cell.fill.start_color and cell.fill.start_color.index != '00000000':
             bg_color = cell.fill.start_color.index

        if cell.value or bg_color or b_data:
            cell_data.append({
                'row': row, 'col': col, 'value': cell.value,
                'fmt': cell.number_format,
                'font': {'name': cell.font.name, 'sz': cell.font.size, 'b': cell.font.bold, 'color': f_color},
                'fill': {'c': bg_color, 't': cell.fill.fill_type},
                'align': {'h': cell.alignment.horizontal, 'v': cell.alignment.vertical, 'w': cell.alignment.wrap_text},
                'border': b_data
            })

merged_cells = list(ws_original.merged_cells.ranges)
row_heights = {r: ws_original.row_dimensions[r].height for r in range(min_row, max_row + 1) if ws_original.row_dimensions[r].height}
col_widths = {get_column_letter(c): ws_original.column_dimensions[get_column_letter(c)].width for c in range(min_col, max_col + 1) if ws_original.column_dimensions[get_column_letter(c)].width}

# [2] 새 시트 생성
if '📊 Dashboard_v5' in wb.sheetnames: del wb['📊 Dashboard_v5']
ws_new = wb.create_sheet('📊 Dashboard_v5')
ws_new.sheet_view.showGridLines = False

# [3] 복제 적용
for r, h in row_heights.items(): ws_new.row_dimensions[r].height = h
for c, w in col_widths.items(): ws_new.column_dimensions[c].width = w

for d in cell_data:
    c = ws_new.cell(d['row'], d['col'])
    c.value = d['value']
    if d['fmt']: c.number_format = d['fmt']
    c.font = Font(name=d['font']['name'], size=d['font']['sz'], bold=d['font']['b'], color=d['font']['color'])
    if d['fill']['c']: c.fill = PatternFill(start_color=d['fill']['c'], end_color=d['fill']['c'], fill_type=d['fill']['t'])
    c.alignment = Alignment(horizontal=d['align']['h'], vertical=d['align']['v'], wrap_text=d['align']['w'])
    
    borders = {}
    for k, v in d['border'].items():
        if v: borders[k] = Side(style=v['style'], color=v['color'])
    if borders: c.border = Border(**borders)

# 병합 (값 입력 후)
for m in merged_cells: ws_new.merge_cells(str(m))

# [4] 차트 생성

print("차트 생성 중...")

# -----------------------------------------------------------
# 차트 1: 콤보 차트 (주요지표)
# -----------------------------------------------------------
# 막대 (수입, 지출)
c1 = BarChart()
c1.type = "col"
c1.grouping = "clustered"
c1.overlap = 100
c1.y_axis.title = '금액'
c1.x_axis.title = '월'

# 데이터: D(수입), E(지출) 10:23 (9행 헤더)
# *titles_from_data=True* 사용
data_bar = Reference(ws_new, min_col=4, min_row=9, max_col=5, max_row=23)
c1.add_data(data_bar, titles_from_data=True)

cats = Reference(ws_new, min_col=3, min_row=10, max_row=23)
c1.set_categories(cats)

# 선 (합계)
c2 = LineChart()
# 데이터: F(합계) 10:23 (9행 헤더)
data_line = Reference(ws_new, min_col=6, min_row=9, max_col=6, max_row=23)
c2.add_data(data_line, titles_from_data=True)

# 합계 선 스타일 (노란색 등) - 기본값 사용 (안전)
# 필요시: c2.series[0].graphicalProperties.line.solidFill = "FFC000" (Hex Str)

c1 += c2 # 결합

c1.anchor = "H6"
c1.height = 13
c1.width = 18
c1.title = "주요 지표 추이"

ws_new.add_chart(c1)

# -----------------------------------------------------------
# 차트 2: 지출 구조 차트 (묶은 가로 막대)
# -----------------------------------------------------------
c3 = BarChart()
c3.type = "bar" # 가로 막대
c3.style = 10
c3.title = "지출 구조 차트"

# 데이터: D(금액), F(월평균) 37:47 (9행 X -> 37행이 헤더)
# 사용자 요청: D37:D47, F37:F47
# C열(37:47)이 카테고리(항목명)일 것임.
data_c3 = Reference(ws_new, min_col=4, min_row=37, max_col=4, max_row=47) # D열
data_c3_2 = Reference(ws_new, min_col=6, min_row=37, max_col=6, max_row=47) # F열

# add_data를 두 번 호출하거나 범위를 합칠 수 있음 (떨어져 있으니 두 번)
c3.add_data(data_c3, titles_from_data=True)
c3.add_data(data_c3_2, titles_from_data=True)

cats_c3 = Reference(ws_new, min_col=3, min_row=38, max_row=47) # 38행부터 데이터
c3.set_categories(cats_c3)

c3.anchor = "K30"
c3.height = 15
c3.width = 16

ws_new.add_chart(c3)

print("차트 생성 완료")

# 저장
output_file = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_Dashboard_v5.xlsx'
wb.save(output_file)
print(f"저장 완료: {output_file}")
