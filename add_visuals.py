import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference, Series
from openpyxl.utils import get_column_letter

INPUT_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-07~2025-12-07_v5_LinkDB.xlsx"
OUTPUT_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-07~2025-12-07_v6_Visuals.xlsx"

def add_visuals():
    print(f"Loading {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE)
    
    # 1. Formatting DB_Raw (Table Style)
    if "DB_Raw" in wb.sheetnames:
        ws = wb["DB_Raw"]
        print("Formatting 'DB_Raw'...")
        
        # Header Style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Apply to Row 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
            
        # Freeze Panes (Keep header visible)
        ws.freeze_panes = "A2"
        
    # 2. Add 'Dashboard_XLS' Sheet with Charts
    # Note: We need data to chart. We can create a pivot-like summary table first.
    
    dash_name = "Dashboard_Visual"
    if dash_name in wb.sheetnames:
        del wb[dash_name]
    ws_dash = wb.create_sheet(dash_name, 0) # Make it first tab
    
    print("Creating Dashboard...")
    ws_dash.sheet_view.showGridLines = False # Clean look
    
    # Title
    ws_dash["B2"] = "📊 Financial Dashboard"
    ws_dash["B2"].font = Font(size=20, bold=True, color="2F5597")
    
    # --- Summary Data Section (Hidden or Side) ---
    # We will build a small summary table in Columns K, L, M to feed the charts
    # Or just put it in the dashboard sheet
    
    # Let's create a Summary by Category for the Chart
    # Since openpyxl doesn't calculation formulas instantly, we can't easily "read" the values to pivot them.
    # But we can write SumIf formulas that the Chart refers to.
    
    # Category List (Static for now, or commonly used ones)
    categories = ["식비", "쇼핑", "교통", "주거", "통신", "의료", "저축", "기타"]
    
    # Summary Table Location on Dashboard
    ws_dash["B5"] = "Category Summary (Expense)"
    ws_dash["B5"].font = Font(bold=True)
    
    ws_dash["B6"] = "Category"
    ws_dash["C6"] = "Total"
    
    for i, cat in enumerate(categories):
        r = 7 + i
        ws_dash[f"B{r}"] = cat
        # Formula: SUMIFS(Amount, Type="지출", MainCat=Category)
        # DB_Raw Columns: F=Amount, C=Type, D=MainCat
        formula = f'=SUMIFS(DB_Raw!F:F, DB_Raw!C:C, "지출", DB_Raw!D:D, "{cat}")'
        ws_dash[f"C{r}"] = formula
        ws_dash[f"C{r}"].number_format = '#,##0'

    # --- Chart 1: Pie Chart (Expenses) ---
    pie = PieChart()
    labels = Reference(ws_dash, min_col=2, min_row=7, max_row=6+len(categories))
    data = Reference(ws_dash, min_col=3, min_row=6, max_row=6+len(categories))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Expense by Category"
    
    # Position
    ws_dash.add_chart(pie, "E5")
    
    # --- Chart 2: Income vs Expense Bar Chart ---
    # Data area
    ws_dash["B20"] = "Total Flow"
    ws_dash["B21"] = "Type"
    ws_dash["C21"] = "Amount"
    
    ws_dash["B22"] = "Income"
    ws_dash["C22"] = '=SUMIF(DB_Raw!C:C, "수입", DB_Raw!F:F)'
    
    ws_dash["B23"] = "Expense"
    ws_dash["C23"] = '=SUMIF(DB_Raw!C:C, "지출", DB_Raw!F:F)'
    
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Income vs Expense"
    bar.y_axis.title = 'Amount'
    
    labels_bar = Reference(ws_dash, min_col=2, min_row=22, max_row=23)
    data_bar = Reference(ws_dash, min_col=3, min_row=21, max_row=23)
    
    bar.add_data(data_bar, titles_from_data=True)
    bar.set_categories(labels_bar)
    bar.shape = 4
    
    ws_dash.add_chart(bar, "M5")
    
    print(f"Saving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print("Done! Visuals added.")

if __name__ == "__main__":
    add_visuals()
