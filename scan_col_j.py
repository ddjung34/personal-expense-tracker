
import openpyxl
from openpyxl import load_workbook

file_path = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_Dashboard_v22_labels_fix.xlsx'
wb = load_workbook(file_path, data_only=True)
ws = wb['📋 T_RawData']

print("=== Scanning Column J (Index 9) for Data ===")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True)):
    val = row[9] if len(row) > 9 else None
    if val is not None:
        print(f"Row {i+1}: {val} (type: {type(val)})")
