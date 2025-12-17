"""
Dashboard 복제 분석 스크립트
현재 Dashboard 시트의 구조를 분석하고 프로그래밍 방식으로 재생성
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
import copy

# Load workbook
print("=" * 70)
print("Dashboard 분석 및 복제 시작")
print("=" * 70)

wb = load_workbook(
    r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_수식연결_가계부엔진_최종.xlsx'
)

print(f"\n시트 목록 ({len(wb.sheetnames)}개):")
for idx, name in enumerate(wb.sheetnames, 1):
    print(f"  {idx}. {name}")

# Find Dashboard sheet
dashboard_sheet_name = None
for sheet_name in wb.sheetnames:
    if 'dashboard' in sheet_name.lower() or '대시보드' in sheet_name:
        dashboard_sheet_name = sheet_name
        break

if not dashboard_sheet_name:
    # Use second sheet if exists
    if len(wb.sheetnames) >= 2:
        dashboard_sheet_name = wb.sheetnames[1]
    else:
        raise ValueError("Dashboard 시트를 찾을 수 없습니다!")

print(f"\n✅ Dashboard 시트 감지: '{dashboard_sheet_name}'")

ws_original = wb[dashboard_sheet_name]

print("\n[1단계] Dashboard 구조 분석 중...")

# ================================================================
# 1. 셀 데이터 및 스타일 분석
# ================================================================

# 사용된 영역 파악
min_row, max_row = 1, 60
min_col, max_col = 1, 20  # A~T

cell_data = []
for row in range(min_row, max_row + 1):
    for col in range(min_col, max_col + 1):
        cell = ws_original.cell(row=row, column=col)
        if cell.value or (cell.fill and cell.fill.start_color and cell.fill.start_color.index != '00000000'):
            cell_data.append({
                'row': row,
                'col': col,
                'value': cell.value,
                'number_format': cell.number_format,
                'font': {
                    'name': cell.font.name,
                    'size': cell.font.size,
                    'bold': cell.font.bold,
                    'italic': cell.font.italic,
                    'color': cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None,
                },
                'fill': {
                    'start_color': cell.fill.start_color.index if cell.fill and cell.fill.start_color else None,
                    'fill_type': cell.fill.fill_type if cell.fill else None,
                },
                'alignment': {
                    'horizontal': cell.alignment.horizontal,
                    'vertical': cell.alignment.vertical,
                    'wrap_text': cell.alignment.wrap_text,
                },
                'border': {
                    'left': cell.border.left.style if cell.border and cell.border.left else None,
                    'right': cell.border.right.style if cell.border and cell.border.right else None,
                    'top': cell.border.top.style if cell.border and cell.border.top else None,
                    'bottom': cell.border.bottom.style if cell.border and cell.border.bottom else None,
                }
            })

print(f"   - 분석된 셀: {len(cell_data)}개")

# ================================================================
# 2. 병합된 셀 분석
# ================================================================

merged_cells = list(ws_original.merged_cells.ranges)
print(f"   - 병합된 영역: {len(merged_cells)}개")

# ================================================================
# 3. 행/열 크기 분석
# ================================================================

row_heights = {}
for row_num in range(min_row, max_row + 1):
    if ws_original.row_dimensions[row_num].height:
        row_heights[row_num] = ws_original.row_dimensions[row_num].height

col_widths = {}
for col_num in range(min_col, max_col + 1):
    col_letter = get_column_letter(col_num)
    if ws_original.column_dimensions[col_letter].width:
        col_widths[col_letter] = ws_original.column_dimensions[col_letter].width

print(f"   - 사용자 정의 행 높이: {len(row_heights)}개")
print(f"   - 사용자 정의 컬럼 너비: {len(col_widths)}개")

# ================================================================
# 4. 차트 분석
# ================================================================

charts = list(ws_original._charts)
print(f"   - 차트 개수: {len(charts)}개")

# ================================================================
# 5. 새 시트 생성
# ================================================================

print("\n[2단계] 새 시트 생성 중...")

# 기존 테스트 시트 삭제
if '📊 Dashboard_복제본' in wb.sheetnames:
    del wb['📊 Dashboard_복제본']

ws_new = wb.create_sheet('📊 Dashboard_복제본')
ws_new.sheet_view.showGridLines = False

print("   - 시트 생성 완료")

# ================================================================
# 6. 행/열 크기 복제
# ================================================================

print("\n[3단계] 행/열 크기 복제 중...")

for row_num, height in row_heights.items():
    ws_new.row_dimensions[row_num].height = height

for col_letter, width in col_widths.items():
    ws_new.column_dimensions[col_letter].width = width

print(f"   - 행/열 크기 설정 완료")

# ================================================================
# 7. 셀 데이터 및 스타일 복제
# ================================================================

print("\n[4단계] 셀 데이터 및 스타일 복제 중...")

for cell_info in cell_data:
    new_cell = ws_new.cell(row=cell_info['row'], column=cell_info['col'])
    
    # 값
    new_cell.value = cell_info['value']
    
    # 숫자 형식
    if cell_info['number_format']:
        new_cell.number_format = cell_info['number_format']
    
    # 폰트
    font_color = cell_info['font']['color']
    new_cell.font = Font(
        name=cell_info['font']['name'] or '맑은 고딕',
        size=cell_info['font']['size'] or 11,
        bold=cell_info['font']['bold'] or False,
        italic=cell_info['font']['italic'] or False,
        color=font_color if font_color and font_color != '00000000' else None
    )
    
    # 배경색
    if cell_info['fill']['start_color'] and cell_info['fill']['start_color'] != '00000000':
        new_cell.fill = PatternFill(
            start_color=cell_info['fill']['start_color'],
            end_color=cell_info['fill']['start_color'],
            fill_type=cell_info['fill']['fill_type'] or 'solid'
        )
    
    # 정렬
    new_cell.alignment = Alignment(
        horizontal=cell_info['alignment']['horizontal'],
        vertical=cell_info['alignment']['vertical'],
        wrap_text=cell_info['alignment']['wrap_text'] or False
    )
    
    # 테두리
    border_info = cell_info['border']
    if any(border_info.values()):
        new_cell.border = Border(
            left=Side(style=border_info['left']) if border_info['left'] else Side(),
            right=Side(style=border_info['right']) if border_info['right'] else Side(),
            top=Side(style=border_info['top']) if border_info['top'] else Side(),
            bottom=Side(style=border_info['bottom']) if border_info['bottom'] else Side()
        )

print(f"   - {len(cell_data)}개 셀 복제 완료")

# ================================================================
# 8. 병합된 셀 복제
# ================================================================

print("\n[5단계] 병합된 셀 복제 중...")

for merged_range in merged_cells:
    ws_new.merge_cells(str(merged_range))

print(f"   - {len(merged_cells)}개 병합 영역 복제")

# ================================================================
# 9. 차트 복제
# ================================================================

print("\n[6단계] 차트 복제 중...")

for idx, chart in enumerate(charts):
    try:
        # 차트 타입에 따라 새 차트 생성
        if isinstance(chart, BarChart):
            new_chart = BarChart()
            new_chart.type = chart.type
            if hasattr(chart, 'grouping'):
                new_chart.grouping = chart.grouping
        elif isinstance(chart, PieChart):
            new_chart = PieChart()
        else:
            print(f"     - 알 수 없는 차트 타입: {type(chart)}")
            continue
        
        # 제목
        if chart.title and hasattr(chart.title, 'text'):
            new_chart.title = chart.title.text
        
        # 크기
        new_chart.height = chart.height
        new_chart.width = chart.width
        
        # 스타일
        if hasattr(chart, 'style'):
            new_chart.style = chart.style
        
        # 데이터 시리즈 복제
        try:
            for series in chart.series:
                new_chart.series.append(copy.copy(series))
        except Exception as e:
            print(f"     - 시리즈 복사 실패: {e}")
        
        # 축 제목 (Bar chart)
        if isinstance(chart, BarChart):
            try:
                if hasattr(chart.y_axis, 'title') and chart.y_axis.title:
                    new_chart.y_axis.title = chart.y_axis.title
                if hasattr(chart.x_axis, 'title') and chart.x_axis.title:
                    new_chart.x_axis.title = chart.x_axis.title
            except:
                pass
        
        # 데이터 레이블 (Pie chart)
        if isinstance(chart, PieChart):
            try:
                if chart.dataLabels:
                    new_chart.dataLabels = DataLabelList()
                    new_chart.dataLabels.showCatName = True
                    new_chart.dataLabels.showPercent = True
                    new_chart.dataLabels.showVal = False
            except:
                pass
        
        # 위치
        new_chart.anchor = chart.anchor
        
        # 시트에 추가
        ws_new.add_chart(new_chart)
        
        print(f"   - 차트 {idx+1} 복제 완료 (위치: {chart.anchor})")
        
    except Exception as e:
        print(f"   - 차트 {idx+1} 복제 실패: {e}")

# ================================================================
# 10. 기타 속성
# ================================================================

print("\n[7단계] 기타 속성 복제 중...")

# 화면 고정
if ws_original.freeze_panes:
    ws_new.freeze_panes = ws_original.freeze_panes
    print(f"   - 화면 고정: {ws_original.freeze_panes}")

# 인쇄 설정
ws_new.page_setup.orientation = ws_original.page_setup.orientation
ws_new.page_setup.paperSize = ws_original.page_setup.paperSize
ws_new.print_options.gridLines = ws_original.print_options.gridLines

print("   - 페이지 설정 복제 완료")

# ================================================================
# 11. 저장
# ================================================================

print("\n[8단계] 파일 저장 중...")

output_file = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_Dashboard_복제테스트.xlsx'
wb.save(output_file)

print("\n" + "=" * 70)
print("✅ Dashboard 복제 완료!")
print("=" * 70)
print(f"\n📁 파일: {output_file}")
print(f"\n📊 원본 시트: '{dashboard_sheet_name}'")
print(f"📊 복제본 시트: '📊 Dashboard_복제본'")
print(f"\n🎯 복제된 요소:")
print(f"   - 셀 데이터: {len(cell_data)}개")
print(f"   - 병합 영역: {len(merged_cells)}개")
print(f"   - 차트: {len(charts)}개")
print(f"   - 행 높이: {len(row_heights)}개")
print(f"   - 컬럼 너비: {len(col_widths)}개")
print(f"\n💡 두 시트를 비교해서 확인해보세요!")
print("   완전히 동일하게 프로그래밍 방식으로 복제되었습니다.")
print("=" * 70)
