
import openpyxl
from openpyxl import load_workbook

file_path = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_Dashboard_v22_labels_fix.xlsx'
wb = load_workbook(file_path, data_only=True)
ws = wb['📋 T_RawData']

unique_c = set()
unique_d = set()

# Row 4부터 100까지 스캔
for row in ws.iter_rows(min_row=4, max_row=100, values_only=True):
    # C열: Index 2, D열: Index 3
    if row[2]: unique_c.add(str(row[2]).strip())
    if row[3]: unique_d.add(str(row[3]).strip())

print("=== Unique C (Flow) ===")
print(unique_c)
print("\n=== Unique D (Category) ===")
print(unique_d)
