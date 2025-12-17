import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

INPUT_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-07~2025-12-07.xlsx"
OUTPUT_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\Dashboard_Analysis.xlsx"

def create_excel_dashboard():
    print(f"📊 Loading data from: {INPUT_FILE}")
    
    # Read the data
    df = pd.read_excel(INPUT_FILE, sheet_name='가계부 내역', engine='openpyxl')
    
    
    # Data cleaning
    df['date'] = pd.to_datetime(df['날짜'], errors='coerce')
    df['amount'] = pd.to_numeric(df['금액'], errors='coerce')
    df['type'] = df['타입']
    df['main_category'] = df['대분류']
    df['sub_category'] = df['소분류']
    df['payment_method'] = df['결제수단']
    df['merchant'] = df['내용']
    
    # Remove invalid data
    df = df.dropna(subset=['date', 'amount', 'type'])
    
    print(f"✅ Loaded {len(df)} transactions")
    
    # Create new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # ===== SHEET 1: Dashboard =====
    ws_dash = wb.create_sheet("Dashboard", 0)
    ws_dash.sheet_view.showGridLines = False
    
    # Title
    ws_dash['B2'] = "💎 재정 분석 대시보드"
    ws_dash['B2'].font = Font(size=24, bold=True, color="1F4E78")
    ws_dash.merge_cells('B2:G2')
    
    # Date range info
    ws_dash['B3'] = f"분석 기간: {df['date'].min().strftime('%Y-%m-%d')} ~ {df['date'].max().strftime('%Y-%m-%d')}"
    ws_dash['B3'].font = Font(size=11, italic=True, color="7F7F7F")
    ws_dash.merge_cells('B3:G3')
    
    # ===== KPI Cards Area =====
    row_start = 5
    
    # Calculate KPIs
    total_income = df[df['type'] == '수입']['amount'].sum()
    total_expense = df[df['type'] == '지출']['amount'].sum()
    net_income = total_income - total_expense
    transaction_count = len(df)
    
    # KPI styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    value_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    value_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    
    kpis = [
        ("총 수입", total_income, "B"),
        ("총 지출", total_expense, "D"),
        ("순수익", net_income, "F"),
        ("거래 건수", transaction_count, "H")
    ]
    
    for label, value, col in kpis:
        # Header
        ws_dash[f'{col}{row_start}'] = label
        ws_dash[f'{col}{row_start}'].fill = header_fill
        ws_dash[f'{col}{row_start}'].font = header_font
        ws_dash[f'{col}{row_start}'].alignment = center
        
        # Value
        ws_dash[f'{col}{row_start+1}'] = f"{value:,.0f}" if isinstance(value, (int, float)) else value
        ws_dash[f'{col}{row_start+1}'].fill = value_fill
        ws_dash[f'{col}{row_start+1}'].font = value_font
        ws_dash[f'{col}{row_start+1}'].alignment = center
    
    # ===== SHEET 2: Monthly Analysis =====
    ws_monthly = wb.create_sheet("월별분석", 1)
    
    # Monthly grouping
    monthly = df.groupby([df['date'].dt.to_period('M'), 'type'])['amount'].sum().unstack(fill_value=0)
    
    if '수입' not in monthly.columns:
        monthly['수입'] = 0
    if '지출' not in monthly.columns:
        monthly['지출'] = 0
        
    monthly['순수익'] = monthly.get('수입', 0) - monthly.get('지출', 0)
    monthly = monthly.reset_index()
    monthly['월'] = monthly['date'].astype(str)
    
    # Write headers
    headers = ['월', '수입', '지출', '순수익']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_monthly.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = center
    
    # Write data
    for row_idx, row_data in enumerate(monthly.itertuples(index=False), 2):
        ws_monthly.cell(row=row_idx, column=1, value=row_data.월)
        ws_monthly.cell(row=row_idx, column=2, value=row_data.수입)
        ws_monthly.cell(row=row_idx, column=3, value=row_data.지출)
        ws_monthly.cell(row=row_idx, column=4, value=row_data.순수익)
        
        # Format numbers
        for col in range(2, 5):
            ws_monthly.cell(row=row_idx, column=col).number_format = '#,##0'
    
    # Create Line Chart for Monthly Trends
    chart_monthly = LineChart()
    chart_monthly.title = "월별 재정 추이"
    chart_monthly.style = 10
    chart_monthly.height = 10
    chart_monthly.width = 20
    
    # Data for chart
    data = Reference(ws_monthly, min_col=2, min_row=1, max_row=len(monthly)+1, max_col=4)
    cats = Reference(ws_monthly, min_col=1, min_row=2, max_row=len(monthly)+1)
    
    chart_monthly.add_data(data, titles_from_data=True)
    chart_monthly.set_categories(cats)
    
    ws_monthly.add_chart(chart_monthly, "F2")
    
    # ===== SHEET 3: Category Analysis =====
    ws_category = wb.create_sheet("카테고리분석", 2)
    
    # Category breakdown (expenses only)
    expense_df = df[df['type'] == '지출']
    category_summary = expense_df.groupby('main_category')['amount'].sum().reset_index()
    category_summary = category_summary.sort_values('amount', ascending=False)
    
    # Write headers
    ws_category['B1'] = "카테고리"
    ws_category['C1'] = "지출액"
    ws_category['B1'].font = header_font
    ws_category['C1'].font = header_font
    ws_category['B1'].fill = header_fill
    ws_category['C1'].fill = header_fill
    ws_category['B1'].alignment = center
    ws_category['C1'].alignment = center
    
    # Write data
    for row_idx, row_data in enumerate(category_summary.itertuples(index=False), 2):
        ws_category.cell(row=row_idx, column=2, value=row_data.main_category)
        ws_category.cell(row=row_idx, column=3, value=row_data.amount)
        ws_category.cell(row=row_idx, column=3).number_format = '#,##0'
    
    # Create Pie Chart
    pie = PieChart()
    pie.title = "카테고리별 지출 비중"
    pie.height = 12
    pie.width = 12
    
    labels = Reference(ws_category, min_col=2, min_row=2, max_row=len(category_summary)+1)
    data = Reference(ws_category, min_col=3, min_row=1, max_row=len(category_summary)+1)
    
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    
    ws_category.add_chart(pie, "E2")
    
    # ===== SHEET 4: Raw Data (Cleaned) =====
    ws_data = wb.create_sheet("거래내역", 3)
    
    # Select relevant columns
    clean_df = df[['date', 'type', 'main_category', 'sub_category', 'amount', 'payment_method']].copy()
    clean_df['date'] = clean_df['date'].dt.strftime('%Y-%m-%d')
    
    # Write headers
    headers_data = ['날짜', '구분', '대분류', '소분류', '금액', '결제수단']
    for col_idx, header in enumerate(headers_data, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    
    # Write data
    for row_idx, row_data in enumerate(clean_df.itertuples(index=False), 2):
        ws_data.cell(row=row_idx, column=1, value=row_data.date)
        ws_data.cell(row=row_idx, column=2, value=row_data.type)
        ws_data.cell(row=row_idx, column=3, value=row_data.main_category)
        ws_data.cell(row=row_idx, column=4, value=row_data.sub_category)
        ws_data.cell(row=row_idx, column=5, value=row_data.amount)
        ws_data.cell(row=row_idx, column=6, value=row_data.payment_method)
        
        # Format
        ws_data.cell(row=row_idx, column=5).number_format = '#,##0'
    
    # Freeze panes
    ws_data.freeze_panes = "A2"
    
    # Auto-fit columns
    for ws in [ws_dash, ws_monthly, ws_category, ws_data]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    # Save
    print(f"💾 Saving to: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    print("✅ Excel Dashboard created successfully!")
    print(f"\n📊 Summary:")
    print(f"   - Total Sheets: 4 (Dashboard, 월별분석, 카테고리분석, 거래내역)")
    print(f"   - Charts: 2 (Line Chart for Monthly, Pie Chart for Categories)")
    print(f"   - Transactions: {len(df)}")

if __name__ == "__main__":
    create_excel_dashboard()
