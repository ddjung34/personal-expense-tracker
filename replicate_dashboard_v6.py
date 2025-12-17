"""
Dashboard 정밀 복제 v6 - 차트 배치 및 디테일 수정
1. 차트 위치 및 크기 정밀 조정 (화면 꽉 차게)
2. 차트 세부 내용 (범례, 데이터 레이블) 추가
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

print("=" * 70)
print("Dashboard 정밀 복제 v6")
print("=" * 70)

wb = load_workbook(
    r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_수식연결_가계부엔진_최종.xlsx'
)

# Dashboard 시트 찾기
dashboard_sheet_name = None
for s in wb.sheetnames:
    if 'dashboard' in s.lower() and '복제' not in s and '최종' not in s and 'v' not in s:
        dashboard_sheet_name = s; break
if not dashboard_sheet_name: dashboard_sheet_name = wb.sheetnames[1]

print(f"✅ 원본 시트: '{dashboard_sheet_name}'")
ws_original = wb[dashboard_sheet_name]

# [1] 데이터 및 스타일 복제 (v5와 동일)
cell_data = []
min_row, max_row = 1, 60
min_col, max_col = 1, 25

for row in range(min_row, max_row + 1):
    for col in range(min_col, max_col + 1):
        cell = ws_original.cell(row=row, column=col)
        
        # 테두리
        b_data = {}
        if cell.border:
            for s in ['left', 'right', 'top', 'bottom']:
                side = getattr(cell.border, s)
                if side:
                    c_val = None
                    if side.color and hasattr(side.color, 'rgb') and isinstance(side.color.rgb, str) and len(side.color.rgb) <= 8:
                        c_val = side.color.rgb
                    b_data[s] = {'style': side.style, 'color': c_val}

        # 폰트 색상
        f_color = None
        if cell.font.color and hasattr(cell.font.color, 'rgb') and isinstance(cell.font.color.rgb, str):
            f_color = cell.font.color.rgb

        # 배경 색상
        bg_color = None
        if cell.fill and cell.fill.start_color and cell.fill.start_color.index != '00000000':
             bg_color = cell.fill.start_color.index

        if cell.value or bg_color or b_data:
            cell_data.append({
                'row': row, 'col': col, 'value': cell.value,
                'fmt': cell.number_format,
                'font': {'name': cell.font.name, 'sz': cell.font.size, 'b': cell.font.bold, 'color': f_color},
                'fill': {'c': bg_color, 't': cell.fill.fill_type},
                'align': {'h': cell.alignment.horizontal, 'v': cell.alignment.vertical, 'w': cell.alignment.wrap_text},
                'border': b_data
            })

merged_cells = list(ws_original.merged_cells.ranges)
row_heights = {r: ws_original.row_dimensions[r].height for r in range(min_row, max_row + 1) if ws_original.row_dimensions[r].height}
col_widths = {get_column_letter(c): ws_original.column_dimensions[get_column_letter(c)].width for c in range(min_col, max_col + 1) if ws_original.column_dimensions[get_column_letter(c)].width}

# [2] 새 시트 생성
if '📊 Dashboard_v6' in wb.sheetnames: del wb['📊 Dashboard_v6']
ws_new = wb.create_sheet('📊 Dashboard_v6')
ws_new.sheet_view.showGridLines = False

# [3] 복제
for r, h in row_heights.items(): ws_new.row_dimensions[r].height = h
for c, w in col_widths.items(): ws_new.column_dimensions[c].width = w

for d in cell_data:
    c = ws_new.cell(d['row'], d['col'])
    c.value = d['value']
    if d['fmt']: c.number_format = d['fmt']
    c.font = Font(name=d['font']['name'], size=d['font']['sz'], bold=d['font']['b'], color=d['font']['color'])
    if d['fill']['c']: c.fill = PatternFill(start_color=d['fill']['c'], end_color=d['fill']['c'], fill_type=d['fill']['t'])
    c.alignment = Alignment(horizontal=d['align']['h'], vertical=d['align']['v'], wrap_text=d['align']['w'])
    
    borders = {}
    for k, v in d['border'].items():
        if v: borders[k] = Side(style=v['style'], color=v['color'])
    if borders: c.border = Border(**borders)

for m in merged_cells: ws_new.merge_cells(str(m))

# [4] 차트 생성 (개선됨)

# -----------------------------------------------------------
# 차트 1: 콤보 차트 (주요지표) - 배치 개선
# -----------------------------------------------------------
c1 = BarChart()
c1.type = "col"
c1.grouping = "clustered"
c1.overlap = 100
c1.y_axis.title = '금액'
c1.title = "주요 지표 추이"

# 데이터
data_bar = Reference(ws_new, min_col=4, min_row=9, max_col=5, max_row=23)
c1.add_data(data_bar, titles_from_data=True)
cats = Reference(ws_new, min_col=3, min_row=10, max_row=23)
c1.set_categories(cats)

c2 = LineChart()
data_line = Reference(ws_new, min_col=6, min_row=9, max_col=6, max_row=23)
c2.add_data(data_line, titles_from_data=True)

c1 += c2

# **배치 조절**
# 왼쪽 테이블이 C~G열. 차트는 H열부터 시작.
# 높이는 테이블 높이(10행~23행)보다 조금 더 넉넉하게.
# 너비는 T열까지 꽉 차게.
c1.anchor = "H6"
c1.height = 14  # 높이 증가
c1.width = 25   # 너비 대폭 증가 (T열까지 닿도록)

# **범례 표시**
c1.legend.position = "tr" # top-right

ws_new.add_chart(c1)

# -----------------------------------------------------------
# 차트 2: 지출 구조 차트 (묶은 가로 막대) - 상세 내용 추가
# -----------------------------------------------------------
c3 = BarChart()
c3.type = "bar" # 가로 막대
c3.style = 10
c3.title = "지출 구조 차트"

# 데이터: D(금액), F(월평균) 37:47 
data_c3 = Reference(ws_new, min_col=4, min_row=37, max_col=4, max_row=47)
data_c3_2 = Reference(ws_new, min_col=6, min_row=37, max_col=6, max_row=47)

c3.add_data(data_c3, titles_from_data=True)
c3.add_data(data_c3_2, titles_from_data=True) # 월평균도 추가

cats_c3 = Reference(ws_new, min_col=3, min_row=38, max_row=47)
c3.set_categories(cats_c3)

# **배치 조절**
# 왼쪽 테이블이 C~F열. 차트는 G열 또는 H열부터 시작 가능.
# 원본 이미지(2)를 보면 I열 정도에서 시작해서 T열까지 꽉 참.
# 높이는 테이블 높이(38~47행)보다 훨씬 큼 (30행~50행 덮음).
c3.anchor = "I30" # 위치 조정
c3.height = 16 # 높이 증가
c3.width = 22  # 너비 대폭 증가

# **데이터 레이블 추가**
c3.dataLabels = DataLabelList()
c3.dataLabels.showCatName = True # 카테고리 이름
c3.dataLabels.showVal = True     # 값
c3.dataLabels.showPercent = True # 백분율 (BarChart에서 동작할지 확인 필요, 보통 Pie)
# BarChart에서 Percent는 Stacked 100%가 아니면 계산 안 될 수 있음.
# 일단 활성화.

ws_new.add_chart(c3)

# 저장
output_file = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_Dashboard_v6.xlsx'
wb.save(output_file)
print(f"저장 완료: {output_file}")
