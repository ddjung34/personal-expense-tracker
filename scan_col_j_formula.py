
import openpyxl
from openpyxl import load_workbook

file_path = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_Dashboard_v22_labels_fix.xlsx'
# data_only=False (default) to see formulas
wb = load_workbook(file_path) 
ws = wb['📋 T_RawData']

print("=== Scanning Column J (formulas) ===")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False)):
    cell = row[9] # J column
    print(f"Row {i+1}: {cell.value}")
