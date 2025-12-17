import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

INPUT_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-07~2025-12-07.xlsx"
OUTPUT_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\{date}_가계부_종합분석.xlsx".format(
    date=datetime.now().strftime("%Y%m%d")
)

def create_excel_report():
    print(f"📊 Loading data from: {INPUT_FILE}")
    
    # Read the data
    df = pd.read_excel(INPUT_FILE, sheet_name='가계부 내역', engine='openpyxl')
    
    # Data cleaning
    df['date'] = pd.to_datetime(df['날짜'], errors='coerce')
    df['amount'] = pd.to_numeric(df['금액'], errors='coerce')
    df['type'] = df['타입']
    df['main_category'] = df['대분류']
    df['sub_category'] = df['소분류']
    df['payment_method'] = df['결제수단']
    df['merchant'] = df['내용']
    df['memo'] = df['메모'].fillna('')
    
    # Remove invalid data
    df = df.dropna(subset=['date', 'amount', 'type'])
    
    print(f"✅ Loaded {len(df)} transactions")
    
    # Calculate KPIs
    total_income = df[df['type'] == '수입']['amount'].sum()
    total_expense = df[df['type'] == '지출']['amount'].sum()
    net_income = total_income - total_expense
    
    if not df[df['type'] == '지출'].empty:
        top_category = df[df['type'] == '지출'].groupby('main_category')['amount'].sum().idxmax()
        top_category_amount = df[df['type'] == '지출'].groupby('main_category')['amount'].sum().max()
    else:
        top_category = "N/A"
        top_category_amount = 0
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # ==================================================
    # SHEET 1: 📊 대시보드 요약 (Dashboard Summary)
    # ==================================================
    ws_dash = wb.create_sheet("📊 대시보드 요약", 0)
    ws_dash.sheet_view.showGridLines = False
    
    # Title
    ws_dash['B2'] = "가계부 종합 분석 대시보드"
    ws_dash['B2'].font = Font(size=20, bold=True, color="1F4E78")
    ws_dash.merge_cells('B2:H2')
    ws_dash['B2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Date range
    ws_dash['B3'] = f"분석 기간: {df['date'].min().strftime('%Y-%m-%d')} ~ {df['date'].max().strftime('%Y-%m-%d')}"
    ws_dash['B3'].font = Font(size=11, italic=True, color="7F7F7F")
    ws_dash.merge_cells('B3:H3')
    ws_dash['B3'].alignment = Alignment(horizontal='center')
    
    # KPI Section
    row = 5
    kpi_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    kpi_font = Font(bold=True, color="FFFFFF", size=12)
    value_font = Font(bold=True, size=16, color="2C3E50")
    
    kpis = [
        ("💰 총 순수익", net_income, "B"),
        ("💸 총 지출액", total_expense, "D"),
        ("🔥 최대 지출 카테고리", f"{top_category}\n({top_category_amount:,.0f}원)", "F")
    ]
    
    for label, value, col in kpis:
        # Header
        ws_dash[f'{col}{row}'] = label
        ws_dash[f'{col}{row}'].fill = kpi_fill
        ws_dash[f'{col}{row}'].font = kpi_font
        ws_dash[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dash[f'{col}{row}'].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Value
        ws_dash[f'{col}{row+1}'] = f"{value:,.0f}원" if isinstance(value, (int, float)) else value
        ws_dash[f'{col}{row+1}'].font = value_font
        ws_dash[f'{col}{row+1}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws_dash[f'{col}{row+1}'].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        ws_dash.row_dimensions[row+1].height = 40
    
    # Monthly data for chart
    monthly = df.groupby([df['date'].dt.to_period('M'), 'type'])['amount'].sum().unstack(fill_value=0)
    if '수입' not in monthly.columns:
        monthly['수입'] = 0
    if '지출' not in monthly.columns:
        monthly['지출'] = 0
    monthly = monthly.reset_index()
    monthly['월'] = monthly['date'].astype(str)
    
    # Write monthly data (hidden area for chart)
    data_row = 10
    ws_dash['K10'] = "월"
    ws_dash['L10'] = "수입"
    ws_dash['M10'] = "지출"
    
    for i, row_data in enumerate(monthly.itertuples(index=False), data_row + 1):
        ws_dash[f'K{i}'] = row_data.월
        ws_dash[f'L{i}'] = row_data.수입
        ws_dash[f'M{i}'] = row_data.지출
    
    # Line Chart: Monthly Trends
    chart1 = LineChart()
    chart1.title = "월별 수입/지출 추이"
    chart1.style = 10
    chart1.y_axis.title = "금액 (원)"
    chart1.x_axis.title = "월"
    chart1.height = 10
    chart1.width = 18
    
    data = Reference(ws_dash, min_col=12, min_row=data_row, max_row=data_row + len(monthly), max_col=13)
    cats = Reference(ws_dash, min_col=11, min_row=data_row + 1, max_row=data_row + len(monthly))
    
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    
    ws_dash.add_chart(chart1, "B9")
    
    # Category data for pie chart
    expense_df = df[df['type'] == '지출']
    if not expense_df.empty:
        category_data = expense_df.groupby('main_category')['amount'].sum().reset_index()
        category_data = category_data.sort_values('amount', ascending=False)
        
        # Write category data (hidden area for chart)
        cat_row = 10
        ws_dash['O10'] = "카테고리"
        ws_dash['P10'] = "금액"
        
        for i, row_data in enumerate(category_data.itertuples(index=False), cat_row + 1):
            ws_dash[f'O{i}'] = row_data.main_category
            ws_dash[f'P{i}'] = row_data.amount
        
        # Pie Chart: Category Distribution
        pie = PieChart()
        pie.title = "카테고리별 지출 비중"
        pie.height = 10
        pie.width = 12
        
        labels = Reference(ws_dash, min_col=15, min_row=cat_row + 1, max_row=cat_row + len(category_data))
        data = Reference(ws_dash, min_col=16, min_row=cat_row, max_row=cat_row + len(category_data))
        
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        
        ws_dash.add_chart(pie, "B24")
    
    # ==================================================
    # SHEET 2: 📈 월별 및 카테고리 피벗 분석
    # ==================================================
    ws_pivot = wb.create_sheet("📈 피벗 분석", 1)
    
    # Title
    ws_pivot['B2'] = "월별 및 카테고리 피벗 분석"
    ws_pivot['B2'].font = Font(size=16, bold=True, color="2C3E50")
    ws_pivot.merge_cells('B2:F2')
    
    # Monthly Pivot
    ws_pivot['B4'] = "월별 재정 요약"
    ws_pivot['B4'].font = Font(size=14, bold=True)
    
    monthly_pivot = df.groupby([df['date'].dt.to_period('M'), 'type'])['amount'].sum().unstack(fill_value=0)
    if '수입' not in monthly_pivot.columns:
        monthly_pivot['수입'] = 0
    if '지출' not in monthly_pivot.columns:
        monthly_pivot['지출'] = 0
    monthly_pivot['순수익'] = monthly_pivot['수입'] - monthly_pivot['지출']
    monthly_pivot = monthly_pivot.reset_index()
    monthly_pivot['월'] = monthly_pivot['date'].astype(str)
    monthly_pivot = monthly_pivot[['월', '수입', '지출', '순수익']]
    
    # Write monthly pivot
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for r_idx, row in enumerate(dataframe_to_rows(monthly_pivot, index=False, header=True), 5):
        for c_idx, value in enumerate(row, 2):
            cell = ws_pivot.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 5:  # Header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center')
                if c_idx > 2:  # Number columns
                    cell.number_format = '#,##0'
    
    # Category Pivot
    pivot_start_row = 5 + len(monthly_pivot) + 3
    ws_pivot[f'B{pivot_start_row}'] = "카테고리별 지출 분석"
    ws_pivot[f'B{pivot_start_row}'].font = Font(size=14, bold=True)
    
    if not expense_df.empty:
        cat_pivot = expense_df.groupby(['main_category', 'sub_category'])['amount'].sum().reset_index()
        cat_pivot = cat_pivot.sort_values('amount', ascending=False)
        cat_pivot.columns = ['대분류', '소분류', '총 지출액']
        
        # Write category pivot
        for r_idx, row in enumerate(dataframe_to_rows(cat_pivot, index=False, header=True), pivot_start_row + 1):
            for c_idx, value in enumerate(row, 2):
                cell = ws_pivot.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == pivot_start_row + 1:  # Header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    if c_idx == 4:  # Amount column
                        cell.number_format = '#,##0'
    
    # ==================================================
    # SHEET 3: 📋 정리된 Raw Data
    # ==================================================
    ws_data = wb.create_sheet("📋 Raw Data", 2)
    
    # Prepare clean data
    clean_df = df[['date', 'type', 'main_category', 'sub_category', 'amount', 'payment_method', 'merchant', 'memo']].copy()
    clean_df.columns = ['날짜', '구분', '대분류', '소분류', '금액', '결제수단', '거래처', '메모']
    
    # Write to sheet
    for r_idx, row in enumerate(dataframe_to_rows(clean_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                if c_idx == 1:  # Date column
                    if isinstance(value, pd.Timestamp):
                        cell.value = value.to_pydatetime()
                        cell.number_format = 'YYYY-MM-DD'
                elif c_idx == 5:  # Amount column
                    cell.number_format = '#,##0'
    
    # Freeze panes
    ws_data.freeze_panes = "A2"
    
    # Auto-fit columns for all sheets
    for ws in [ws_dash, ws_pivot, ws_data]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    # Save
    print(f"💾 Saving to: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    print("✅ Excel Report created successfully!")
    print(f"\n📊 Summary:")
    print(f"   - 시트 1: 📊 대시보드 요약 (KPIs + 2 Charts)")
    print(f"   - 시트 2: 📈 피벗 분석 (Monthly + Category Pivots)")
    print(f"   - 시트 3: 📋 Raw Data ({len(clean_df)} transactions)")
    print(f"\n✅ File saved: {OUTPUT_FILE}")

if __name__ == "__main__":
    create_excel_report()
