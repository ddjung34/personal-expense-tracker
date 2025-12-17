import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load the existing workbook
wb = openpyxl.load_workbook(
    r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251213_수식연결_가계부엔진.xlsx',
    data_only=False
)

# Get Dashboard sheet
ws_dash = wb['📊 Dashboard Summary']

# Colors
HEADER_COLOR = "2C3E50"
ACCENT_COLOR = "3498DB"
SUCCESS_COLOR = "27AE60"
DANGER_COLOR = "E74C3C"
WARNING_COLOR = "F39C12"
LIGHT_BG = "ECF0F1"

# Add Insights section starting at row 45 (below charts)
insight_row = 45

# Title
ws_dash[f'B{insight_row}'] = "💡 지출 분석 인사이트"
ws_dash[f'B{insight_row}'].font = Font(size=14, bold=True, color=ACCENT_COLOR, name="맑은 고딕")
ws_dash.merge_cells(f'B{insight_row}:G{insight_row}')
ws_dash[f'B{insight_row}'].alignment = Alignment(horizontal='center', vertical='center')
ws_dash[f'B{insight_row}'].fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
ws_dash.row_dimensions[insight_row].height = 25

# Key findings
findings = [
    ("🔥 최대 지출", "월세 (19.7%)", "6,540,000원"),
    ("🍽️ 식생활", "식비+카페 (23.6%)", "7,881,234원"),
    ("🚗 고정비", "월세+자동차+주거 (35.9%)", "11,953,834원"),
    ("📊 거래 빈도 1위", "생활 (761건)", "평균 3,489원"),
    ("⚠️ 주의 필요", "카페/간식 (264건)", "월 평균 20만원"),
]

for idx, (label, value, detail) in enumerate(findings, insight_row + 2):
    # Label
    ws_dash[f'B{idx}'] = label
    ws_dash[f'B{idx}'].font = Font(size=10, bold=True, name="맑은 고딕")
    ws_dash[f'B{idx}'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Value
    ws_dash[f'D{idx}'] = value
    ws_dash[f'D{idx}'].font = Font(size=10, color=DANGER_COLOR, name="맑은 고딕")
    ws_dash[f'D{idx}'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Detail
    ws_dash[f'F{idx}'] = detail
    ws_dash[f'F{idx}'].font = Font(size=9, italic=True, color="7F7F7F", name="맑은 고딕")
    ws_dash[f'F{idx}'].alignment = Alignment(horizontal='right', vertical='center')
    
    # Borders
    for col in ['B', 'D', 'F']:
        ws_dash[f'{col}{idx}'].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

# Recommendations section
rec_row = insight_row + 9
ws_dash[f'B{rec_row}'] = "🎯 개선 제안"
ws_dash[f'B{rec_row}'].font = Font(size=14, bold=True, color=WARNING_COLOR, name="맑은 고딕")
ws_dash.merge_cells(f'B{rec_row}:G{rec_row}')
ws_dash[f'B{rec_row}'].alignment = Alignment(horizontal='center', vertical='center')
ws_dash[f'B{rec_row}'].fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
ws_dash.row_dimensions[rec_row].height = 25

# Recommendations
recs = [
    ("1️⃣", "카페/간식 빈도 50% 감축", "→ 월 10만원 절약"),
    ("2️⃣", "온라인쇼핑 충동구매 자제", "→ 월 5만원 절약"),
    ("3️⃣", "외식 vs 자취 비율 조정", "→ 월 15만원 절약"),
    ("4️⃣", "고정비 재검토 (통신비, 구독)", "→ 월 3만원 절약"),
]

for idx, (num, action, saving) in enumerate(recs, rec_row + 2):
    ws_dash[f'B{idx}'] = num
    ws_dash[f'B{idx}'].font = Font(size=10, name="맑은 고딕")
    ws_dash[f'B{idx}'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws_dash[f'C{idx}'] = action
    ws_dash[f'C{idx}'].font = Font(size=10, name="맑은 고딕")
    ws_dash[f'C{idx}'].alignment = Alignment(horizontal='left', vertical='center')
    ws_dash.merge_cells(f'C{idx}:E{idx}')
    
    ws_dash[f'F{idx}'] = saving
    ws_dash[f'F{idx}'].font = Font(size=10, bold=True, color=SUCCESS_COLOR, name="맑은 고딕")
    ws_dash[f'F{idx}'].alignment = Alignment(horizontal='right', vertical='center')
    
    for col in ['B', 'C', 'F']:
        ws_dash[f'{col}{idx}'].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

# Monthly comparison section (right panel)
comp_row = insight_row
ws_dash[f'I{comp_row}'] = "📅 월별 비교 (높은 순)"
ws_dash[f'I{comp_row}'].font = Font(size=14, bold=True, color=HEADER_COLOR, name="맑은 고딕")
ws_dash.merge_cells(f'I{comp_row}:L{comp_row}')
ws_dash[f'I{comp_row}'].alignment = Alignment(horizontal='center', vertical='center')
ws_dash[f'I{comp_row}'].fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")

# Monthly data
monthly = [
    ("2025-10", "4,602,779원", "⚠️ 최대"),
    ("2025-09", "3,525,038원", ""),
    ("2025-03", "2,855,082원", ""),
    ("평균", "~2,700,000원", ""),
    ("목표", "< 2,500,000원", "✅"),
]

for idx, (month, amount, note) in enumerate(monthly, comp_row + 2):
    ws_dash[f'I{idx}'] = month
    ws_dash[f'I{idx}'].font = Font(size=9, name="맑은 고딕")
    ws_dash[f'I{idx}'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws_dash[f'J{idx}'] = amount
    ws_dash[f'J{idx}'].font = Font(size=9, name="맑은 고딕")
    ws_dash[f'J{idx}'].alignment = Alignment(horizontal='right', vertical='center')
    
    ws_dash[f'K{idx}'] = note
    ws_dash[f'K{idx}'].font = Font(size=9, bold=True, color=DANGER_COLOR if "⚠️" in note else SUCCESS_COLOR, name="맑은 고딕")
    ws_dash[f'K{idx}'].alignment = Alignment(horizontal='center', vertical='center')

# Save
output_file = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_수식연결_가계부엔진_분석추가.xlsx'
wb.save(output_file)

print("✅ Dashboard에 인사이트 추가 완료!")
print(f"📁 저장 위치: {output_file}")
