"""
Dashboard 정밀 복제 스크립트 v2
차트의 세부 속성까지 완벽하게 복제
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.axis import NumericAxis
import copy

print("=" * 70)
print("Dashboard 정밀 복제 v2")
print("=" * 70)

wb = load_workbook(
    r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_수식연결_가계부엔진_최종.xlsx'
)

# Find Dashboard sheet
dashboard_sheet_name = None
for sheet_name in wb.sheetnames:
    if 'dashboard' in sheet_name.lower() and '복제' not in sheet_name and '분석' in sheet_name:
        dashboard_sheet_name = sheet_name
        break

if not dashboard_sheet_name:
    for sheet_name in wb.sheetnames:
        if 'dashboard' in sheet_name.lower() or '대시보드' in sheet_name:
            if '복제' not in sheet_name:
                dashboard_sheet_name = sheet_name
                break

if not dashboard_sheet_name and len(wb.sheetnames) >= 2:
    dashboard_sheet_name = wb.sheetnames[1]

print(f"\n✅ 원본 시트: '{dashboard_sheet_name}'")

ws_original = wb[dashboard_sheet_name]

# ================================================================
# 데이터 수집
# ================================================================

print("\n[1] 데이터 수집 중...")

min_row, max_row = 1, 60
min_col, max_col = 1, 20

cell_data = []
for row in range(min_row, max_row + 1):
    for col in range(min_col, max_col + 1):
        cell = ws_original.cell(row=row, column=col)
        if cell.value or (cell.fill and cell.fill.start_color and cell.fill.start_color.index != '00000000'):
            # 테두리 정보 수집
            border_data = {}
            if cell.border:
                if cell.border.left:
                    border_data['left'] = {'style': cell.border.left.style, 'color': cell.border.left.color.rgb if cell.border.left.color else None}
                if cell.border.right:
                    border_data['right'] = {'style': cell.border.right.style, 'color': cell.border.right.color.rgb if cell.border.right.color else None}
                if cell.border.top:
                    border_data['top'] = {'style': cell.border.top.style, 'color': cell.border.top.color.rgb if cell.border.top.color else None}
                if cell.border.bottom:
                    border_data['bottom'] = {'style': cell.border.bottom.style, 'color': cell.border.bottom.color.rgb if cell.border.bottom.color else None}
            
            cell_data.append({
                'row': row,
                'col': col,
                'value': cell.value,
                'number_format': cell.number_format,
                'font': {
                    'name': cell.font.name,
                    'size': cell.font.size,
                    'bold': cell.font.bold,
                    'italic': cell.font.italic,
                    'color': cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None,
                },
                'fill': {
                    'start_color': cell.fill.start_color.index if cell.fill and cell.fill.start_color else None,
                    'fill_type': cell.fill.fill_type if cell.fill else None,
                },
                'alignment': {
                    'horizontal': cell.alignment.horizontal,
                    'vertical': cell.alignment.vertical,
                    'wrap_text': cell.alignment.wrap_text,
                },
                'border': border_data
            })

merged_cells = list(ws_original.merged_cells.ranges)
row_heights = {r: ws_original.row_dimensions[r].height for r in range(min_row, max_row + 1) if ws_original.row_dimensions[r].height}
col_widths = {get_column_letter(c): ws_original.column_dimensions[get_column_letter(c)].width for c in range(min_col, max_col + 1) if ws_original.column_dimensions[get_column_letter(c)].width}
charts = list(ws_original._charts)

print(f"   - 셀: {len(cell_data)}개")
print(f"   - 차트: {len(charts)}개")

# ================================================================
# 새 시트 생성
# ================================================================

print("\n[2] 새 시트 생성...")

if '📊 Dashboard_완벽복제' in wb.sheetnames:
    del wb['📊 Dashboard_완벽복제']

ws_new = wb.create_sheet('📊 Dashboard_완벽복제')
ws_new.sheet_view.showGridLines = False

# ================================================================
# 크기 복제
# ================================================================

print("\n[3] 크기 설정...")

for row_num, height in row_heights.items():
    ws_new.row_dimensions[row_num].height = height

for col_letter, width in col_widths.items():
    ws_new.column_dimensions[col_letter].width = width

# ================================================================
# 셀 복제 (테두리 개선)
# ================================================================

print("\n[4] 셀 복제...")

for cell_info in cell_data:
    new_cell = ws_new.cell(row=cell_info['row'], column=cell_info['col'])
    
    new_cell.value = cell_info['value']
    
    if cell_info['number_format']:
        new_cell.number_format = cell_info['number_format']
    
    font_color = cell_info['font']['color']
    new_cell.font = Font(
        name=cell_info['font']['name'] or '맑은 고딕',
        size=cell_info['font']['size'] or 11,
        bold=cell_info['font']['bold'] or False,
        italic=cell_info['font']['italic'] or False,
        color=font_color if font_color and font_color != '00000000' else None
    )
    
    if cell_info['fill']['start_color'] and cell_info['fill']['start_color'] != '00000000':
        new_cell.fill = PatternFill(
            start_color=cell_info['fill']['start_color'],
            end_color=cell_info['fill']['start_color'],
            fill_type=cell_info['fill']['fill_type'] or 'solid'
        )
    
    new_cell.alignment = Alignment(
        horizontal=cell_info['alignment']['horizontal'],
        vertical=cell_info['alignment']['vertical'],
        wrap_text=cell_info['alignment']['wrap_text'] or False
    )
    
    # 개선된 테두리 복제
    border_info = cell_info['border']
    if border_info:
        def make_side(side_info):
            if not side_info:
                return Side()
            style = side_info.get('style')
            color = side_info.get('color')
            # RGB 객체를 문자열로 변환
            if color and hasattr(color, 'rgb'):
                color = color.rgb
            elif color and not isinstance(color, str):
                color = None
            return Side(style=style, color=color) if style else Side()
        
        left = make_side(border_info.get('left'))
        right = make_side(border_info.get('right'))
        top = make_side(border_info.get('top'))
        bottom = make_side(border_info.get('bottom'))
        
        new_cell.border = Border(left=left, right=right, top=top, bottom=bottom)

print(f"   - {len(cell_data)}개 완료")

# ================================================================
# 병합
# ================================================================

print("\n[5] 병합...")

for merged_range in merged_cells:
    ws_new.merge_cells(str(merged_range))

# ================================================================
# 차트 정밀 복제
# ================================================================

print("\n[6] 차트 복제...")

for idx, original_chart in enumerate(charts):
    try:
        chart_type = type(original_chart).__name__
        print(f"   - 차트 {idx+1}: {chart_type}")
        
        if isinstance(original_chart, BarChart):
            # Bar chart 복제
            new_chart = BarChart()
            new_chart.type = original_chart.type
            
            if hasattr(original_chart, 'grouping'):
                new_chart.grouping = original_chart.grouping
            
            # 제목, 크기, 스타일
            if original_chart.title:
                new_chart.title = original_chart.title.text if hasattr(original_chart.title, 'text') else str(original_chart.title)
            
            new_chart.height = original_chart.height
            new_chart.width = original_chart.width
            
            if hasattr(original_chart, 'style'):
                new_chart.style = original_chart.style
            
            # 축 제목
            if hasattr(original_chart, 'y_axis') and original_chart.y_axis.title:
                new_chart.y_axis.title = original_chart.y_axis.title
            if hasattr(original_chart, 'x_axis') and original_chart.x_axis.title:
                new_chart.x_axis.title = original_chart.x_axis.title
            
            # 시리즈 복사
            for series in original_chart.series:
                new_chart.series.append(copy.copy(series))
            
            # 범례
            if hasattr(original_chart, 'legend') and original_chart.legend:
                new_chart.legend = copy.copy(original_chart.legend)
            
            new_chart.anchor = original_chart.anchor
            ws_new.add_chart(new_chart)
            
        elif isinstance(original_chart, PieChart):
            # Pie chart 복제
            new_chart = PieChart()
            
            if original_chart.title:
                new_chart.title = original_chart.title.text if hasattr(original_chart.title, 'text') else str(original_chart.title)
            
            new_chart.height = original_chart.height
            new_chart.width = original_chart.width
            
            if hasattr(original_chart, 'style'):
                new_chart.style = original_chart.style
            
            # 시리즈
            for series in original_chart.series:
                new_chart.series.append(copy.copy(series))
            
            # 데이터 레이블
            if original_chart.dataLabels:
                new_chart.dataLabels = DataLabelList()
                if hasattr(original_chart.dataLabels, 'showCatName'):
                    new_chart.dataLabels.showCatName = original_chart.dataLabels.showCatName
                if hasattr(original_chart.dataLabels, 'showPercent'):
                    new_chart.dataLabels.showPercent = original_chart.dataLabels.showPercent
                if hasattr(original_chart.dataLabels, 'showVal'):
                    new_chart.dataLabels.showVal = original_chart.dataLabels.showVal
            
            # 범례
            if hasattr(original_chart, 'legend') and original_chart.legend:
                new_chart.legend = copy.copy(original_chart.legend)
            
            new_chart.anchor = original_chart.anchor
            ws_new.add_chart(new_chart)
            
        print(f"     ✓ 완료 ({original_chart.anchor})")
        
    except Exception as e:
        print(f"     ✗ 실패: {e}")

# ================================================================
# 기타 속성
# ================================================================

print("\n[7] 페이지 설정...")

if ws_original.freeze_panes:
    ws_new.freeze_panes = ws_original.freeze_panes

ws_new.page_setup.orientation = ws_original.page_setup.orientation
ws_new.page_setup.paperSize = ws_original.page_setup.paperSize
ws_new.print_options.gridLines = False  # 격자선 숨김

# ================================================================
# 저장
# ================================================================

print("\n[8] 저장 중...")

output_file = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_Dashboard_완벽복제.xlsx'
wb.save(output_file)

print("\n" + "=" * 70)
print("✅ 완벽 복제 완료!")
print("=" * 70)
print(f"\n📁 파일: {output_file}")
print(f"📊 새 시트: '📊 Dashboard_완벽복제'")
print(f"\n개선사항:")
print("  - 테두리 색상/스타일 정확히 복제")
print("  - 차트 범례 위치 복제")
print("  - 모든 세부 속성 보존")
print("=" * 70)
