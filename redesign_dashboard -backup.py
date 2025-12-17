import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

# Load workbook
print("Loading workbook...")
wb = load_workbook(
    r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251213_수식연결_가계부엔진.xlsx'
)

# Load data for analysis - to get top categories
print("Loading data...")
ws_raw = wb['📋 T_RawData']
data = []
for row in ws_raw.iter_rows(min_row=3, values_only=True):
    if row and len(row) > 9 and row[0] and row[2] and row[6]:
        data.append({
            '구분': row[2],
            '대분류': row[3] if row[3] else "",
            '금액': row[6],
            'Flow_Filter': row[9] if row[9] is not None else 1
        })

df = pd.DataFrame(data)
if len(df) == 0:
    print("No data found!")
    top_categories = ["월세", "식비", "자동차", "생활", "카페/간식", "온라인쇼핑", "금융", "주거/통신", "문화/여가", "가족"]
    num_months = 13
else:
    df_filtered = df[df['Flow_Filter'] == 1].copy()
    expense = df_filtered[df_filtered['구분'] == '지출'].copy()
    
    # Get top 10 categories
    if len(expense) > 0:
        cat_summary = expense.groupby('대분류')['금액'].sum().sort_values(ascending=False).head(10)
        top_categories = cat_summary.index.tolist()
    else:
        top_categories = ["월세", "식비", "자동차", "생활", "카페/간식", "온라인쇼핑", "금융", "주거/통신", "문화/여가", "가족"]
    
    num_months = 13  # Fixed value

print(f"Top 10 categories: {top_categories}")

# Get or create Dashboard sheet
if '📊 Dashboard 분석' in wb.sheetnames:
    del wb['📊 Dashboard 분석']

ws_dash = wb.create_sheet('📊 Dashboard 분석', 1)
ws_dash.sheet_view.showGridLines = False

# Colors
HEADER_COLOR = "2C3E50"
ACCENT_COLOR = "3498DB"
SUCCESS_COLOR = "27AE60"
DANGER_COLOR = "E74C3C"
WARNING_COLOR = "F39C12"
LIGHT_BG = "ECF0F1"
TABLE_HEADER = "34495E"

print(f"분석 기간: {num_months}개월")

# ==================================================================
# Header
# ==================================================================
ws_dash['B2'] = "📊 가계부 분석 대시보드"
ws_dash['B2'].font = Font(size=20, bold=True, color=HEADER_COLOR, name="맑은 고딕")
ws_dash.merge_cells('B2:O2')
ws_dash['B2'].alignment = Alignment(horizontal='center', vertical='center')
ws_dash.row_dimensions[2].height = 35

ws_dash['B3'] = f"분석 기간: {num_months}개월 | T_RawData 기반 실시간 연동"
ws_dash['B3'].font = Font(size=10, italic=True, color="7F7F7F", name="맑은 고딕")
ws_dash.merge_cells('B3:O3')
ws_dash['B3'].alignment = Alignment(horizontal='center')

# ==================================================================
# 1. Category Top 10 Table with FORMULAS
# ==================================================================
cat_section_start = 5

# Title
ws_dash[f'B{cat_section_start}'] = "🔥 대분류별 지출 순위 Top 10"
ws_dash[f'B{cat_section_start}'].font = Font(size=14, bold=True, color=DANGER_COLOR, name="맑은 고딕")
ws_dash.merge_cells(f'B{cat_section_start}:G{cat_section_start}')
ws_dash[f'B{cat_section_start}'].alignment = Alignment(horizontal='center', vertical='center')
ws_dash[f'B{cat_section_start}'].fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
ws_dash.row_dimensions[cat_section_start].height = 25

# Headers
cat_headers_row = cat_section_start + 1
headers = ['순위', '카테고리', '금액', '비율', '월평균']
for col_idx, header in enumerate(headers, 2):
    cell = ws_dash.cell(row=cat_headers_row, column=col_idx)
    cell.value = header
    cell.fill = PatternFill(start_color=TABLE_HEADER, end_color=TABLE_HEADER, fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF", size=10, name="맑은 고딕")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'), bottom=Side(style='medium')
    )

# Write data with FORMULAS
for rank, category in enumerate(top_categories, 1):
    idx = cat_headers_row + rank
    
    # Rank
    cell = ws_dash.cell(row=idx, column=2)
    cell.value = rank
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True, size=10, color=ACCENT_COLOR, name="맑은 고딕")
    
    # Category name
    cell = ws_dash.cell(row=idx, column=3)
    cell.value = category
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.font = Font(size=10, name="맑은 고딕")
    
    # Amount - SUMIFS formula
    cell = ws_dash.cell(row=idx, column=4)
    cell.value = f'=SUMIFS(\'📋 T_RawData\'!G:G,\'📋 T_RawData\'!C:C,"지출",\'📋 T_RawData\'!D:D,"{category}",\'📋 T_RawData\'!J:J,1)'
    cell.number_format = '₩#,##0'
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.font = Font(size=10, name="맑은 고딕")
    
    # Percentage - formula (amount / total)
    # Total is in cell D17 (sum of all expenses)
    cell = ws_dash.cell(row=idx, column=5)
    cell.value = f'=D{idx}/$D$17'
    cell.number_format = '0.0%'
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=10, name="맑은 고딕")
    
    # Monthly Average - formula (amount / num_months)
    cell = ws_dash.cell(row=idx, column=6)
    cell.value = f'=D{idx}/{num_months}'
    cell.number_format = '₩#,##0'
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.font = Font(size=10, name="맑은 고딕")
    
    # Borders
    for col in range(2, 7):
        ws_dash.cell(row=idx, column=col).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

# Total row
total_row = cat_headers_row + 11
ws_dash.cell(row=total_row, column=2, value="합계")
ws_dash.cell(row=total_row, column=2).font = Font(bold=True, size=10, name="맑은 고딕")
ws_dash.cell(row=total_row, column=2).alignment = Alignment(horizontal='center')

ws_dash.cell(row=total_row, column=3, value="전체 지출")
ws_dash.cell(row=total_row, column=3).font = Font(bold=True, size=10, name="맑은 고딕")

# Total amount formula
ws_dash.cell(row=total_row, column=4).value = '=SUMIFS(\'📋 T_RawData\'!G:G,\'📋 T_RawData\'!C:C,"지출",\'📋 T_RawData\'!J:J,1)'
ws_dash.cell(row=total_row, column=4).number_format = '₩#,##0'
ws_dash.cell(row=total_row, column=4).font = Font(bold=True, size=11, color=DANGER_COLOR, name="맑은 고딕")
ws_dash.cell(row=total_row, column=4).alignment = Alignment(horizontal='right')

ws_dash.cell(row=total_row, column=5, value="100.0%")
ws_dash.cell(row=total_row, column=5).font = Font(bold=True, size=10, name="맑은 고딕")
ws_dash.cell(row=total_row, column=5).alignment = Alignment(horizontal='center')

for col in range(2, 7):
    ws_dash.cell(row=total_row, column=col).border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'), bottom=Side(style='medium')
    )

# ==================================================================
# 2. Pie Chart (Top 5)
# ==================================================================
chart_row = cat_section_start

ws_dash[f'I{chart_row}'] = "💰 지출 구조 (Top 5)"
ws_dash[f'I{chart_row}'].font = Font(size=14, bold=True, color=DANGER_COLOR, name="맑은 고딕")
ws_dash.merge_cells(f'I{chart_row}:M{chart_row}')
ws_dash[f'I{chart_row}'].alignment = Alignment(horizontal='center', vertical='center')
ws_dash[f'I{chart_row}'].fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")

# Create pie chart
pie = PieChart()
pie.title = "카테고리별 지출 비중"
pie.height = 13
pie.width = 14
pie.style = 10

# Reference to category names and amounts
pie_labels = Reference(ws_dash, min_col=3, min_row=cat_headers_row+1, max_row=cat_headers_row+5)
pie_data = Reference(ws_dash, min_col=4, min_row=cat_headers_row, max_row=cat_headers_row+5)

pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(pie_labels)

# Data labels
pie.dataLabels = DataLabelList()
pie.dataLabels.showCatName = True
pie.dataLabels.showPercent = True
pie.dataLabels.showVal = False

ws_dash.add_chart(pie, f"I{chart_row + 2}")

# ==================================================================
# 3. Spending Pattern Summary with FORMULAS
# ==================================================================
pattern_start = 18

ws_dash[f'B{pattern_start}'] = "📈 소비 패턴 분석"
ws_dash[f'B{pattern_start}'].font = Font(size=14, bold=True, color=ACCENT_COLOR, name="맑은 고딕")
ws_dash.merge_cells(f'B{pattern_start}:G{pattern_start}')
ws_dash[f'B{pattern_start}'].alignment = Alignment(horizontal='center', vertical='center')
ws_dash[f'B{pattern_start}'].fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
ws_dash.row_dimensions[pattern_start].height = 25

# Headers
pattern_headers = pattern_start + 1
for col_idx, header in enumerate(['구분', '카테고리', '금액', '비중'], 2):
    cell = ws_dash.cell(row=pattern_headers, column=col_idx)
    cell.value = header
    cell.fill = PatternFill(start_color=TABLE_HEADER, end_color=TABLE_HEADER, fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF", size=10, name="맑은 고딕")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Pattern definitions
patterns = [
    ("🏠 고정비", ["월세", "주거/통신", "자동차"]),
    ("🍽️ 식생활", ["식비", "카페/간식"]),
    ("💳 유동비", ["온라인쇼핑", "생활"]),
    ("🎭 여가", ["문화/여가", "술/유흥"]),
]

for idx, (type_, categories) in enumerate(patterns, pattern_headers + 1):
    # Type
    ws_dash.cell(row=idx, column=2, value=type_)
    ws_dash.cell(row=idx, column=2).font = Font(size=10, bold=True, name="맑은 고딕")
    ws_dash.cell(row=idx, column=2).alignment = Alignment(horizontal='left', vertical='center')
    
    # Categories list
    cat_list = " + ".join(categories)
    ws_dash.cell(row=idx, column=3, value=cat_list)
    ws_dash.cell(row=idx, column=3).font = Font(size=9, name="맑은 고딕")
    ws_dash.cell(row=idx, column=3).alignment = Alignment(horizontal='left', vertical='center')
    
    # Amount - SUMIFS for multiple categories
    formula_parts = []
    for cat in categories:
        formula_parts.append(f'SUMIFS(\'📋 T_RawData\'!G:G,\'📋 T_RawData\'!C:C,"지출",\'📋 T_RawData\'!D:D,"{cat}",\'📋 T_RawData\'!J:J,1)')
    
    formula = "=" + "+".join(formula_parts)
    ws_dash.cell(row=idx, column=4, value=formula)
    ws_dash.cell(row=idx, column=4).number_format = '₩#,##0'
    ws_dash.cell(row=idx, column=4).font = Font(size=10, name="맑은 고딕")
    ws_dash.cell(row=idx, column=4).alignment = Alignment(horizontal='right', vertical='center')
    
    # Percentage
    ws_dash.cell(row=idx, column=5, value=f'=D{idx}/$D$17')
    ws_dash.cell(row=idx, column=5).number_format = '0.0%'
    ws_dash.cell(row=idx, column=5).font = Font(size=10, bold=True, color=DANGER_COLOR, name="맑은 고딕")
    ws_dash.cell(row=idx, column=5).alignment = Alignment(horizontal='center', vertical='center')
    
    # Borders
    for col in range(2, 6):
        ws_dash.cell(row=idx, column=col).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

# Column widths
ws_dash.column_dimensions['B'].width = 15
ws_dash.column_dimensions['C'].width = 18
ws_dash.column_dimensions['D'].width = 18
ws_dash.column_dimensions['E'].width = 12
ws_dash.column_dimensions['F'].width = 12
ws_dash.column_dimensions['G'].width = 15

# Save
output_file = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_수식연결_가계부엔진_최종.xlsx'
wb.save(output_file)

print("\n" + "=" * 70)
print("✅ Dashboard 완성!")
print("=" * 70)
print(f"\n📁 파일: {output_file}")
print(f"\n✨ 모든 데이터가 T_RawData 기반 수식으로 연결되었습니다!")
print(f"   - Top 10 카테고리: SUMIFS 수식")
print(f"   - 비율: 자동 계산 수식")
print(f"   - 월평균: 자동 계산 수식")
print(f"   - 소비 패턴: SUMIFS 수식")
print("\n" + "=" * 70)
