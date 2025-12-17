
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

file_path = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_Dashboard_v25_direct_let.xlsx'
wb = load_workbook(file_path, data_only=True)
ws = wb['📋 T_RawData']

print("=== Grid Scan A1:L5 ===")
for r in range(1, 6):
    row_vals = []
    for c in range(1, 13): # A to L
        val = ws.cell(row=r, column=c).value
        # If val is None, search properly?
        row_vals.append(str(val))
    print(f"Row {r}: {row_vals}")
