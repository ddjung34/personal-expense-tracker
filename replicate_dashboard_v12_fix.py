"""
Dashboard 정밀 복제 v12 (Fix: Titles via Cell Modification)
1. TypeError 방지: series.title 직접 할당 대신, 시트의 헤더 셀 값을 변경하여 자동 반영
2. 축 제목, 색상 등 나머지 설정 유지
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.axis import TextAxis

print("=" * 70)
print("Dashboard 정밀 복제 v12 (Fix)")
print("=" * 70)

wb = load_workbook(
    r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_수식연결_가계부엔진_최종.xlsx'
)
dashboard_sheet_name = None
for s in wb.sheetnames:
    if 'dashboard' in s.lower() and '복제' not in s and '최종' not in s and 'v' not in s:
        dashboard_sheet_name = s; break
if not dashboard_sheet_name: dashboard_sheet_name = wb.sheetnames[1]
print(f"✅ 원본 시트: '{dashboard_sheet_name}'")
ws_original = wb[dashboard_sheet_name]

# [1] 데이터 복제
cell_data = []
min_row, max_row = 1, 60
min_col, max_col = 1, 25
for row in range(min_row, max_row + 1):
    for col in range(min_col, max_col + 1):
        cell = ws_original.cell(row=row, column=col)
        b_data = {}
        if cell.border:
            for s in ['left', 'right', 'top', 'bottom']:
                side = getattr(cell.border, s)
                if side:
                    c_val = None
                    if side.color and hasattr(side.color, 'rgb') and isinstance(side.color.rgb, str) and len(side.color.rgb) <= 8:
                        c_val = side.color.rgb
                    b_data[s] = {'style': side.style, 'color': c_val}
        f_color = None
        if cell.font.color and hasattr(cell.font.color, 'rgb') and isinstance(cell.font.color.rgb, str):
            f_color = cell.font.color.rgb
        bg_color = None
        if cell.fill and cell.fill.start_color and cell.fill.start_color.index != '00000000':
             bg_color = cell.fill.start_color.index
        if cell.value or bg_color or b_data:
            cell_data.append({
                'row': row, 'col': col, 'value': cell.value, 'fmt': cell.number_format,
                'font': {'name': cell.font.name, 'sz': cell.font.size, 'b': cell.font.bold, 'color': f_color},
                'fill': {'c': bg_color, 't': cell.fill.fill_type},
                'align': {'h': cell.alignment.horizontal, 'v': cell.alignment.vertical, 'w': cell.alignment.wrap_text},
                'border': b_data
            })

merged_cells = list(ws_original.merged_cells.ranges)
row_heights = {r: ws_original.row_dimensions[r].height for r in range(min_row, max_row + 1) if ws_original.row_dimensions[r].height}
col_widths = {get_column_letter(c): ws_original.column_dimensions[get_column_letter(c)].width for c in range(min_col, max_col + 1) if ws_original.column_dimensions[get_column_letter(c)].width}

if '📊 Dashboard_v12' in wb.sheetnames: del wb['📊 Dashboard_v12']
ws_new = wb.create_sheet('📊 Dashboard_v12')
ws_new.sheet_view.showGridLines = False

for r, h in row_heights.items(): ws_new.row_dimensions[r].height = h
for c, w in col_widths.items(): ws_new.column_dimensions[c].width = w
for d in cell_data:
    c = ws_new.cell(d['row'], d['col'])
    c.value = d['value']
    if d['fmt']: c.number_format = d['fmt']
    c.font = Font(name=d['font']['name'], size=d['font']['sz'], bold=d['font']['b'], color=d['font']['color'])
    if d['fill']['c']: c.fill = PatternFill(start_color=d['fill']['c'], end_color=d['fill']['c'], fill_type=d['fill']['t'])
    c.alignment = Alignment(horizontal=d['align']['h'], vertical=d['align']['v'], wrap_text=d['align']['w'])
    borders = {}
    for k, v in d['border'].items():
        if v: borders[k] = Side(style=v['style'], color=v['color'])
    if borders: c.border = Border(**borders)
for m in merged_cells: ws_new.merge_cells(str(m))

# ==================================================================================
# [4] 차트 생성 (v12 Fix)
# ==================================================================================
print("차트 v12 생성 중...")

# [중요] 범례 이름을 위해 실제 셀 값 변경 (TypeError 방지)
# 38행 D열: '금액' -> '카테고리별 지출'
# 38행 F열: '월평균' -> '월평균 지출' (이미 비슷할 수 있음)
ws_new.cell(row=38, column=4).value = "카테고리별 지출"
ws_new.cell(row=38, column=6).value = "월평균 지출"
# 11행은 이미 '수입', '지출', '합계'이므로 변경 불필요

# --- 1. 주요지표 콤보 차트 ---
c1 = BarChart()
c1.type = "col"
c1.grouping = "clustered"
c1.title = "주요 지표 추이"
c1.legend.position = "tr"
c1.x_axis.title = "월"   
c1.y_axis.title = "금액" 
c1.y_axis.tickLblPos = "nextTo"
c1.y_axis.majorTickMark = "out"

# 데이터 (Row 11: Header, Row 12~23: Data)
data_bar = Reference(ws_new, min_col=4, min_row=11, max_col=5, max_row=23)
c1.add_data(data_bar, titles_from_data=True)
cats = Reference(ws_new, min_col=3, min_row=12, max_row=23)
c1.set_categories(cats)

# 색상 (수입:파랑, 지출:빨강)
try:
    c1.series[0].graphicalProperties.solidFill = "4472C4"
    c1.series[0].graphicalProperties.line.solidFill = "4472C4"
    c1.series[1].graphicalProperties.solidFill = "FF0000"
    c1.series[1].graphicalProperties.line.solidFill = "FF0000"
    c1.series[1].invertIfNegative = False 
except: pass

c2 = LineChart()
data_line = Reference(ws_new, min_col=6, min_row=11, max_col=6, max_row=23)
c2.add_data(data_line, titles_from_data=True)
try: c2.series[0].graphicalProperties.line.solidFill = "92D050" 
except: pass

c1 += c2
c1.anchor = "H6"
c1.height = 13.5
c1.width = 24
ws_new.add_chart(c1)

# --- 2. 지출 구조 차트 ---
c3 = BarChart()
c3.type = "bar"
c3.style = 10
c3.title = "지출 구조 차트"
c3.legend.position = "r"
c3.x_axis.majorGridlines = None 

# 데이터 (Row 38: Header -> '카테고리별 지출' 등)
data_c3 = Reference(ws_new, min_col=4, min_row=38, max_col=4, max_row=47)
data_c3_2 = Reference(ws_new, min_col=6, min_row=38, max_col=6, max_row=47)
c3.add_data(data_c3, titles_from_data=True)
c3.add_data(data_c3_2, titles_from_data=True)
cats_c3 = Reference(ws_new, min_col=3, min_row=39, max_row=47) 
c3.set_categories(cats_c3)

# Data Labels
c3.dataLabels = DataLabelList()
c3.dataLabels.showVal = True      
c3.dataLabels.showCatName = True  

c3.anchor = "I30"
c3.height = 14
c3.width = 21

ws_new.add_chart(c3)

output_path = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_Dashboard_v12_final_fix.xlsx'
wb.save(output_path)
print(f"저장 완료: {output_path}")
