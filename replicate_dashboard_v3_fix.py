"""
Dashboard 정밀 복제 v3.1 (Fix)
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import Series
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.colors import ColorChoice

print("=" * 70)
print("Dashboard 정밀 복제 v3.1")
print("=" * 70)

wb = load_workbook(
    r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_수식연결_가계부엔진_최종.xlsx'
)

# Find Dashboard sheet
dashboard_sheet_name = None
for sheet_name in wb.sheetnames:
    if 'dashboard' in sheet_name.lower() and '복제' not in sheet_name and '분석' in sheet_name:
        dashboard_sheet_name = sheet_name
        break

if not dashboard_sheet_name:
    for sheet_name in wb.sheetnames:
        if 'dashboard' in sheet_name.lower() or '대시보드' in sheet_name:
            if '복제' not in sheet_name:
                dashboard_sheet_name = sheet_name
                break

if not dashboard_sheet_name and len(wb.sheetnames) >= 2:
    dashboard_sheet_name = wb.sheetnames[1]

print(f"✅ 원본 시트: '{dashboard_sheet_name}'")
ws_original = wb[dashboard_sheet_name]

# [1] 데이터 수집
print("[1] 데이터 수집 중...")
min_row, max_row = 1, 60
min_col, max_col = 1, 20

cell_data = []
for row in range(min_row, max_row + 1):
    for col in range(min_col, max_col + 1):
        cell = ws_original.cell(row=row, column=col)
        # 테두리 정보 안전하게 추출 함수
        def get_side_info(side):
            if not side: return None
            color_val = None
            if side.color:
                if hasattr(side.color, 'rgb'):
                    color_val = side.color.rgb
                elif hasattr(side.color, 'theme'):
                    color_val = side.color.theme # 테마 컬러는 복잡할 수 있어 일단 무시하거나 처리 필요
            return {'style': side.style, 'color': color_val}

        border_data = {}
        if cell.border:
            border_data['left'] = get_side_info(cell.border.left)
            border_data['right'] = get_side_info(cell.border.right)
            border_data['top'] = get_side_info(cell.border.top)
            border_data['bottom'] = get_side_info(cell.border.bottom)
            
        color_rgb = None
        if cell.font.color:
             if hasattr(cell.font.color, 'rgb'):
                 color_rgb = cell.font.color.rgb
             # theme color는 단순화 위해 건너뜀 (필요시 추가)
        
        fill_color = None
        if cell.fill and cell.fill.start_color and cell.fill.start_color.index != '00000000':
             fill_color = cell.fill.start_color.index
            
        if cell.value or fill_color or any(border_data.values()):
            cell_data.append({
                'row': row, 'col': col, 'value': cell.value,
                'number_format': cell.number_format,
                'font': {
                    'name': cell.font.name, 'size': cell.font.size,
                    'bold': cell.font.bold, 'italic': cell.font.italic,
                    'color': color_rgb
                },
                'fill': {
                    'start_color': fill_color,
                    'fill_type': cell.fill.fill_type if cell.fill else None
                },
                'alignment': {
                    'horizontal': cell.alignment.horizontal,
                    'vertical': cell.alignment.vertical,
                    'wrap_text': cell.alignment.wrap_text
                },
                'border': border_data
            })

merged_cells = list(ws_original.merged_cells.ranges)
row_heights = {r: ws_original.row_dimensions[r].height for r in range(min_row, max_row + 1) if ws_original.row_dimensions[r].height}
col_widths = {get_column_letter(c): ws_original.column_dimensions[get_column_letter(c)].width for c in range(min_col, max_col + 1) if ws_original.column_dimensions[get_column_letter(c)].width}

print(f"   - 셀: {len(cell_data)}개")

# [2] 새 시트 생성
print("[2] 새 시트 생성...")
if '📊 Dashboard_최종' in wb.sheetnames:
    del wb['📊 Dashboard_최종']
ws_new = wb.create_sheet('📊 Dashboard_최종')
ws_new.sheet_view.showGridLines = False

# [3] 크기 설정
print("[3] 크기 설정...")
for r, h in row_heights.items(): ws_new.row_dimensions[r].height = h
for c, w in col_widths.items(): ws_new.column_dimensions[c].width = w

# [4] 셀 복제
print("[4] 셀 복제...")
for item in cell_data:
    cell = ws_new.cell(row=item['row'], column=item['col'])
    cell.value = item['value']
    if item['number_format']: cell.number_format = item['number_format']
    
    # Font
    font_color = item['font']['color']
    # RGB값이 문자열이 아닌 경우 안전처리
    if font_color and not isinstance(font_color, str): font_color = None
    
    cell.font = Font(
        name=item['font']['name'], size=item['font']['size'],
        bold=item['font']['bold'], italic=item['font']['italic'],
        color=font_color
    )
    
    # Fill
    fill_c = item['fill']['start_color']
    if fill_c and isinstance(fill_c, str) and fill_c != '00000000':
        cell.fill = PatternFill(start_color=fill_c, end_color=fill_c, fill_type=item['fill']['fill_type'])
        
    # Alignment
    cell.alignment = Alignment(
        horizontal=item['alignment']['horizontal'], vertical=item['alignment']['vertical'],
        wrap_text=item['alignment']['wrap_text']
    )
    
    # Border
    sides = {}
    for side_key in ['left', 'right', 'top', 'bottom']:
        s_info = item['border'].get(side_key)
        if s_info:
            s_style = s_info.get('style')
            s_color = s_info.get('color')
            # Color 안전처리: 문자열이 아니거나 rgb 속성 처리
            safe_color = None
            if isinstance(s_color, str):
                safe_color = s_color
            elif hasattr(s_color, 'rgb'):
                safe_color = s_color.rgb
                
            sides[side_key] = Side(style=s_style, color=safe_color) if s_style else Side()
        else:
            sides[side_key] = Side()
            
    cell.border = Border(**sides)

# [5] 병합
print("[5] 병합...")
for r in merged_cells: ws_new.merge_cells(str(r))

# [6] Q48 테두리 추가
print("[6] Q48 테두리 추가...")
for row_num in range(38, 50):
    cell = ws_new.cell(row=row_num, column=17) # Q열
    if cell.value:
        cur = cell.border
        cell.border = Border(
            left=cur.left, right=Side(style='thin', color='000000'),
            top=cur.top, bottom=cur.bottom
        )

# [7] 굵은 테두리
print("[7] 굵은 테두리 추가...")
thick = Side(style='medium') # color='000000' 제거 (기본값 검정)
sections = ['C3:T26', 'C27:T52']
from openpyxl.utils import coordinate_to_tuple
for rng in sections:
    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(rng)
    # Top
    for c in range(min_col, max_col+1):
        cell = ws_new.cell(min_row, c)
        cell.border = Border(left=cell.border.left, right=cell.border.right, top=thick, bottom=cell.border.bottom)
    # Bottom
    for c in range(min_col, max_col+1):
        cell = ws_new.cell(max_row, c)
        cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=thick)
    # Left
    for r in range(min_row, max_row+1):
        cell = ws_new.cell(r, min_col)
        cell.border = Border(left=thick, right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom)
    # Right
    for r in range(min_row, max_row+1):
        cell = ws_new.cell(r, max_col)
        cell.border = Border(left=cell.border.left, right=thick, top=cell.border.top, bottom=cell.border.bottom)

# [8] 콤보 차트 생성
print("[8] 콤보 차트 생성...")
try:
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.overlap = 100
    chart.title = "주요 지표 추이"
    
    # Data
    income_ref = Reference(ws_new, min_col=4, min_row=10, max_row=23)
    expense_ref = Reference(ws_new, min_col=5, min_row=10, max_row=23)
    chart.add_data(income_ref, titles_from_data=False)
    chart.add_data(expense_ref, titles_from_data=False)
    
    # Line series - Simplified creation
    net_ref = Reference(ws_new, min_col=6, min_row=10, max_row=23)
    net_series = Series(net_ref)
    net_series.title = "합계"
    
    # Styling line - Simplified to avoid errors
    # 기본 선 스타일 사용 (색상/두께 지정 없이)
    # openpyxl이 자동으로 기본 스타일 적용
    pass
    
    # net_series.graphicalProperties.line ... (Removed to avoid TypeError)
    
    chart.series.append(net_series)
    
    cats = Reference(ws_new, min_col=3, min_row=10, max_row=23)
    chart.set_categories(cats)
    
    chart.legend = None
    chart.height = 13
    chart.width = 14
    chart.anchor = "H6"
    chart.y_axis.title = "금액 (원)"
    chart.x_axis.title = "월"
    
    ws_new.add_chart(chart)
    print("   - 콤보 차트 생성 성공")

except Exception as e:
    print(f"   - ⚠️ 차트 생성 실패: {e}")
    import traceback
    traceback.print_exc()

# [9] 기타 차트 복제 (간소화)
print("[9] 기타 차트 복제...")
import copy
try:
    charts = list(ws_original._charts)
    for idx, c in enumerate(charts):
        if idx == 0: continue # Skip first
        if "pie" in str(type(c)).lower():
            new_c = copy.deepcopy(c)
            new_c.anchor = c.anchor
            ws_new.add_chart(new_c)
            print(f"   - 차트 {idx} 복제")
except Exception as e:
    print(f"   - 차트 복제 중 오류: {e}")

# [10] 저장
print("[10] 저장 중...")
wb.save(r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_Dashboard_최종수정.xlsx')
print("✅ 완료!")
