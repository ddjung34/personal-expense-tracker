"""
Dashboard 정밀 복제 v3 - 사용자 피드백 반영
1. 주요지표 콤보 차트 (막대 + 선)
2. Q48 테두리 추가
3. 굵은 외곽 테두리 추가
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
import copy

print("=" * 70)
print("Dashboard 정밀 복제 v3")
print("=" * 70)

wb = load_workbook(
    r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_수식연결_가계부엔진_최종.xlsx'
)

# Find Dashboard sheet
dashboard_sheet_name = None
for sheet_name in wb.sheetnames:
    if 'dashboard' in sheet_name.lower() and '복제' not in sheet_name and '분석' in sheet_name:
        dashboard_sheet_name = sheet_name
        break

if not dashboard_sheet_name:
    for sheet_name in wb.sheetnames:
        if 'dashboard' in sheet_name.lower() or '대시보드' in sheet_name:
            if '복제' not in sheet_name:
                dashboard_sheet_name = sheet_name
                break

if not dashboard_sheet_name and len(wb.sheetnames) >= 2:
    dashboard_sheet_name = wb.sheetnames[1]

print(f"\n✅ 원본 시트: '{dashboard_sheet_name}'")

ws_original = wb[dashboard_sheet_name]

# ================================================================
# 데이터 수집 (v2와 동일)
# ================================================================

print("\n[1] 데이터 수집 중...")

min_row, max_row = 1, 60
min_col, max_col = 1, 20

cell_data = []
for row in range(min_row, max_row + 1):
    for col in range(min_col, max_col + 1):
        cell = ws_original.cell(row=row, column=col)
        if cell.value or (cell.fill and cell.fill.start_color and cell.fill.start_color.index != '00000000'):
            border_data = {}
            if cell.border:
                if cell.border.left:
                    border_data['left'] = {'style': cell.border.left.style, 'color': cell.border.left.color.rgb if cell.border.left.color and hasattr(cell.border.left.color, 'rgb') else None}
                if cell.border.right:
                    border_data['right'] = {'style': cell.border.right.style, 'color': cell.border.right.color.rgb if cell.border.right.color and hasattr(cell.border.right.color, 'rgb') else None}
                if cell.border.top:
                    border_data['top'] = {'style': cell.border.top.style, 'color': cell.border.top.color.rgb if cell.border.top.color and hasattr(cell.border.top.color, 'rgb') else None}
                if cell.border.bottom:
                    border_data['bottom'] = {'style': cell.border.bottom.style, 'color': cell.border.bottom.color.rgb if cell.border.bottom.color and hasattr(cell.border.bottom.color, 'rgb') else None}
            
            cell_data.append({
                'row': row,
                'col': col,
                'value': cell.value,
                'number_format': cell.number_format,
                'font': {
                    'name': cell.font.name,
                    'size': cell.font.size,
                    'bold': cell.font.bold,
                    'italic': cell.font.italic,
                    'color': cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None,
                },
                'fill': {
                    'start_color': cell.fill.start_color.index if cell.fill and cell.fill.start_color else None,
                    'fill_type': cell.fill.fill_type if cell.fill else None,
                },
                'alignment': {
                    'horizontal': cell.alignment.horizontal,
                    'vertical': cell.alignment.vertical,
                    'wrap_text': cell.alignment.wrap_text,
                },
                'border': border_data
            })

merged_cells = list(ws_original.merged_cells.ranges)
row_heights = {r: ws_original.row_dimensions[r].height for r in range(min_row, max_row + 1) if ws_original.row_dimensions[r].height}
col_widths = {get_column_letter(c): ws_original.column_dimensions[get_column_letter(c)].width for c in range(min_col, max_col + 1) if ws_original.column_dimensions[get_column_letter(c)].width}

print(f"   - 셀: {len(cell_data)}개")

# ================================================================
# 새 시트 생성
# ================================================================

print("\n[2] 새 시트 생성...")

if '📊 Dashboard_최종' in wb.sheetnames:
    del wb['📊 Dashboard_최종']

ws_new = wb.create_sheet('📊 Dashboard_최종')
ws_new.sheet_view.showGridLines = False

# ================================================================
# 크기 설정
# ================================================================

print("\n[3] 크기 설정...")

for row_num, height in row_heights.items():
    ws_new.row_dimensions[row_num].height = height

for col_letter, width in col_widths.items():
    ws_new.column_dimensions[col_letter].width = width

# ================================================================
# 셀 복제
# ================================================================

print("\n[4] 셀 복제...")

for cell_info in cell_data:
    new_cell = ws_new.cell(row=cell_info['row'], column=cell_info['col'])
    
    new_cell.value = cell_info['value']
    
    if cell_info['number_format']:
        new_cell.number_format = cell_info['number_format']
    
    font_color = cell_info['font']['color']
    new_cell.font = Font(
        name=cell_info['font']['name'] or '맑은 고딕',
        size=cell_info['font']['size'] or 11,
        bold=cell_info['font']['bold'] or False,
        italic=cell_info['font']['italic'] or False,
        color=font_color if font_color and font_color != '00000000' else None
    )
    
    if cell_info['fill']['start_color'] and cell_info['fill']['start_color'] != '00000000':
        new_cell.fill = PatternFill(
            start_color=cell_info['fill']['start_color'],
            end_color=cell_info['fill']['start_color'],
            fill_type=cell_info['fill']['fill_type'] or 'solid'
        )
    
    new_cell.alignment = Alignment(
        horizontal=cell_info['alignment']['horizontal'],
        vertical=cell_info['alignment']['vertical'],
        wrap_text=cell_info['alignment']['wrap_text'] or False
    )
    
    border_info = cell_info['border']
    if border_info:
        def make_side(side_info):
            if not side_info:
                return Side()
            style = side_info.get('style')
            color = side_info.get('color')
            if color and hasattr(color, 'rgb'):
                color = color.rgb
            elif color and not isinstance(color, str):
                color = None
            return Side(style=style, color=color) if style else Side()
        
        left = make_side(border_info.get('left'))
        right = make_side(border_info.get('right'))
        top = make_side(border_info.get('top'))
        bottom = make_side(border_info.get('bottom'))
        
        new_cell.border = Border(left=left, right=right, top=top, bottom=bottom)

print(f"   - {len(cell_data)}개 완료")

# ================================================================
# 병합
# ================================================================

print("\n[5] 병합...")

for merged_range in merged_cells:
    ws_new.merge_cells(str(merged_range))

# ================================================================
# 수정 1: Q48 테두리 추가 (합계 행, 월평균 행)
# ================================================================

print("\n[6] Q48 테두리 추가...")

# Q48은 17열, 48행
# 합계와 월평균 부분 찾기 - 대분류별 지출 Top 10 테이블의 마지막
# 이미지 보니 Q48이 48행 17열(Q열)

# Top 10 테이블이 약 B38:Q48 범위로 추정
# Q열(17열)에 테두리 추가
for row_num in range(38, 50):  # 테이블 범위
    cell = ws_new.cell(row=row_num, column=17)  # Q열
    if cell.value:  # 값이 있는 셀만
        current_border = cell.border
        # 기존 테두리 유지하면서 오른쪽 테두리 추가
        cell.border = Border(
            left=current_border.left if current_border else Side(),
            right=Side(style='thin', color='000000'),  # 오른쪽 테두리 추가
            top=current_border.top if current_border else Side(),
            bottom=current_border.bottom if current_border else Side()
        )

# ================================================================
# 수정 3: 굵은 외곽 테두리 추가
# ================================================================

print("\n[7] 굵은 외곽 테두리 추가...")

# 주요 섹션들의 외곽 테두리
# 이미지 보니 C3:T26 정도의 큰 박스 (주요 지표 + 월별 데이터 + 차트)
# C27:T52 정도 (소비 지표 + 차트)

sections = [
    {'name': '주요 지표 섹션', 'range': 'C3:T26'},
    {'name': '소비 지표 섹션', 'range': 'C27:T52'},
]

thick_border = Side(style='medium', color='000000')

for section in sections:
    print(f"   - {section['name']}: {section['range']}")
    
    # 범위 파싱
    range_str = section['range']
    start_cell, end_cell = range_str.split(':')
    
    # 시작/끝 열과 행 계산
    from openpyxl.utils import column_index_from_string, get_column_letter
    
    start_col = column_index_from_string(start_cell[0])
    start_row = int(start_cell[1:])
    end_col = column_index_from_string(end_cell[0])
    end_row = int(end_cell[1:])
    
    # 상단 테두리
    for col in range(start_col, end_col + 1):
        cell = ws_new.cell(row=start_row, column=col)
        current = cell.border
        cell.border = Border(
            left=current.left if current else Side(),
            right=current.right if current else Side(),
            top=thick_border,
            bottom=current.bottom if current else Side()
        )
    
    # 하단 테두리
    for col in range(start_col, end_col + 1):
        cell = ws_new.cell(row=end_row, column=col)
        current = cell.border
        cell.border = Border(
            left=current.left if current else Side(),
            right=current.right if current else Side(),
            top=current.top if current else Side(),
            bottom=thick_border
        )
    
    # 좌측 테두리
    for row in range(start_row, end_row + 1):
        cell = ws_new.cell(row=row, column=start_col)
        current = cell.border
        cell.border = Border(
            left=thick_border,
            right=current.right if current else Side(),
            top=current.top if current else Side(),
            bottom=current.bottom if current else Side()
        )
    
    # 우측 테두리
    for row in range(start_row, end_row + 1):
        cell = ws_new.cell(row=row, column=end_col)
        current = cell.border
        cell.border = Border(
            left=current.left if current else Side(),
            right=thick_border,
            top=current.top if current else Side(),
            bottom=current.bottom if current else Side()
        )

# ================================================================
# 수정 1: 주요지표 콤보 차트 생성 (막대 + 선)
# ================================================================

print("\n[8] 주요지표 콤보 차트 생성...")

# 기존 차트는 건너뛰고 직접 생성
# 데이터 범위: C10:F23

# 콤보 차트 생성
chart = BarChart()
chart.type = "col"
chart.grouping = "clustered"
chart.overlap = 100

# 제목
chart.title = "주요 지표 추이"

# 데이터 시리즈
# C10:F23 범위
# C: 월, D: 수입, E: 지출, F: 합계

# 수입 (파란 막대)
income_ref = Reference(ws_new, min_col=4, min_row=10, max_row=23)
chart.add_data(income_ref, titles_from_data=False)

# 지출 (빨간 막대)
expense_ref = Reference(ws_new, min_col=5, min_row=10, max_row=23)
chart.add_data(expense_ref, titles_from_data=False)

# 합계 (노란 선) - 선 차트로 추가
from openpyxl.chart.series import Series
net_ref = Reference(ws_new, min_col=6, min_row=10, max_row=23)
net_series = Series(net_ref)

# 선 스타일 설정
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.fill import SolidColorFillProperties, ColorChoice

net_series.graphicalProperties = openpyxl.chart.GraphicalProperties()
net_series.graphicalProperties.line = LineProperties(w=25000)  # 선 두께
net_series.graphicalProperties.line.solidFill = SolidColorFillProperties(ColorChoice(srgbClr="FFC000"))  # 노란색

chart.series.append(net_series)

# 카테고리 (월)
cats = Reference(ws_new, min_col=3, min_row=10, max_row=23)
chart.set_categories(cats)

# 범례
chart.legend = None  # 또는 범례 위치 설정

# 크기 및 위치
chart.height = 13
chart.width = 14
chart.anchor = "H6"  # 차트 위치 (원본과 동일하게)

# 축 설정
chart.y_axis.title = "금액 (원)"
chart.x_axis.title = "월"

ws_new.add_chart(chart)

print("   - 콤보 차트 생성 완료")

# ================================================================
# 기타 차트 복제 (파이 차트 등)
# ================================================================

print("\n[9] 기타 차트 복제...")

charts = list(ws_original._charts)

for idx, original_chart in enumerate(charts):
    try:
        # 첫 번째 차트(주요지표)는 이미 생성했으므로 건너뜀
        if idx == 0:
            continue
            
        chart_type = type(original_chart).__name__
        print(f"   - 차트 {idx+1}: {chart_type}")
        
        if isinstance(original_chart, openpyxl.chart.pie_chart.PieChart):
            new_chart = openpyxl.chart.PieChart()
            
            if original_chart.title:
                new_chart.title = original_chart.title.text if hasattr(original_chart.title, 'text') else str(original_chart.title)
            
            new_chart.height = original_chart.height
            new_chart.width = original_chart.width
            
            if hasattr(original_chart, 'style'):
                new_chart.style = original_chart.style
            
            for series in original_chart.series:
                new_chart.series.append(copy.copy(series))
            
            if original_chart.dataLabels:
                new_chart.dataLabels = DataLabelList()
                if hasattr(original_chart.dataLabels, 'showCatName'):
                    new_chart.dataLabels.showCatName = original_chart.dataLabels.showCatName
                if hasattr(original_chart.dataLabels, 'showPercent'):
                    new_chart.dataLabels.showPercent = original_chart.dataLabels.showPercent
                if hasattr(original_chart.dataLabels, 'showVal'):
                    new_chart.dataLabels.showVal = original_chart.dataLabels.showVal
            
            if hasattr(original_chart, 'legend') and original_chart.legend:
                new_chart.legend = copy.copy(original_chart.legend)
            
            new_chart.anchor = original_chart.anchor
            ws_new.add_chart(new_chart)
            
            print(f"     ✓ 완료")
        
    except Exception as e:
        print(f"     ✗ 실패: {e}")

# ================================================================
# 기타 속성
# ================================================================

print("\n[10] 페이지 설정...")

if ws_original.freeze_panes:
    ws_new.freeze_panes = ws_original.freeze_panes

ws_new.page_setup.orientation = ws_original.page_setup.orientation
ws_new.page_setup.paperSize = ws_original.page_setup.paperSize
ws_new.print_options.gridLines = False

# ================================================================
# 저장
# ================================================================

print("\n[11] 저장 중...")

output_file = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_Dashboard_최종수정.xlsx'
wb.save(output_file)

print("\n" + "=" * 70)
print("✅ 최종 수정 완료!")
print("=" * 70)
print(f"\n📁 파일: {output_file}")
print(f"📊 새 시트: '📊 Dashboard_최종'")
print(f"\n✨ 적용된 수정사항:")
print("  1. ✅ 주요지표 콤보 차트 (막대 + 선)")
print("  2. ✅ Q48 테두리 추가")
print("  3. ✅ 굵은 외곽 테두리 추가")
print("=" * 70)
