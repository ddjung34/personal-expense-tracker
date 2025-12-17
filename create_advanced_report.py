import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import LineChart, PieChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

INPUT_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-10~2025-12-10.xlsx"
OUTPUT_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\{date}_가계부_전문분석보고서.xlsx".format(
    date=datetime.now().strftime("%Y%m%d")
)

def create_advanced_excel_report():
    print(f"📊 Loading data from: {INPUT_FILE}")
    
    
    # Read the data
    df = pd.read_excel(INPUT_FILE, sheet_name='Rawdata', engine='openpyxl')
    
    # Data cleaning & mapping
    df['date'] = pd.to_datetime(df['날짜'], errors='coerce')
    df['time'] = df['시간'].astype(str) if '시간' in df.columns else ''
    df['type'] = df['타입']
    df['main_category'] = df['대분류'].fillna('')
    df['sub_category'] = df['소분류'].fillna('')
    df['amount'] = pd.to_numeric(df['금액'], errors='coerce')
    df['currency'] = df['화폐'] if '화폐' in df.columns else 'KRW'
    df['payment_method'] = df['결제수단'].fillna('')
    df['merchant'] = df['내용'].fillna('')
    df['memo'] = df['메모'].fillna('')
    
    # Remove invalid data
    df = df.dropna(subset=['date', 'amount', 'type'])
    df = df.sort_values('date')
    
    print(f"✅ Loaded {len(df)} transactions")
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Define color scheme (consistent throughout)
    HEADER_COLOR = "34495E"  # Dark gray-blue
    ACCENT_COLOR = "3498DB"  # Blue
    SUCCESS_COLOR = "27AE60"  # Green
    DANGER_COLOR = "E74C3C"   # Red
    LIGHT_BG = "ECF0F1"       # Light gray
    
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="맑은 고딕", size=11)
    normal_font = Font(name="맑은 고딕", size=10)
    
    # ==================================================
    # SHEET 1: 📋 Clean Raw Data with Flow_Filter
    # ==================================================
    ws_data = wb.create_sheet("📋 Clean Raw Data", 0)
    ws_data.sheet_view.showGridLines = False
    
    # Prepare data with additional Flow_Filter column
    clean_df = df[['date', 'time', 'type', 'main_category', 'sub_category', 'merchant', 'amount', 'payment_method', 'memo']].copy()
    clean_df.columns = ['날짜', '시간', '구분', '대분류', '소분류', '내용', '금액', '결제수단', '메모']
    
    # Add Flow_Filter placeholder (will add formula)
    clean_df['Flow_Filter'] = 1  # Default to 1 (normal transaction)
    
    # Write headers
    for c_idx, col_name in enumerate(clean_df.columns, 1):
        cell = ws_data.cell(row=1, column=c_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for r_idx, row in enumerate(clean_df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws_data.cell(row=r_idx, column=c_idx)
            if c_idx == 1:  # Date
                if isinstance(value, pd.Timestamp):
                    cell.value = value.to_pydatetime()
                else:
                    cell.value = value
                cell.number_format = 'YYYY-MM-DD'
            elif c_idx == 7:  # Amount
                cell.value = float(value) if pd.notna(value) else 0
                cell.number_format = '#,##0'
            elif c_idx == 10:  # Flow_Filter column
                # Formula to determine if transaction is special (0) or normal (1)
                # Logic: IF main_category is '이동' AND sub_category is '이체', OR memo contains keywords -> 0, else 1
                formula = f'=IF(OR(AND(D{r_idx}="이동",E{r_idx}="이체"),OR(ISNUMBER(SEARCH("투자",I{r_idx})),ISNUMBER(SEARCH("이체",I{r_idx})),ISNUMBER(SEARCH("충전",I{r_idx})))),0,1)'
                cell.value = formula
            else:
                cell.value = value if pd.notna(value) else ""
            cell.font = normal_font
    
    # Convert to Table
    last_row = len(clean_df) + 1
    table_ref = f"A1:J{last_row}"
    table = Table(displayName="T_RawData", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws_data.add_table(table)
    
    # Freeze panes
    ws_data.freeze_panes = "A2"
    
    # Auto-fit columns
    for col in ws_data.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_data.column_dimensions[column].width = min(max_length + 2, 40)
    
    # ==================================================
    # SHEET 2: 📈 Pivot Analysis
    # ==================================================
    ws_pivot = wb.create_sheet("📈 Pivot Analysis", 1)
    ws_pivot.sheet_view.showGridLines = False
    
    # Title
    ws_pivot['B2'] = "피벗 분석 (특수 거래 제외)"
    ws_pivot['B2'].font = Font(size=16, bold=True, color=HEADER_COLOR, name="맑은 고딕")
    ws_pivot.merge_cells('B2:F2')
    
    # Monthly Pivot (excluding special transactions)
    ws_pivot['B4'] = "월별 재정 요약"
    ws_pivot['B4'].font = Font(size=13, bold=True, color=ACCENT_COLOR, name="맑은 고딕")
    
    # Filter out special transactions for pivot
    # In real scenario, Flow_Filter would be calculated. We'll simulate manually
    special_keywords = ['투자', '이체', '충전']
    df['is_special'] = df.apply(
        lambda x: (x['main_category'] == '이동' and x['sub_category'] == '이체') or 
                  any(keyword in str(x['memo']) for keyword in special_keywords),
        axis=1
    )
    
    normal_df = df[~df['is_special']].copy()
    
    monthly_pivot = normal_df.groupby([normal_df['date'].dt.to_period('M'), 'type'])['amount'].sum().unstack(fill_value=0)
    if '수입' not in monthly_pivot.columns:
        monthly_pivot['수입'] = 0
    if '지출' not in monthly_pivot.columns:
        monthly_pivot['지출'] = 0
    monthly_pivot['순수익'] = monthly_pivot['수입'] - monthly_pivot['지출']
    monthly_pivot = monthly_pivot.reset_index()
    monthly_pivot['월'] = monthly_pivot['date'].astype(str)
    monthly_pivot = monthly_pivot[['월', '수입', '지출', '순수익']]
    
    # Write monthly pivot
    for r_idx, row in enumerate(dataframe_to_rows(monthly_pivot, index=False, header=True), 5):
        for c_idx, value in enumerate(row, 2):
            cell = ws_pivot.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 5:  # Header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='center')
                if c_idx > 2:  # Number columns
                    cell.number_format = '#,##0'
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Category Pivot
    expense_df = normal_df[normal_df['type'] == '지출']
    pivot_start_row = 5 + len(monthly_pivot) + 4
    
    ws_pivot[f'B{pivot_start_row}'] = "카테고리별 지출 분석"
    ws_pivot[f'B{pivot_start_row}'].font = Font(size=13, bold=True, color=ACCENT_COLOR, name="맑은 고딕")
    
    if not expense_df.empty:
        cat_pivot = expense_df.groupby(['main_category', 'sub_category'])['amount'].sum().reset_index()
        cat_pivot = cat_pivot.sort_values('amount', ascending=False)
        cat_pivot.columns = ['대분류', '소분류', '총 지출액']
        
        # Write category pivot
        for r_idx, row in enumerate(dataframe_to_rows(cat_pivot, index=False, header=True), pivot_start_row + 1):
            for c_idx, value in enumerate(row, 2):
                cell = ws_pivot.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == pivot_start_row + 1:  # Header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = normal_font
                    if c_idx == 4:  # Amount column
                        cell.number_format = '#,##0'
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
    
    # ==================================================
    # SHEET 3: 📊 Dashboard Summary (Professional Design)
    # ==================================================
    ws_dash = wb.create_sheet("📊 Dashboard Summary", 2)
    ws_dash.sheet_view.showGridLines = False
    ws_dash.page_setup.paperSize = 9  # A4
    ws_dash.page_setup.fitToPage = True
    ws_dash.page_setup.fitToWidth = 1
    ws_dash.page_setup.fitToHeight = 1
    
    # Title
    ws_dash['B2'] = "가계부 종합 대시보드"
    ws_dash['B2'].font = Font(size=22, bold=True, color=HEADER_COLOR, name="맑은 고딕")
    ws_dash.merge_cells('B2:H2')
    ws_dash['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws_dash.row_dimensions[2].height = 30
    
    # Subtitle
    ws_dash['B3'] = f"분석 기간: {normal_df['date'].min().strftime('%Y-%m-%d')} ~ {normal_df['date'].max().strftime('%Y-%m-%d')}"
    ws_dash['B3'].font = Font(size=11, italic=True, color="7F7F7F", name="맑은 고딕")
    ws_dash.merge_cells('B3:H3')
    ws_dash['B3'].alignment = Alignment(horizontal='center')
    
    # Calculate KPIs (excluding special transactions)
    total_income = normal_df[normal_df['type'] == '수입']['amount'].sum()
    total_expense = normal_df[normal_df['type'] == '지출']['amount'].sum()
    net_income = total_income - total_expense
    
    # KPI Cards with professional design
    kpi_row = 5
    ws_dash.row_dimensions[kpi_row].height = 25
    ws_dash.row_dimensions[kpi_row + 1].height = 35
    
    kpi_cards = [
        ("💰 정확한 총 순수익", net_income, "B", SUCCESS_COLOR if net_income >= 0 else DANGER_COLOR),
        ("💸 총 지출액", total_expense, "E", DANGER_COLOR),
        ("📊 거래 건수", len(normal_df), "H", ACCENT_COLOR)
    ]
    
    for label, value, col, color in kpi_cards:
        # Label
        ws_dash[f'{col}{kpi_row}'] = label
        ws_dash[f'{col}{kpi_row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws_dash[f'{col}{kpi_row}'].font = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")
        ws_dash[f'{col}{kpi_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dash.merge_cells(f'{col}{kpi_row}:{chr(ord(col)+1)}{kpi_row}')
        
        # Value
        ws_dash[f'{col}{kpi_row+1}'] = f"{value:,.0f}{'원' if isinstance(value, (int, float)) else ''}"
        ws_dash[f'{col}{kpi_row+1}'].fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
        ws_dash[f'{col}{kpi_row+1}'].font = Font(bold=True, size=18, color=color, name="맑은 고딕")
        ws_dash[f'{col}{kpi_row+1}'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dash.merge_cells(f'{col}{kpi_row+1}:{chr(ord(col)+1)}{kpi_row+1}')
        
        # Border
        for r in [kpi_row, kpi_row+1]:
            for c in [col, chr(ord(col)+1)]:
                ws_dash[f'{c}{r}'].border = Border(
                    left=Side(style='medium'), right=Side(style='medium'),
                    top=Side(style='medium'), bottom=Side(style='medium')
                )
    
    # Chart Data Area (hidden)
    chart_data_row = 10
    ws_dash[f'K{chart_data_row}'] = "월"
    ws_dash[f'L{chart_data_row}'] = "수입"
    ws_dash[f'M{chart_data_row}'] = "지출"
    
    for i, row_data in enumerate(monthly_pivot.itertuples(index=False), chart_data_row + 1):
        ws_dash[f'K{i}'] = row_data.월
        ws_dash[f'L{i}'] = row_data.수입
        ws_dash[f'M{i}'] = row_data.지출
    
    # Line Chart with professional styling
    chart1 = LineChart()
    chart1.title = "월별 수입/지출 추이 (특수 거래 제외)"
    chart1.style = 12
    chart1.y_axis.title = "금액 (원)"
    chart1.x_axis.title = "월"
    chart1.height = 12
    chart1.width = 20
    
    data = Reference(ws_dash, min_col=12, min_row=chart_data_row, max_row=chart_data_row + len(monthly_pivot), max_col=13)
    cats = Reference(ws_dash, min_col=11, min_row=chart_data_row + 1, max_row=chart_data_row + len(monthly_pivot))
    
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    
    # Customize series colors
    s1 = chart1.series[0]  # 수입
    s1.graphicalProperties.line.solidFill = SUCCESS_COLOR
    s1.graphicalProperties.line.width = 30000
    
    s2 = chart1.series[1]  # 지출
    s2.graphicalProperties.line.solidFill = DANGER_COLOR
    s2.graphicalProperties.line.width = 30000
    
    ws_dash.add_chart(chart1, "B9")
    
    # Category Pie Chart
    if not expense_df.empty:
        cat_summary = expense_df.groupby('main_category')['amount'].sum().reset_index()
        cat_summary = cat_summary.sort_values('amount', ascending=False)
        
        cat_data_row = chart_data_row
        ws_dash[f'O{cat_data_row}'] = "카테고리"
        ws_dash[f'P{cat_data_row}'] = "금액"
        
        for i, row_data in enumerate(cat_summary.itertuples(index=False), cat_data_row + 1):
            ws_dash[f'O{i}'] = row_data.main_category
            ws_dash[f'P{i}'] = row_data.amount
        
        pie = PieChart()
        pie.title = "카테고리별 지출 비중 (2D)"
        pie.height = 12
        pie.width = 14
        
        labels = Reference(ws_dash, min_col=15, min_row=cat_data_row + 1, max_row=cat_data_row + len(cat_summary))
        data = Reference(ws_dash, min_col=16, min_row=cat_data_row, max_row=cat_data_row + len(cat_summary))
        
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        
        ws_dash.add_chart(pie, "B27")
    
    # ==================================================
    # SHEET 4: 💰 Investment & Transfer Tracking
    # ==================================================
    ws_invest = wb.create_sheet("💰 Investment Tracking", 3)
    ws_invest.sheet_view.showGridLines = False
    
    # Title
    ws_invest['B2'] = "투자 및 이동 내역 추적"
    ws_invest['B2'].font = Font(size=16, bold=True, color=HEADER_COLOR, name="맑은 고딕")
    ws_invest.merge_cells('B2:F2')
    
    # Special transactions only
    special_df = df[df['is_special']].copy()
    
    ws_invest['B4'] = f"총 특수 거래 건수: {len(special_df)}건"
    ws_invest['B4'].font = Font(size=12, bold=True, color=ACCENT_COLOR, name="맑은 고딕")
    
    if not special_df.empty:
        # Monthly investment summary
        monthly_invest = special_df.groupby(special_df['date'].dt.to_period('M'))['amount'].sum().reset_index()
        monthly_invest['월'] = monthly_invest['date'].astype(str)
        monthly_invest = monthly_invest[['월', 'amount']]
        monthly_invest.columns = ['월', '총 이동/투자액']
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(monthly_invest, index=False, header=True), 6):
            for c_idx, value in enumerate(row, 2):
                cell = ws_invest.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 6:  # Header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = normal_font
                    cell.alignment = Alignment(horizontal='center')
                    if c_idx == 3:
                        cell.number_format = '#,##0'
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
        
        # Bar Chart for monthly investment
        bar = BarChart()
        bar.type = "col"
        bar.title = "월별 투자/이동 규모"
        bar.y_axis.title = "금액 (원)"
        bar.x_axis.title = "월"
        bar.height = 10
        bar.width = 16
        
        labels = Reference(ws_invest, min_col=2, min_row=7, max_row=6 + len(monthly_invest))
        data = Reference(ws_invest, min_col=3, min_row=6, max_row=6 + len(monthly_invest))
        
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(labels)
        
        ws_invest.add_chart(bar, "B15")
    
    # Auto-fit columns for all sheets
    for ws in [ws_pivot, ws_dash, ws_invest]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 40)
    
    # Save
    print(f"💾 Saving to: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    print("✅ Advanced Excel Report created successfully!")
    print(f"\n📊 Summary:")
    print(f"   - 시트 1: 📋 Clean Raw Data (Flow_Filter 수식 포함, {len(clean_df)} rows)")
    print(f"   - 시트 2: 📈 Pivot Analysis (특수 거래 제외)")
    print(f"   - 시트 3: 📊 Dashboard Summary (Professional Design, A4 layout)")
    print(f"   - 시트 4: 💰 Investment Tracking ({len(special_df)} special transactions)")
    print(f"\n✅ File saved: {OUTPUT_FILE}")

if __name__ == "__main__":
    create_advanced_excel_report()
