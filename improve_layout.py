import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load workbook
print("Loading workbook...")
wb = load_workbook(
    r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_수식연결_가계부엔진_최종.xlsx'
)

ws = wb['📊 Dashboard Summary']

# Colors
HEADER_COLOR = "2C3E50"
ACCENT_COLOR = "3498DB"
SUCCESS_COLOR = "27AE60"
DANGER_COLOR = "E74C3C"
WARNING_COLOR = "F39C12"
LIGHT_BG = "ECF0F1"

print("Reorganizing layout...")

# ================================================================
# LAYOUT REORGANIZATION
# ================================================================

# Move charts to better positions
old_charts = list(ws._charts)
ws._charts.clear()

# Repositen charts
chart_count = 0
for chart in old_charts:
    if chart_count == 0:  # First chart (likely monthly trend)
        chart.anchor = "K5"
        chart.height = 15
        chart.width = 22
    elif chart_count == 1:  # Second chart (likely pie)
        chart.anchor = "K22"
        chart.height = 14
        chart.width = 18
    
    ws.add_chart(chart)
    chart_count += 1

print(f"Repositioned {chart_count} charts")

# Adjust row heights for better spacing
ws.row_dimensions[2].height = 40  # Title
ws.row_dimensions[3].height = 18  # Subtitle
ws.row_dimensions[5].height = 25  # Section headers

# Adjust column widths for better readability
column_widths = {
    'A': 3,   # Margin
    'B': 12,  # Labels
    'C': 15,  # Values 1
    'D': 12,  # Values 2
    'E': 12,  # Values 3
    'F': 15,  # Category names
    'G': 15,  # Amounts
    'H': 10,  # Percentages
    'I': 15,  # Monthly avg
    'J': 3,   # Margin
    'K': 12,  # Chart area
    'L': 15,
    'M': 15,
    'N': 15,
    'O': 15,
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Add visual separators (merge cells for section breaks)
# Add spacing row between sections
for spacing_row in [4, 12, 25, 38]:
    ws.row_dimensions[spacing_row].height = 8

# Enhance title styling
title_cell = ws['B2']
title_cell.font = Font(size=24, bold=True, color=HEADER_COLOR, name="맑은 고딕")

# Add subtle background to main sections
# KPI section background
for row in range(5, 11):
    for col in range(2, 5):
        cell = ws.cell(row=row, column=col)
        if cell.fill.start_color.index == '00000000':  # No fill
            cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

# Adjust table borders to be more prominent
# Top 10 table - make borders thicker
for row in range(13, 24):  # Top 10 table approximate range
    for col in range(2, 7):
        cell = ws.cell(row=row, column=col)
        if cell.value:
            cell.border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )

# Monthly data table - similar treatment
for row in range(6, 20):  # Monthly table approximate range
    for col in range(9, 16):
        cell = ws.cell(row=row, column=col)
        if cell.value:
            cell.border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )

# Pattern table borders
for row in range(26, 32):  # Pattern table approximate range
    for col in range(2, 6):
        cell = ws.cell(row=row, column=col)
        if cell.value:
            cell.border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )

# Freeze panes at row 4 for better scrolling
ws.freeze_panes = 'A4'

# Set print area for better printing
ws.print_area = 'A1:P50'

# Page setup for better printing
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToHeight = 1
ws.page_setup.fitToWidth = 1

# Add gridlines for print
ws.print_options.gridLines = False
ws.print_options.gridLinesSet = True

print("Layout reorganization complete!")

# Save
output_file = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_수식연결_가계부엔진_레이아웃개선.xlsx'
wb.save(output_file)

print("\n" + "=" * 70)
print("✅ Dashboard 레이아웃 개선 완료!")
print("=" * 70)
print(f"\n📁 파일: {output_file}")
print(f"\n✨ 개선사항:")
print("  - 차트 위치 최적화 (상단/하단 배치)")
print("  - 차트 크기 확대 (가시성 향상)")
print("  - 컬럼 너비 조정 (가독성 향상)")
print("  - 테이블 간 간격 추가")
print("  - 테두리 스타일 개선")
print("  - 인쇄 레이아웃 최적화")
print("  - 화면 고정 (스크롤 편의성)")
print("\n" + "=" * 70)
