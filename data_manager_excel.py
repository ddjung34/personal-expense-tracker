import pandas as pd
import openpyxl
from datetime import datetime
import os
from copy import copy
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.views import SheetView, Selection

# Configuration
DATA_FILE = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_수식연결_가계부엔진_최종.xlsx'
SHEET_NAME = '📋 T_RawData'

def load_data():
    """
    Loads raw data from the Excel file with SMART HEADER DETECTION.
    Scans looking for '날짜', '구분', '금액' to find the correct header row.
    Returns a pandas DataFrame with an 'Active' flag.
    """
    if not os.path.exists(DATA_FILE):
        print(f"File not found: {DATA_FILE}")
        return pd.DataFrame() 
        
    try:
        # 1. Read first few rows to find the header
        preview_df = pd.read_excel(DATA_FILE, sheet_name=SHEET_NAME, header=None, nrows=10)
        
        header_row_index = None
        required_cols = {'날짜', '구분', '금액'}
        
        for idx, row in preview_df.iterrows():
            row_values = set(row.astype(str).tolist())
            if required_cols.issubset(row_values):
                header_row_index = idx
                break
        
        if header_row_index is None:
            print("Warning: Could not detect header row automatically. Trying default Header=1 (Row 2).")
            header_row_index = 1
            
        # 2. Re-load with correct header
        df = pd.read_excel(DATA_FILE, sheet_name=SHEET_NAME, header=header_row_index)
        
        if '날짜' in df.columns:
            df['날짜'] = pd.to_datetime(df['날짜'], errors='coerce')
            
        def parse_time(val):
            if pd.isna(val): return None
            if isinstance(val, str):
                try: return pd.to_datetime(val, format='%H:%M:%S').time()
                except:
                    try: return pd.to_datetime(val).time()
                    except: return None
            if hasattr(val, 'time'): return val.time()
            if type(val).__name__ == 'time': return val
            return None
            
        if '시간' in df.columns:
            df['시간'] = df['시간'].apply(parse_time)
            
        if '금액' in df.columns:
            df['금액'] = pd.to_numeric(df['금액'], errors='coerce').fillna(0)
            
        if 'Flow_Filter' in df.columns:
            def is_active(val):
                try: return str(val).strip().split('.')[0] == '1'
                except: return False
            df['Is_Active'] = df['Flow_Filter'].apply(is_active)
        else:
            df['Is_Active'] = True
            
        return df
    except Exception as e:
        print(f"Error loading data: {e}")
        return pd.DataFrame()

def save_data(df):
    """
    Saves the data back to Excel with openpyxl, preserving styling.
    """
    try:
        if os.path.exists(DATA_FILE):
             wb = openpyxl.load_workbook(DATA_FILE)
        else:
             wb = openpyxl.Workbook()
             
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.create_sheet(SHEET_NAME)
            
        # Clear data from Row 3 downwards (Header is Row 2 usually)
        # We assume header is at Row 2 based on previous logic finding.
        start_row = 3
        max_row = ws.max_row
        if max_row >= start_row:
             ws.delete_rows(start_row, max_row - start_row + 1)
             
        save_df = df.copy()
        
        # Sync Flow_Filter
        if 'Is_Active' in save_df.columns:
            save_df['Flow_Filter'] = save_df['Is_Active'].apply(lambda x: 1 if x else 0)
            save_df = save_df.drop(columns=['Is_Active'])
            
        # Column Order
        target_order = ['날짜', '시간', '구분', '대분류', '소분류', '내용', '금액', '결제수단', '메모', 'Flow_Filter']
        final_columns = [c for c in target_order if c in save_df.columns]
        save_df = save_df[final_columns]
        
        # Styles
        border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        fill_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        fill_white = PatternFill(fill_type=None)
        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
            
        # Write
        current_row = start_row # 3
        for row_data in save_df.itertuples(index=False, name=None):
            ws.append(row_data) # This might append at max_row + 1. 
            # If we deleted rows, max_row should be 2. So append starts at 3. Correct.
            
            is_blue_row = (current_row % 2 == 0) # Even rows blue
            current_fill = fill_blue if is_blue_row else fill_white
            
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = Font(name='맑은 고딕', size=11)
                cell.border = border_all
                cell.fill = current_fill
                
                col_name = final_columns[col_idx-1]
                if col_name == '날짜':
                    cell.number_format = 'yyyy-mm-dd'
                    cell.alignment = align_center
                elif col_name == '금액':
                    cell.number_format = '#,##0' 
                    cell.alignment = align_right
                elif col_name in ['시간', '구분', '대분류', '소분류', '결제수단', 'Flow_Filter']:
                    cell.alignment = align_center
                else: 
                    cell.alignment = align_left
            
            current_row += 1
        
        # Freeze and Reset View
        ws.freeze_panes = 'A2'
        ws.sheet_view.topLeftCell = 'A1'
        ws.sheet_view.selection = [Selection(activeCell='A1', sqref='A1')]
        
        wb.save(DATA_FILE)
        return True
        
    except Exception as e:
        print(f"Error saving data: {e}")
        return False

def get_kpi_metrics(df):
    if df.empty: return {'income':0, 'expense':0, 'net':0, 'other':0}
    active_df = df[df['Is_Active'] == True]
    income = active_df[active_df['구분'] == '수입']['금액'].sum()
    expense = active_df[active_df['구분'] == '지출']['금액'].sum()
    other_df = df[(df['Is_Active'] == False) & (df['Flow_Filter'] == 1)]
    other = other_df['금액'].sum() if not other_df.empty else 0
    return {'income': income, 'expense': expense, 'net': income + expense, 'other': other}
