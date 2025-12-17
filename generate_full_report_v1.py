"""
Expense Report Generator v1 (Phase 3)
사용자 요청: "프로그램" 방식의 가계부 (연간 대시보드 + 월별 상세 시트)

Architecture:
1. Template: `Dashboard_v40_perfect.xlsx` (연간 대시보드 포함된 완벽한 파일)
2. Data Source: `2013_수식연결_가계부엔진.xlsx` (데이터 신뢰성 높음)
3. Processing:
   - 데이터 로드 및 월별(YYYY-MM) 그룹화
   - 월별 시트 자동 생성 (예: '2024-12', '2025-01')
   - 각 시트에 해당 월의 내역 기입
4. Output: `Expense_Report_Full_v1.xlsx`

이 프로그램은 "버튼 하나로" 전체 리포트를 생성하는 구조의 프로토타입입니다.
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

print("=" * 70)
print("가계부 리포트 생성 프로그램 v1 (연간 + 월별)")
print("=" * 70)

# 1. 설정
TEMPLATE_FILE = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_Dashboard_v40_perfect.xlsx'
DATA_FILE = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251213_수식연결_가계부엔진.xlsx'
OUTPUT_FILE = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\Expense_Report_Full_v1.xlsx'

# 2. 데이터 로드 및 그룹화
print("1. 데이터 로드 및 월별 분류 중...")
wb_data = load_workbook(DATA_FILE, data_only=True)
ws_raw = wb_data['📋 T_RawData']

monthly_data = {} # {'2024-12': [row_data, ...], ...}

# Raw Data Header (Row 2) 제외하고 3행부터 읽기
headers = ["날짜", "시간", "구분", "대분류", "소분류", "내용", "금액", "결제수단", "메모"]
raw_header_indices = [0, 1, 2, 3, 4, 5, 6, 7, 8] # A~I

for row in ws_raw.iter_rows(min_row=3, values_only=True):
    date_val = row[0]
    if isinstance(date_val, datetime):
        month_key = date_val.strftime("%Y-%m")
        if month_key not in monthly_data:
            monthly_data[month_key] = []
        
        # 필요한 컬럼만 추출
        row_items = [row[i] for i in raw_header_indices]
        monthly_data[month_key].append(row_items)
    elif date_val is not None:
        # 날짜 형식이 아닌 경우 (텍스트 등) - 로깅만 하고 건너뜀
        # print(f"Skip invalid date row: {date_val}")
        pass

wb_data.close()

sorted_months = sorted(monthly_data.keys())
print(f"   - 분류된 월: {sorted_months}")


# 3. 리포트 생성 (Template 복사)
print(f"2. 리포트 생성 중 (Template: {TEMPLATE_FILE})...")
wb_report = load_workbook(TEMPLATE_FILE)

# 연간 대시보드 시트 이름 변경 (명확하게)
if '📊 Dashboard_v13' in wb_report.sheetnames:
    ws_main = wb_report['📊 Dashboard_v13']
    ws_main.title = "🏆 연간 대시보드"
    # 탭 색상 변경 (Gold)
    ws_main.sheet_properties.tabColor = "FFD700"
    
# 4. 월별 시트 생성
print("3. 월별 시트 생성 및 데이터 기입...")

# 스타일 정의
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")
border_style = Side(style='thin')
thin_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

for month in sorted_months:
    print(f"   - 시트 생성: {month}")
    # 기존 시트가 있다면 삭제 (충돌/병합 방지)
    if month in wb_report.sheetnames:
        del wb_report[month]
    
    ws_month = wb_report.create_sheet(title=month)
        
    # 헤더 기입
    for col_idx, header in enumerate(headers, 1):
        cell = ws_month.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # 데이터 기입
    transactions = monthly_data[month]
    # 날짜순 정렬
    transactions.sort(key=lambda x: x[0] if isinstance(x[0], datetime) else datetime.min)
    
    for r_idx, row_data in enumerate(transactions, 2):
        for c_idx, val in enumerate(row_data, 1):
            try:
                cell = ws_month.cell(row=r_idx, column=c_idx, value=val)
                # Style application...
                cell.border = thin_border
                
                # 포맷팅
                if c_idx == 1: # 날짜
                    cell.number_format = 'yyyy-mm-dd'
                    cell.alignment = center_align
                elif c_idx == 2: # 시간
                    cell.number_format = 'hh:mm:ss'
                    cell.alignment = center_align
                elif c_idx == 7: # 금액
                    cell.number_format = '#,##0'
            except Exception as e:
                print(f"Error at Sheet {month}, Row {r_idx}, Col {c_idx}, Val: {val} ({type(val)})")
                print(f"Error details: {e}")
                # Don't exit, try next

                
    # 컬럼 너비 자동 조정 (대략)
    ws_month.column_dimensions['A'].width = 12 # 날짜
    ws_month.column_dimensions['B'].width = 10 # 시간
    ws_month.column_dimensions['C'].width = 6  # 구분
    ws_month.column_dimensions['D'].width = 10 # 대분류
    ws_month.column_dimensions['E'].width = 10 # 소분류
    ws_month.column_dimensions['F'].width = 25 # 내용
    ws_month.column_dimensions['G'].width = 12 # 금액
    ws_month.column_dimensions['H'].width = 15 # 결제수단
    ws_month.column_dimensions['I'].width = 20 # 메모

# 5. 저장
wb_report.save(OUTPUT_FILE)
print(f"✅ 리포트 생성 완료: {OUTPUT_FILE}")
print("   - 시트 구성을 확인하세요: [연간 대시보드] + [월별 시트들]")
