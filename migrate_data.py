import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd

SOURCE_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-07~2025-12-07.xlsx"
OUTPUT_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-07~2025-12-07_v3.xlsx"

# 1. Define Target Headers (Analysis Result)
HEADERS = [
    "date", "time", "type", "main_category", "sub_category", 
    "amount", "payment_method", "merchant", "memo"
]

# 2. Define Column Mapping (Source -> Target Name)
# Source Headers based on inspection: 
# ['날짜', '시간', '수입/지출', '대분류', '소분류', '내용', '금액', '통화', '결제수단', '메모']
MAPPING = {
    '날짜': 'date',
    '시간': 'time',
    '수입/지출': 'type',
    '대분류': 'main_category',
    '소분류': 'sub_category',
    '금액': 'amount',
    '결제수단': 'payment_method',
    '내용': 'merchant',
    '메모': 'memo'
}

def migrate_data():
    print(f"Loading Source Data from: {SOURCE_FILE}")
    try:
        # Read source data using pandas for easy manipulation
        df = pd.read_excel(SOURCE_FILE, sheet_name="가계부 내역", engine='openpyxl')
        print(f"Read {len(df)} rows from '가계부 내역'.")
        
        # Load workbook to modify
        wb = openpyxl.load_workbook(SOURCE_FILE)
        
    except Exception as e:
        print(f"Error loading file: {e}")
        return

    # Check/Create DB_Raw
    if "DB_Raw" in wb.sheetnames:
        print("Removing existing DB_Raw to clean start...")
        del wb["DB_Raw"]
        
    print("Creating new DB_Raw sheet...")
    ws = wb.create_sheet("DB_Raw", 0) # Index 0

    # Styling for Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    # Write Headers
    for col_num, header_title in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Write Data
    print("Migrating rows...")
    for idx, row in df.iterrows():
        # Map source row to target schema
        target_row = []
        for header in HEADERS:
            # Find which source column maps to this target header
            source_col = None
            for s_col, t_col in MAPPING.items():
                if t_col == header:
                    source_col = s_col
                    break
            
            value = row.get(source_col) if source_col else None
            
            # Data Cleaning
            if header == 'amount':
                # Convert negative amounts to positive if needed, or keep as is?
                # Usually DB stores absolute values and 'type' defines sign.
                # But let's check source. Source has -8000 for Expense.
                # Let's keep it consistent: Expense should probably be positive number in DB if Type is Expense,
                # BUT user might prefer seeing - for math.
                # Standard practice: Store positive, Type determines math.
                # Let's converting to absolute for cleaner DB if type is known.
                pass 
                
            target_row.append(value)
            
        # Write to sheet (row idx + 2 because: 1-based index, +1 for header)
        for col_idx, val in enumerate(target_row, 1):
            ws.cell(row=idx+2, column=col_idx, value=val)

    print(f"Saving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print("Migration Complete!")

if __name__ == "__main__":
    migrate_data()
