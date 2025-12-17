
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

file_path = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_Dashboard_v25_direct_let.xlsx'
wb = load_workbook(file_path, data_only=True)
ws = wb['📋 T_RawData']

print("=== Scanning Columns J, K, L (Rows 1-50) ===")
found_j = False
found_k = False
found_l = False

for r in range(1, 51):
    j_val = ws.cell(row=r, column=10).value # J
    k_val = ws.cell(row=r, column=11).value # K
    l_val = ws.cell(row=r, column=12).value # L
    
    if j_val is not None and not found_j:
        print(f"J{r}: {j_val}")
        # found_j = True # Keep scanning to see patterns
    if k_val is not None and not found_k:
        print(f"K{r}: {k_val}")
    if l_val is not None and not found_l:
        print(f"L{r}: {l_val}")
