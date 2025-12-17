import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

# Using v3 as input, will save as v4_final
INPUT_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-07~2025-12-07_v3.xlsx"
OUTPUT_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-07~2025-12-07_v4_Auto.xlsx"

def apply_formulas():
    print(f"Loading {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE)
    
    if "뱅샐현황" not in wb.sheetnames:
        print("Error: '뱅샐현황' sheet not found.")
        return

    ws_report = wb["뱅샐현황"]
    
    # Configuration based on analysis
    # Target: Column E (2024-12)
    # Category Source: Column B
    # DB Sheet: DB_Raw
    # DB Columns: F (Amount), D (Main Category), A (Date)
    
    target_col_idx = 5 # Column E
    category_col_idx = 2 # Column B
    header_row = 11 # Data starts from 12
    
    print("Applying formulas to '뱅샐현황' (Column E - 2024-12)...")
    
    count = 0
    for row in range(header_row + 1, ws_report.max_row + 1):
        # GET Category Name
        cat_cell = ws_report.cell(row=row, column=category_col_idx)
        cat_val = cat_cell.value
        
        # Define Target Cell
        target_cell = ws_report.cell(row=row, column=target_col_idx)
        
        # Skip if category is empty or weird
        if not cat_val or not isinstance(cat_val, str):
            continue
            
        # Ref for formula
        cat_ref = f"{get_column_letter(category_col_idx)}{row}" # e.g. B12
        
        # Build Formula
        # SUMIFS(Amount, CategoryCol, CategoryVal, DateCol, ">=Start", DateCol, "<=End")
        formula = f'=SUMIFS(DB_Raw!$F:$F, DB_Raw!$D:$D, {cat_ref}, DB_Raw!$A:$A, ">=2024-12-01", DB_Raw!$A:$A, "<=2024-12-31")'
        
        # Handle Merged Cells
        if isinstance(target_cell, MergedCell):
            continue
            
        try:
            target_cell.value = formula
            count += 1
        except Exception as e:
            print(f"Error at Row {row}: {e}")

    print(f"Applied {count} formulas.")
    print(f"Saving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print("Done! v4_Auto created.")

if __name__ == "__main__":
    apply_formulas()
