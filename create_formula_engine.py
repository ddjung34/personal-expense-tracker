import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, PieChart, BarChart, Reference
from datetime import datetime

INPUT_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-10~2025-12-10.xlsx"
OUTPUT_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\{date}_수식연결_가계부엔진.xlsx".format(
    date=datetime.now().strftime("%Y%m%d")
)

def create_formula_linked_report():
    print(f"Loading data from: {INPUT_FILE}")
    
    # Read data
    df = pd.read_excel(INPUT_FILE, sheet_name='Rawdata', engine='openpyxl')
    
    # Data cleaning
    df['date'] = pd.to_datetime(df['날짜'], errors='coerce')
    df['time'] = df['시간'].astype(str) if '시간' in df.columns else ''
    df['type'] = df['타입']
    df['main_category'] = df['대분류'].fillna('')
    df['sub_category'] = df['소분류'].fillna('')
    df['merchant'] = df['내용'].fillna('')
    df['amount'] = pd.to_numeric(df['금액'], errors='coerce')
    df['payment_method'] = df['결제수단'].fillna('')
    df['memo'] = df['메모'].fillna('')
    
    df = df.dropna(subset=['date', 'amount', 'type'])
    df = df.sort_values(['date', 'time'])
    
    # Add row tracking for T_RawData reference (row 3 is first data row)
    df['raw_row'] = range(3, 3 + len(df))
    
    print(f"Loaded {len(df)} transactions")
    
    # Get months
    df['year_month'] = df['date'].dt.to_period('M')
    months = sorted(df['year_month'].unique())
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Unified Design Colors
    HEADER_COLOR = "2C3E50"      # Dark Gray
    ACCENT_COLOR = "3498DB"       # Blue
    SUCCESS_COLOR = "27AE60"      # Green
    DANGER_COLOR = "E74C3C"       # Red
    WARNING_COLOR = "F39C12"      # Orange
    LIGHT_BG = "ECF0F1"          # Light Gray
    TABLE_HEADER = "34495E"       # Table Header Dark
    
    # ==================================================
    # SHEET 1: 📋 T_RawData (Data Engine)
    # ==================================================
    ws_raw = wb.create_sheet("📋 T_RawData", 0)
    ws_raw.sheet_view.showGridLines = False
    
    # Title
    ws_raw['A1'] = "가계부 데이터 엔진 (T_RawData)"
    ws_raw['A1'].font = Font(size=14, bold=True, color=HEADER_COLOR, name="맑은 고딕")
    ws_raw.merge_cells('A1:K1')
    ws_raw['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_raw.row_dimensions[1].height = 25
    
    # Prepare data with Flow_Filter
    columns = ['날짜', '시간', '구분', '대분류', '소분류', '내용', '금액', '결제수단', '메모', 'Flow_Filter']
    
    # Write headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws_raw.cell(row=2, column=col_idx, value=col_name)
        cell.fill = PatternFill(start_color=TABLE_HEADER, end_color=TABLE_HEADER, fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="맑은 고딕")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    # Write data with ACTUAL date values (not strings)
    for idx, row in enumerate(df.itertuples(), 3):
        # Store as actual date object
        ws_raw.cell(row=idx, column=1, value=row.date)
        ws_raw.cell(row=idx, column=1).number_format = 'YYYY-MM-DD'
        
        ws_raw.cell(row=idx, column=2, value=str(row.time) if row.time else '')
        ws_raw.cell(row=idx, column=3, value=row.type)
        ws_raw.cell(row=idx, column=4, value=row.main_category)
        ws_raw.cell(row=idx, column=5, value=row.sub_category)
        ws_raw.cell(row=idx, column=6, value=row.merchant)
        ws_raw.cell(row=idx, column=7, value=float(row.amount))
        ws_raw.cell(row=idx, column=7).number_format = '#,##0'
        ws_raw.cell(row=idx, column=8, value=row.payment_method)
        ws_raw.cell(row=idx, column=9, value=row.memo)
        
        # Flow_Filter Formula
        # Logic: IF main='이동' AND (sub='이체' OR sub='투자') OR memo contains keywords -> 0, else 1
        formula = f'=IF(OR(AND(D{idx}="이동",OR(E{idx}="이체",E{idx}="투자")),OR(ISNUMBER(SEARCH("투자",I{idx})),ISNUMBER(SEARCH("이체",I{idx})),ISNUMBER(SEARCH("충전",I{idx})))),0,1)'
        ws_raw.cell(row=idx, column=10, value=formula)
        
        # Add internal transfer/investment indicator in memo if applicable
        if (row.main_category == '이동' and row.sub_category == '이체'):
            current_memo = row.memo if row.memo else ''
            if '내부이체' not in current_memo:
                ws_raw.cell(row=idx, column=9, value=f'[내부이체] {current_memo}'.strip())
        elif (row.main_category == '이동' and row.sub_category == '투자'):
            current_memo = row.memo if row.memo else ''
            if '내부거래' not in current_memo:
                ws_raw.cell(row=idx, column=9, value=f'[내부거래-투자] {current_memo}'.strip())
        
        for col_idx in range(1, 11):
            ws_raw.cell(row=idx, column=col_idx).font = Font(name="맑은 고딕", size=9)
            ws_raw.cell(row=idx, column=col_idx).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    last_data_row = len(df) + 2
    
    # Convert to Table
    table_ref = f"A2:J{last_data_row}"
    table = Table(displayName="T_RawData", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws_raw.add_table(table)
    
    # Auto-fit columns
    for col_idx, width in enumerate([12, 10, 8, 12, 12, 20, 15, 15, 25, 10], 1):
        ws_raw.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws_raw.freeze_panes = "A3"
    
    # ==================================================
    # SHEET 2-N: Monthly Detail Sheets (Formula-Linked)
    # ==================================================
    
    for month_period in months:
        month_str = str(month_period)
        
        ws_month = wb.create_sheet(f"{month_str}", len(wb.sheetnames))
        ws_month.sheet_view.showGridLines = False
        
        # Title
        ws_month['B2'] = f"{month_str} 거래 내역"
        ws_month['B2'].font = Font(size=16, bold=True, color=HEADER_COLOR, name="맑은 고딕")
        ws_month.merge_cells('B2:H2')
        ws_month['B2'].alignment = Alignment(horizontal='center', vertical='center')
        ws_month.row_dimensions[2].height = 30
        
        ws_month['B3'] = "(T_RawData 수식 연결 - Flow_Filter=1만 집계)"
        ws_month['B3'].font = Font(size=9, italic=True, color="7F7F7F", name="맑은 고딕")
        ws_month.merge_cells('B3:H3')
        ws_month['B3'].alignment = Alignment(horizontal='center')
        
        # Calculate month start/end dates
        month_date = month_period.to_timestamp()
        year = month_date.year
        month = month_date.month
        
        # Add Summary Section at TOP with FORMULAS
        summary_row = 5
        
        # Summary title
        ws_month[f'B{summary_row}'] = "월별 요약 (Flow_Filter=1만 집계)"
        ws_month[f'B{summary_row}'].font = Font(size=13, bold=True, color=ACCENT_COLOR, name="맑은 고딕")
        ws_month.merge_cells(f'B{summary_row}:D{summary_row}')
        
        # Income summary - SUMIFS formula
        ws_month[f'B{summary_row+2}'] = "💰 수입"
        ws_month[f'B{summary_row+2}'].fill = PatternFill(start_color=SUCCESS_COLOR, end_color=SUCCESS_COLOR, fill_type="solid")
        ws_month[f'B{summary_row+2}'].font = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")
        ws_month[f'B{summary_row+2}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # SUMIFS: Sum amount where type="수입", Flow_Filter=1, and month/year match
        income_formula = f'=SUMIFS(\'📋 T_RawData\'!G:G,\'📋 T_RawData\'!C:C,"수입",\'📋 T_RawData\'!J:J,1,\'📋 T_RawData\'!A:A,">="&DATE({year},{month},1),\'📋 T_RawData\'!A:A,"<"&DATE({year},{month+1 if month<12 else 1},1))'
        if month == 12:
            income_formula = f'=SUMIFS(\'📋 T_RawData\'!G:G,\'📋 T_RawData\'!C:C,"수입",\'📋 T_RawData\'!J:J,1,\'📋 T_RawData\'!A:A,">="&DATE({year},{month},1),\'📋 T_RawData\'!A:A,"<"&DATE({year+1},1,1))'
        
        ws_month[f'C{summary_row+2}'] = income_formula
        ws_month[f'C{summary_row+2}'].number_format = '₩#,##0'
        ws_month[f'C{summary_row+2}'].font = Font(bold=True, size=12, color=SUCCESS_COLOR, name="맑은 고딕")
        ws_month[f'C{summary_row+2}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Expense summary - SUMIFS formula
        ws_month[f'B{summary_row+3}'] = "💸 지출"
        ws_month[f'B{summary_row+3}'].fill = PatternFill(start_color=DANGER_COLOR, end_color=DANGER_COLOR, fill_type="solid")
        ws_month[f'B{summary_row+3}'].font = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")
        ws_month[f'B{summary_row+3}'].alignment = Alignment(horizontal='center', vertical='center')
        
        expense_formula = f'=SUMIFS(\'📋 T_RawData\'!G:G,\'📋 T_RawData\'!C:C,"지출",\'📋 T_RawData\'!J:J,1,\'📋 T_RawData\'!A:A,">="&DATE({year},{month},1),\'📋 T_RawData\'!A:A,"<"&DATE({year},{month+1 if month<12 else 1},1))'
        if month == 12:
            expense_formula = f'=SUMIFS(\'📋 T_RawData\'!G:G,\'📋 T_RawData\'!C:C,"지출",\'📋 T_RawData\'!J:J,1,\'📋 T_RawData\'!A:A,">="&DATE({year},{month},1),\'📋 T_RawData\'!A:A,"<"&DATE({year+1},1,1))'
        
        ws_month[f'C{summary_row+3}'] = expense_formula
        ws_month[f'C{summary_row+3}'].number_format = '₩#,##0'
        ws_month[f'C{summary_row+3}'].font = Font(bold=True, size=12, color=DANGER_COLOR, name="맑은 고딕")
        ws_month[f'C{summary_row+3}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Total - FORMULA (Income - Expense)
        ws_month[f'B{summary_row+4}'] = "✅ 합계"
        ws_month[f'B{summary_row+4}'].fill = PatternFill(start_color=ACCENT_COLOR, end_color=ACCENT_COLOR, fill_type="solid")
        ws_month[f'B{summary_row+4}'].font = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")
        ws_month[f'B{summary_row+4}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws_month[f'C{summary_row+4}'] = f'=SUM(C{summary_row+2},-C{summary_row+3})'
        ws_month[f'C{summary_row+4}'].number_format = '₩#,##0'
        ws_month[f'C{summary_row+4}'].font = Font(bold=True, size=12, color=ACCENT_COLOR, name="맑은 고딕")
        ws_month[f'C{summary_row+4}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Borders for summary
        for r in range(summary_row+2, summary_row+5):
            for c in ['B', 'C']:
                ws_month[f'{c}{r}'].border = Border(
                    left=Side(style='medium'), right=Side(style='medium'),
                    top=Side(style='medium'), bottom=Side(style='medium')
                )
        
        # Transaction detail section - still showing data but noting it's from T_RawData
        detail_start = summary_row + 7
        ws_month[f'B{detail_start}'] = "거래 내역 (T_RawData 참조 - 모든 거래)"
        ws_month[f'B{detail_start}'].font = Font(size=12, bold=True, color=HEADER_COLOR, name="맑은 고딕")
        ws_month.merge_cells(f'B{detail_start}:J{detail_start}')
        
        # Headers
        headers = ['날짜', '금액', '구분', '대분류', '소분류', '내용', '결제수단', '메모', 'Flow_Filter']
        headers_row = detail_start + 1
        for col_idx, header in enumerate(headers, 2):
            cell = ws_month.cell(row=headers_row, column=col_idx, value=header)
            cell.fill = PatternFill(start_color=TABLE_HEADER, end_color=TABLE_HEADER, fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="맑은 고딕")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Get month data for display
        month_data = df[df['year_month'] == month_period].copy()
        
        # Write transaction data - ALL AS FORMULAS referencing T_RawData
        for idx, row in enumerate(month_data.itertuples(), headers_row + 1):
            # All columns reference T_RawData using the tracked row number
            raw_row = row.raw_row
            
            # 날짜 - Reference to T_RawData column A
            ws_month.cell(row=idx, column=2, value=f"='📋 T_RawData'!A{raw_row}")
            ws_month.cell(row=idx, column=2).number_format = 'YYYY-MM-DD'
            
            # 금액 - Reference to T_RawData column G
            ws_month.cell(row=idx, column=3, value=f"='📋 T_RawData'!G{raw_row}")
            ws_month.cell(row=idx, column=3).number_format = '₩#,##0'
            
            # 구분 - Reference to T_RawData column C
            ws_month.cell(row=idx, column=4, value=f"='📋 T_RawData'!C{raw_row}")
            
            # 대분류 - Reference to T_RawData column D
            ws_month.cell(row=idx, column=5, value=f"='📋 T_RawData'!D{raw_row}")
            
            # 소분류 - Reference to T_RawData column E
            ws_month.cell(row=idx, column=6, value=f"='📋 T_RawData'!E{raw_row}")
            
            # 내용 - Reference to T_RawData column F
            ws_month.cell(row=idx, column=7, value=f"='📋 T_RawData'!F{raw_row}")
            
            # 결제수단 - Reference to T_RawData column H
            ws_month.cell(row=idx, column=8, value=f"='📋 T_RawData'!H{raw_row}")
            
            # 메모 - Reference to T_RawData column I
            ws_month.cell(row=idx, column=9, value=f"='📋 T_RawData'!I{raw_row}")
            
            # Flow_Filter - Reference to T_RawData column J
            ws_month.cell(row=idx, column=10, value=f"='📋 T_RawData'!J{raw_row}")
            
            for col in range(2, 11):
                ws_month.cell(row=idx, column=col).font = Font(name="맑은 고딕", size=9)
                ws_month.cell(row=idx, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
        
        # Column widths
        ws_month.column_dimensions['B'].width = 12
        ws_month.column_dimensions['C'].width = 13
        ws_month.column_dimensions['D'].width = 10
        ws_month.column_dimensions['E'].width = 12
        ws_month.column_dimensions['F'].width = 12
        ws_month.column_dimensions['G'].width = 25
        ws_month.column_dimensions['H'].width = 15
        ws_month.column_dimensions['I'].width = 15
        ws_month.column_dimensions['J'].width = 12
    
    # ==================================================
    # Dashboard Summary with SUMIFS formulas
    # ==================================================
    ws_dash = wb.create_sheet("📊 Dashboard Summary", 1)
    ws_dash.sheet_view.showGridLines = False
    
    # Title
    ws_dash['B2'] = "가계부 종합 대시보드"
    ws_dash['B2'].font = Font(size=20, bold=True, color=HEADER_COLOR, name="맑은 고딕")
    ws_dash.merge_cells('B2:O2')
    ws_dash['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws_dash.row_dimensions[2].height = 35
    
    ws_dash['B3'] = f"분석 기간: {df['date'].min().strftime('%Y-%m-%d')} ~ {df['date'].max().strftime('%Y-%m-%d')}"
    ws_dash['B3'].font = Font(size=10, italic=True, color="7F7F7F", name="맑은 고딕")
    ws_dash.merge_cells('B3:O3')
    ws_dash['B3'].alignment = Alignment(horizontal='center')
    
    # ==================================================================
    # LEFT PANEL: Summary Information (Columns B-G)
    # ==================================================================
    
    summary_start = 5
    
    # Section 1: Main KPIs (Vertical Layout)
    ws_dash['B5'] = "📊 주요 지표"
    ws_dash['B5'].font = Font(size=14, bold=True, color=HEADER_COLOR, name="맑은 고딕")
    ws_dash.merge_cells('B5:G5')
    ws_dash['B5'].alignment = Alignment(horizontal='center', vertical='center')
    ws_dash['B5'].fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    ws_dash.row_dimensions[5].height = 25
    
    # KPI items (키/값 형태)
    kpi_items = [
        ("수입", f"=SUMIFS('📋 T_RawData'!G:G,'📋 T_RawData'!C:C,\"수입\",'📋 T_RawData'!J:J,1)", SUCCESS_COLOR),
        ("지출", f"=SUMIFS('📋 T_RawData'!G:G,'📋 T_RawData'!C:C,\"지출\",'📋 T_RawData'!J:J,1)", DANGER_COLOR),
        ("합계", "=C7-C8", ACCENT_COLOR),
    ]
    
    kpi_start_row = 7
    for idx, (label, formula, color) in enumerate(kpi_items):
        row = kpi_start_row + idx
        # Label
        ws_dash[f'B{row}'] = label
        ws_dash[f'B{row}'].font = Font(bold=True, size=11, name="맑은 고딕")
        ws_dash[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Value
        ws_dash[f'C{row}'] = formula
        ws_dash[f'C{row}'].number_format = '₩#,##0'
        ws_dash[f'C{row}'].font = Font(bold=True, size=12, color=color, name="맑은 고딕")
        ws_dash[f'C{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        for col in ['B', 'C']:
            ws_dash[f'{col}{row}'].border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Section 2: Category Breakdown (Top 5)
    cat_section_start = kpi_start_row + len(kpi_items) + 2
    ws_dash[f'B{cat_section_start}'] = "📊 카테고리별 지출 Top 5"
    ws_dash[f'B{cat_section_start}'].font = Font(size=12, bold=True, color=HEADER_COLOR, name="맑은 고딕")
    ws_dash.merge_cells(f'B{cat_section_start}:G{cat_section_start}')
    ws_dash[f'B{cat_section_start}'].alignment = Alignment(horizontal='center', vertical='center')
    ws_dash[f'B{cat_section_start}'].fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    
    # Category table headers
    cat_headers_row = cat_section_start + 1
    cat_headers = ['카테고리', '금액', '비율']
    for col_idx, header in enumerate(cat_headers, 2):
        cell = ws_dash.cell(row=cat_headers_row, column=col_idx, value=header)
        cell.fill = PatternFill(start_color=TABLE_HEADER, end_color=TABLE_HEADER, fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="맑은 고딕")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    # Calculate and display category data
    expense_df = df[df['type'] == '지출']
    if not expense_df.empty:
        cat_summary = expense_df.groupby('main_category')['amount'].sum().reset_index()
        cat_summary = cat_summary.sort_values('amount', ascending=False).head(5)
        total_expense = expense_df['amount'].sum()
        
        for idx, row_data in enumerate(cat_summary.itertuples(index=False), cat_headers_row + 1):
            # Category
            ws_dash.cell(row=idx, column=2, value=row_data.main_category)
            ws_dash.cell(row=idx, column=2).alignment = Alignment(horizontal='left', vertical='center')
            ws_dash.cell(row=idx, column=2).font = Font(name="맑은 고딕", size=10)
            
            # Amount
            ws_dash.cell(row=idx, column=3, value=row_data.amount)
            ws_dash.cell(row=idx, column=3).number_format = '₩#,##0'
            ws_dash.cell(row=idx, column=3).alignment = Alignment(horizontal='right', vertical='center')
            ws_dash.cell(row=idx, column=3).font = Font(name="맑은 고딕", size=10)
            
            # Percentage
            percentage = (row_data.amount / total_expense) if total_expense > 0 else 0
            ws_dash.cell(row=idx, column=4, value=percentage)
            ws_dash.cell(row=idx, column=4).number_format = '0.0%'
            ws_dash.cell(row=idx, column=4).alignment = Alignment(horizontal='center', vertical='center')
            ws_dash.cell(row=idx, column=4).font = Font(name="맑은 고딕", size=10)
            
            for col_idx in range(2, 5):
                ws_dash.cell(row=idx, column=col_idx).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
    
    # ==================================================================
    # RIGHT PANEL: Detailed Monthly Data (Columns I onwards)
    # ==================================================================
    
    # Title for detailed section
    ws_dash['I5'] = "월별 상세 데이터"
    ws_dash['I5'].font = Font(size=14, bold=True, color=HEADER_COLOR, name="맑은 고딕")
    ws_dash.merge_cells('I5:O5')
    ws_dash['I5'].alignment = Alignment(horizontal='center', vertical='center')
    ws_dash['I5'].fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    ws_dash.row_dimensions[5].height = 25
    
    # Monthly detail table headers
    detail_headers_row = 6
    detail_headers = ['월', '수입', '지출', '합계', '거래건수', '평균지출', '비고']
    for col_idx, header in enumerate(detail_headers, 9):  # Start from column I (9)
        cell = ws_dash.cell(row=detail_headers_row, column=col_idx, value=header)
        cell.fill = PatternFill(start_color=TABLE_HEADER, end_color=TABLE_HEADER, fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="맑은 고딕")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )
    
    # Monthly data with SUMIFS formulas
    for idx, month_period in enumerate(months, detail_headers_row + 1):
        month_str = str(month_period)
        year = month_period.year
        month = month_period.month
        
        # Calculate next month and year correctly
        if month == 12:
            next_month = 1
            next_year = year + 1
        else:
            next_month = month + 1
            next_year = year
        
        # Month name (Column I)
        ws_dash.cell(row=idx, column=9, value=month_str)
        ws_dash.cell(row=idx, column=9).font = Font(name="맑은 고딕", size=10)
        ws_dash.cell(row=idx, column=9).alignment = Alignment(horizontal='center', vertical='center')
        
        # Income (Column J)
        income_formula = f'=SUMIFS(\'📋 T_RawData\'!G:G,\'📋 T_RawData\'!A:A,">="&DATE({year},{month},1),\'📋 T_RawData\'!A:A,"<"&DATE({next_year},{next_month},1),\'📋 T_RawData\'!C:C,"수입",\'📋 T_RawData\'!J:J,1)'
        ws_dash.cell(row=idx, column=10, value=income_formula)
        ws_dash.cell(row=idx, column=10).number_format = '₩#,##0'
        ws_dash.cell(row=idx, column=10).font = Font(name="맑은 고딕", size=10, color=SUCCESS_COLOR)
        ws_dash.cell(row=idx, column=10).alignment = Alignment(horizontal='right', vertical='center')
        
        # Expense (Column K)
        expense_formula = f'=SUMIFS(\'📋 T_RawData\'!G:G,\'📋 T_RawData\'!A:A,">="&DATE({year},{month},1),\'📋 T_RawData\'!A:A,"<"&DATE({next_year},{next_month},1),\'📋 T_RawData\'!C:C,"지출",\'📋 T_RawData\'!J:J,1)'
        ws_dash.cell(row=idx, column=11, value=expense_formula)
        ws_dash.cell(row=idx, column=11).number_format = '₩#,##0'
        ws_dash.cell(row=idx, column=11).font = Font(name="맑은 고딕", size=10, color=DANGER_COLOR)
        ws_dash.cell(row=idx, column=11).alignment = Alignment(horizontal='right', vertical='center')
        
        # Total (Column L)
        ws_dash.cell(row=idx, column=12, value=f"=J{idx}-K{idx}")
        ws_dash.cell(row=idx, column=12).number_format = '₩#,##0'
        ws_dash.cell(row=idx, column=12).font = Font(name="맑은 고딕", size=10, color=ACCENT_COLOR, bold=True)
        ws_dash.cell(row=idx, column=12).alignment = Alignment(horizontal='right', vertical='center')
        
        # Transaction Count (Column M)
        count_formula = f'=COUNTIFS(\'📋 T_RawData\'!A:A,">="&DATE({year},{month},1),\'📋 T_RawData\'!A:A,"<"&DATE({next_year},{next_month},1),\'📋 T_RawData\'!J:J,1)'
        ws_dash.cell(row=idx, column=13, value=count_formula)
        ws_dash.cell(row=idx, column=13).number_format = '0'
        ws_dash.cell(row=idx, column=13).font = Font(name="맑은 고딕", size=10)
        ws_dash.cell(row=idx, column=13).alignment = Alignment(horizontal='center', vertical='center')
        
        # Average Expense (Column N)
        ws_dash.cell(row=idx, column=14, value=f"=IF(M{idx}>0,K{idx}/M{idx},0)")
        ws_dash.cell(row=idx, column=14).number_format = '₩#,##0'
        ws_dash.cell(row=idx, column=14).font = Font(name="맑은 고딕", size=10)
        ws_dash.cell(row=idx, column=14).alignment = Alignment(horizontal='right', vertical='center')
        
        # Note (Column O) - empty for now
        ws_dash.cell(row=idx, column=15, value="")
        ws_dash.cell(row=idx, column=15).font = Font(name="맑은 고딕", size=9)
        ws_dash.cell(row=idx, column=15).alignment = Alignment(horizontal='left', vertical='center')
        
        # Add borders
        for col_idx in range(9, 16):
            ws_dash.cell(row=idx, column=col_idx).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Column widths
    ws_dash.column_dimensions['B'].width = 12
    ws_dash.column_dimensions['C'].width = 16
    ws_dash.column_dimensions['D'].width = 10
    ws_dash.column_dimensions['I'].width = 12
    ws_dash.column_dimensions['J'].width = 14
    ws_dash.column_dimensions['K'].width = 14
    ws_dash.column_dimensions['L'].width = 14
    ws_dash.column_dimensions['M'].width = 10
    ws_dash.column_dimensions['N'].width = 14
    ws_dash.column_dimensions['O'].width = 20
    
    # ==================================================================
    # CHARTS - Add visual representation
    # ==================================================================
    
    # Chart 1: Monthly Financial Trend (Bar Chart)
    # Position: Below left panel
    monthly_chart = BarChart()
    monthly_chart.type = "col"
    monthly_chart.grouping = "clustered"
    monthly_chart.title = "월별 재정 추이"
    monthly_chart.y_axis.title = '금액 (원)'
    monthly_chart.x_axis.title = '월'
    monthly_chart.height = 12
    monthly_chart.width = 18
    monthly_chart.style = 11
    
    # Data references for monthly chart (from I:L columns)
    # Categories (months)
    months_ref = Reference(ws_dash, min_col=9, min_row=7, max_row=6+len(months))
    # Data series (Income, Expense, Net)
    monthly_data_ref = Reference(ws_dash, min_col=10, min_row=6, max_row=6+len(months), max_col=12)
    
    monthly_chart.add_data(monthly_data_ref, titles_from_data=True)
    monthly_chart.set_categories(months_ref)
    
    # Add chart to sheet - positioned below the summary tables
    ws_dash.add_chart(monthly_chart, "B25")
    
    # Chart 2: Category Expense Distribution (Donut/Pie Chart)
    # Position: Next to monthly chart
    category_chart = PieChart()
    category_chart.title = "카테고리별 지출 비중"
    category_chart.height = 12
    category_chart.width = 14
    category_chart.style = 10
    
    # Data references for category chart (from B:C columns, Top 5 categories)
    cat_labels_ref = Reference(ws_dash, min_col=2, min_row=cat_headers_row+1, max_row=cat_headers_row+5)
    cat_data_ref = Reference(ws_dash, min_col=3, min_row=cat_headers_row, max_row=cat_headers_row+5)
    
    category_chart.add_data(cat_data_ref, titles_from_data=True)
    category_chart.set_categories(cat_labels_ref)
    
    # Add data labels with percentage
    from openpyxl.chart.label import DataLabelList
    category_chart.dataLabels = DataLabelList()
    category_chart.dataLabels.showCatName = True
    category_chart.dataLabels.showPercent = True
    category_chart.dataLabels.showVal = False
    
    # Add chart to sheet - positioned next to monthly chart
    ws_dash.add_chart(category_chart, "K25")
    
    # ==================================================
    # Investment Tracking (Flow_Filter=0)
    # ==================================================
    ws_invest = wb.create_sheet("💰 Investment Tracking", len(wb.sheetnames))
    ws_invest.sheet_view.showGridLines = False
    
    ws_invest['B2'] = "투자 및 내부 이동 추적 (Flow_Filter=0)"
    ws_invest['B2'].font = Font(size=16, bold=True, color=WARNING_COLOR, name="맑은 고딕")
    ws_invest.merge_cells('B2:F2')
    ws_invest['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws_invest.row_dimensions[2].height = 30
    
    # Special transactions count
    ws_invest['B4'] = "특수 거래 총액"
    ws_invest['B4'].font = Font(size=11, bold=True, name="맑은 고딕")
    ws_invest['C4'] = f"=SUMIFS('📋 T_RawData'!G:G,'📋 T_RawData'!J:J,0)"
    ws_invest['C4'].number_format = '₩#,##0'
    ws_invest['C4'].font = Font(size=12, bold=True, color=WARNING_COLOR, name="맑은 고딕")
    
    # Monthly special transactions table
    ws_invest['B6'] = "월별 특수 거래 내역"
    ws_invest['B6'].font = Font(size=12, bold=True, color=HEADER_COLOR, name="맑은 고딕")
    
    headers = ['월', '특수거래액']
    for col_idx, header in enumerate(headers, 2):
        cell = ws_invest.cell(row=7, column=col_idx, value=header)
        cell.fill = PatternFill(start_color=TABLE_HEADER, end_color=TABLE_HEADER, fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="맑은 고딕")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for idx, month_period in enumerate(months, 8):
        month_str = str(month_period)
        year = month_period.year
        month = month_period.month
        
        ws_invest.cell(row=idx, column=2, value=month_str)
        special_formula = f'=SUMIFS(\'📋 T_RawData\'!G:G,\'📋 T_RawData\'!A:A,\">=\"&DATE({year},{month},1),\'📋 T_RawData\'!A:A,\"<\"&DATE({year},{month+1 if month < 12 else 1},1),\'📋 T_RawData\'!J:J,0)'
        ws_invest.cell(row=idx, column=3, value=special_formula)
        ws_invest.cell(row=idx, column=3).number_format = '₩#,##0'
        
        for col_idx in [2, 3]:
            ws_invest.cell(row=idx, column=col_idx).font = Font(name="맑은 고딕", size=10)
            ws_invest.cell(row=idx, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
            ws_invest.cell(row=idx, column=col_idx).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Save
    print(f"💾 Saving to: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    print("✅ Formula-Linked Excel Report created successfully!")
    print(f"\n📊 Summary:")
    print(f"   - T_RawData Engine: {len(df)} rows with Flow_Filter formulas")
    print(f"   - Monthly Sheets: {len(months)} (formula-linked)")
    print(f"   - Dashboard: SUMIFS formulas (Flow_Filter=1)")
    print(f"   - Investment Tracking: SUMIFS formulas (Flow_Filter=0)")
    print(f"\n✅ File saved: {OUTPUT_FILE}")

if __name__ == "__main__":
    create_formula_linked_report()
