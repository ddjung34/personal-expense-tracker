import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

# ===== DATA LOADING =====
def load_data_from_google_sheets():
    """Load data from Google Sheets"""
    try:
        # Google Sheets credentials
        SCOPES = [
            'https://www.googleapis.com/auth/spreadsheets.readonly',
            'https://www.googleapis.com/auth/drive.readonly'
        ]
        
        SERVICE_ACCOUNT_FILE = r"C:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\service_account.json"
        
        # Authenticate
        creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        client = gspread.authorize(creds)
        
        # Open spreadsheet
        SPREADSHEET_ID = "1DqpTecTdpRKsXTPImM4iKPT2V-KeJixG85-K6MuLOWY"
        spreadsheet = client.open_by_key(SPREADSHEET_ID)
        
        # Get DB_Raw sheet
        sheet = spreadsheet.worksheet("DB_Raw")
        
        # Get all values
        data = sheet.get_all_values()
        
        # Convert to DataFrame
        df = pd.DataFrame(data[1:], columns=data[0])
        
        # Data cleaning
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        df['amount'] = pd.to_numeric(df['amount'], errors='coerce')
        
        # Remove rows with invalid data
        df = df.dropna(subset=['date', 'amount', 'type'])
        
        # Sort by date
        df = df.sort_values('date')
        
        print(f"✅ Google Sheets 데이터 로드 성공: {len(df)}건")
        return df
        
    except Exception as e:
        print(f"⚠️ Google Sheets 연결 실패: {e}")
        print("로컬 Excel 파일로 전환...")
        
        # Fallback to local file
        file_path = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-07~2025-12-07_v5_LinkDB.xlsx"
        df = pd.read_excel(file_path, sheet_name="DB_Raw", engine='openpyxl')
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        df['amount'] = pd.to_numeric(df['amount'], errors='coerce')
        df = df.dropna(subset=['date', 'amount', 'type'])
        df = df.sort_values('date')
        print(f"✅ 로컬 데이터 로드 성공: {len(df)}건")
        return df


def create_excel_dashboard():
    """Create Excel Dashboard matching Streamlit layout"""
    
    print("\n" + "="*60)
    print("📊 Excel Dashboard Report 생성 시작")
    print("="*60 + "\n")
    
    # Load data
    df = load_data_from_google_sheets()
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Colors
    HEADER_COLOR = "2C3E50"
    ACCENT_COLOR = "3498DB"
    SUCCESS_COLOR = "27AE60"
    DANGER_COLOR = "E74C3C"
    LIGHT_BG = "ECF0F1"
    
    # ===== SHEET 1: Dashboard =====
    ws = wb.create_sheet("📊 Dashboard", 0)
    ws.sheet_view.showGridLines = False
    
    # Title
    ws['B2'] = "재정 상태 통합 대시보드"
    ws['B2'].font = Font(size=20, bold=True, color=HEADER_COLOR, name="맑은 고딕")
    ws.merge_cells('B2:K2')
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 35
    
    # Subtitle
    ws['B3'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['B3'].font = Font(size=10, italic=True, color="7F7F7F", name="맑은 고딕")
    ws.merge_cells('B3:K3')
    ws['B3'].alignment = Alignment(horizontal='center')
    
    # ===== KPI Section =====
    print("📈 KPI 계산 중...")
    
    total_income = df[df['type'] == '수입']['amount'].sum()
    total_expense = df[df['type'] == '지출']['amount'].sum()
    net_income = total_income - total_expense
    
    # Find top category
    expense_df = df[df['type'] == '지출']
    if not expense_df.empty:
        top_cat = expense_df.groupby('main_category')['amount'].sum().idxmax()
        top_cat_amount = expense_df.groupby('main_category')['amount'].sum().max()
    else:
        top_cat = "N/A"
        top_cat_amount = 0
    
    kpi_row = 5
    ws.row_dimensions[kpi_row].height = 25
    ws.row_dimensions[kpi_row + 1].height = 40
    
    # KPI 1: Total Income
    ws[f'B{kpi_row}'] = "💰 총 수입"
    ws[f'B{kpi_row}'].fill = PatternFill(start_color=SUCCESS_COLOR, end_color=SUCCESS_COLOR, fill_type="solid")
    ws[f'B{kpi_row}'].font = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")
    ws[f'B{kpi_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'B{kpi_row}:C{kpi_row}')
    
    ws[f'B{kpi_row+1}'] = total_income
    ws[f'B{kpi_row+1}'].number_format = '₩#,##0'
    ws[f'B{kpi_row+1}'].font = Font(bold=True, size=16, color=SUCCESS_COLOR, name="맑은 고딕")
    ws[f'B{kpi_row+1}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'B{kpi_row+1}'].fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    ws.merge_cells(f'B{kpi_row+1}:C{kpi_row+1}')
    
    # KPI 2: Total Expense
    ws[f'E{kpi_row}'] = "💸 총 지출"
    ws[f'E{kpi_row}'].fill = PatternFill(start_color=DANGER_COLOR, end_color=DANGER_COLOR, fill_type="solid")
    ws[f'E{kpi_row}'].font = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")
    ws[f'E{kpi_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'E{kpi_row}:F{kpi_row}')
    
    ws[f'E{kpi_row+1}'] = total_expense
    ws[f'E{kpi_row+1}'].number_format = '₩#,##0'
    ws[f'E{kpi_row+1}'].font = Font(bold=True, size=16, color=DANGER_COLOR, name="맑은 고딕")
    ws[f'E{kpi_row+1}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'E{kpi_row+1}'].fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    ws.merge_cells(f'E{kpi_row+1}:F{kpi_row+1}')
    
    # KPI 3: Net Income
    ws[f'H{kpi_row}'] = "✅ 순수익"
    ws[f'H{kpi_row}'].fill = PatternFill(start_color=ACCENT_COLOR, end_color=ACCENT_COLOR, fill_type="solid")
    ws[f'H{kpi_row}'].font = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")
    ws[f'H{kpi_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'H{kpi_row}:I{kpi_row}')
    
    ws[f'H{kpi_row+1}'] = net_income
    ws[f'H{kpi_row+1}'].number_format = '₩#,##0'
    ws[f'H{kpi_row+1}'].font = Font(bold=True, size=16, color=ACCENT_COLOR, name="맑은 고딕")
    ws[f'H{kpi_row+1}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'H{kpi_row+1}'].fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    ws.merge_cells(f'H{kpi_row+1}:I{kpi_row+1}')
    
    # KPI 4: Top Category
    ws[f'K{kpi_row}'] = "🔥 최대 지출"
    ws[f'K{kpi_row}'].fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    ws[f'K{kpi_row}'].font = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")
    ws[f'K{kpi_row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws[f'K{kpi_row+1}'] = f"{top_cat}\n{top_cat_amount:,.0f}원"
    ws[f'K{kpi_row+1}'].font = Font(bold=True, size=12, color=HEADER_COLOR, name="맑은 고딕")
    ws[f'K{kpi_row+1}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws[f'K{kpi_row+1}'].fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    
    # Borders for KPIs
    for kpi_col in ['B', 'E', 'H', 'K']:
        for r in [kpi_row, kpi_row+1]:
            end_col = chr(ord(kpi_col)+1) if kpi_col != 'K' else 'K'
            for col in range(ord(kpi_col), ord(end_col)+1):
                cell = ws[f'{chr(col)}{r}']
                cell.border = Border(
                    left=Side(style='medium'), right=Side(style='medium'),
                    top=Side(style='medium'), bottom=Side(style='medium')
                )
    
    # ===== Monthly Trend Data & Chart =====
    print("📊 월별 데이터 집계 중...")
    
    monthly_data = df.groupby([pd.Grouper(key='date', freq='ME'), 'type'])['amount'].sum().reset_index()
    monthly_pivot = monthly_data.pivot(index='date', columns='type', values='amount').fillna(0)
    
    if '수입' not in monthly_pivot.columns:
        monthly_pivot['수입'] = 0
    if '지출' not in monthly_pivot.columns:
        monthly_pivot['지출'] = 0
    
    monthly_pivot['순수익'] = monthly_pivot['수입'] - monthly_pivot['지출']
    monthly_pivot = monthly_pivot.reset_index()
    monthly_pivot['월'] = monthly_pivot['date'].dt.strftime('%Y-%m')
    
    # LEFT PANEL: Monthly Trend Chart
    chart_start = 8
    ws[f'B{chart_start}'] = "📈 재정 흐름 분석 (시간)"
    ws[f'B{chart_start}'].font = Font(size=14, bold=True, color=HEADER_COLOR, name="맑은 고딕")
    ws.merge_cells(f'B{chart_start}:G{chart_start}')
    
    # Write monthly data with formulas
    headers_row = chart_start + 1
    headers = ['월', '수입', '지출', '순수익']
    for col_idx, header in enumerate(headers, 2):
        cell = ws.cell(row=headers_row, column=col_idx, value=header)
        cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="맑은 고딕")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for idx, row_data in enumerate(monthly_pivot.itertuples(), headers_row + 1):
        ws.cell(row=idx, column=2, value=row_data.월)
        ws.cell(row=idx, column=3, value=row_data.수입).number_format = '₩#,##0'
        ws.cell(row=idx, column=4, value=row_data.지출).number_format = '₩#,##0'
        
        # Use SUM formula for 순수익
        ws.cell(row=idx, column=5, value=f"=SUM(C{idx},-D{idx})").number_format = '₩#,##0'
        
        for col in range(2, 6):
            ws.cell(row=idx, column=col).font = Font(name="맑은 고딕", size=9)
            ws.cell(row=idx, column=col).alignment = Alignment(horizontal='center')
            ws.cell(row=idx, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Create Bar Chart with better visibility
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "월별 재정 추이 (전체 기간)"
    chart.y_axis.title = '금액 (원)'
    chart.x_axis.title = '월'
    chart.height = 15
    chart.width = 20
    chart.legend.position = 'r'
    chart.dataLabels = None
    
    data_ref = Reference(ws, min_col=3, min_row=headers_row, max_row=headers_row + len(monthly_pivot), max_col=5)
    cats_ref = Reference(ws, min_col=2, min_row=headers_row + 1, max_row=headers_row + len(monthly_pivot))
    
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.grouping = "clustered"
    
    chart_row = headers_row + len(monthly_pivot) + 2
    ws.add_chart(chart, f"B{chart_row}")
    
    # RIGHT PANEL: Category Analysis
    print("🥧 카테고리 분석 중...")
    
    cat_start_col = 9  # Column I
    ws.cell(row=chart_start, column=cat_start_col, value="💸 지출 구조 분석 (비중)")
    ws.cell(row=chart_start, column=cat_start_col).font = Font(size=14, bold=True, color=HEADER_COLOR, name="맑은 고딕")
    ws.merge_cells(start_row=chart_start, start_column=cat_start_col, end_row=chart_start, end_column=cat_start_col+3)
    
    if not expense_df.empty:
        category_expense = expense_df.groupby('main_category')['amount'].sum().reset_index()
        category_expense = category_expense.sort_values('amount', ascending=False).head(10)
        
        # Write category data for pie chart
        cat_headers_row = chart_start + 1
        ws.cell(row=cat_headers_row, column=cat_start_col, value="카테고리").font = Font(bold=True, name="맑은 고딕", size=9)
        ws.cell(row=cat_headers_row, column=cat_start_col+1, value="금액").font = Font(bold=True, name="맑은 고딕", size=9)
        
        for idx, row_data in enumerate(category_expense.itertuples(), cat_headers_row + 1):
            ws.cell(row=idx, column=cat_start_col, value=row_data.main_category).font = Font(name="맑은 고딕", size=8)
            ws.cell(row=idx, column=cat_start_col+1, value=row_data.amount).number_format = '₩#,##0'
            ws.cell(row=idx, column=cat_start_col+1).font = Font(name="맑은 고딕", size=8)
        
        # Create Pie Chart
        pie = PieChart()
        pie.title = "카테고리별 지출 비중"
        pie.height = 8
        pie.width = 12
        
        pie_labels = Reference(ws, min_col=cat_start_col, min_row=cat_headers_row+1, max_row=cat_headers_row+len(category_expense))
        pie_data = Reference(ws, min_col=cat_start_col+1, min_row=cat_headers_row, max_row=cat_headers_row+len(category_expense))
        
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(pie_labels)
        
        ws.add_chart(pie, f"I{chart_start + 2}")
        
        # Sub-category analysis (Top 5)
        sub_start_row = chart_start + 16
        ws.cell(row=sub_start_row, column=cat_start_col, value="상위 5개 세부 카테고리")
        ws.cell(row=sub_start_row, column=cat_start_col).font = Font(size=12, bold=True, color=HEADER_COLOR, name="맑은 고딕")
        ws.merge_cells(start_row=sub_start_row, start_column=cat_start_col, end_row=sub_start_row, end_column=cat_start_col+3)
        
        sub_expense = expense_df.groupby('sub_category')['amount'].sum().reset_index()
        sub_expense = sub_expense.sort_values('amount', ascending=False).head(5)
        
        # Write sub-category data
        sub_headers_row = sub_start_row + 1
        ws.cell(row=sub_headers_row, column=cat_start_col, value="세부 카테고리").font = Font(bold=True, name="맑은 고딕", size=9)
        ws.cell(row=sub_headers_row, column=cat_start_col+1, value="금액").font = Font(bold=True, name="맑은 고딕", size=9)
        
        for idx, row_data in enumerate(sub_expense.itertuples(), sub_headers_row + 1):
            ws.cell(row=idx, column=cat_start_col, value=row_data.sub_category).font = Font(name="맑은 고딕", size=8)
            ws.cell(row=idx, column=cat_start_col+1, value=row_data.amount).number_format = '₩#,##0'
            ws.cell(row=idx, column=cat_start_col+1).font = Font(name="맑은 고딕", size=8)
        
        # Create horizontal bar chart for sub-categories
        sub_chart = BarChart()
        sub_chart.type = "bar"
        sub_chart.style = 11
        sub_chart.title = "지출액 기준 상위 5개 세부 카테고리"
        sub_chart.height = 8
        sub_chart.width = 12
        
        sub_data_ref = Reference(ws, min_col=cat_start_col+1, min_row=sub_headers_row, max_row=sub_headers_row+len(sub_expense))
        sub_cats_ref = Reference(ws, min_col=cat_start_col, min_row=sub_headers_row+1, max_row=sub_headers_row+len(sub_expense))
        
        sub_chart.add_data(sub_data_ref, titles_from_data=True)
        sub_chart.set_categories(sub_cats_ref)
        
        ws.add_chart(sub_chart, f"I{sub_start_row + 2}")
    
    # ===== Transaction Table (Below Charts) =====
    print("📋 거래 내역 작성 중...")
    
    table_start = chart_row + 30
    ws[f'B{table_start}'] = "📋 거래 내역"
    ws[f'B{table_start}'].font = Font(size=14, bold=True, color=HEADER_COLOR, name="맑은 고딕")
    ws.merge_cells(f'B{table_start}:K{table_start}')
    
    # Headers
    table_headers = ['날짜', '구분', '대분류', '소분류', '내용', '금액', '결제수단', '메모']
    table_headers_row = table_start + 1
    for col_idx, header in enumerate(table_headers, 2):
        cell = ws.cell(row=table_headers_row, column=col_idx, value=header)
        cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="맑은 고딕")
        cell.alignment = Alignment(horizontal='center')
    
    # Write all transaction data
    for idx, row in enumerate(df.itertuples(), table_headers_row + 1):
        ws.cell(row=idx, column=2, value=row.date.strftime('%Y-%m-%d'))
        ws.cell(row=idx, column=3, value=row.type)
        ws.cell(row=idx, column=4, value=row.main_category)
        ws.cell(row=idx, column=5, value=row.sub_category)
        ws.cell(row=idx, column=6, value=row.merchant)
        ws.cell(row=idx, column=7, value=row.amount).number_format = '₩#,##0'
        ws.cell(row=idx, column=8, value=row.payment_method)
        ws.cell(row=idx, column=9, value=row.memo)
        
        for col in range(2, 10):
            ws.cell(row=idx, column=col).font = Font(name="맑은 고딕", size=9)
            ws.cell(row=idx, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Auto-filter
    ws.auto_filter.ref = f"B{table_headers_row}:I{table_headers_row + len(df)}"
    
    # Column widths
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 14
    
    # ===== CREATE MONTHLY DETAIL SHEETS =====
    print("\n📅 월별 상세 시트 생성 중...")
    
    for month_period in monthly_pivot['date'].unique():
        month_str = pd.Timestamp(month_period).strftime('%Y-%m')
        month_name = pd.Timestamp(month_period).strftime('%m월')
        
        # Create monthly sheet
        ws_month = wb.create_sheet(month_name)
        ws_month.sheet_view.showGridLines = False
        
        # Title
        ws_month['B2'] = f"{month_str} 거래 내역"
        ws_month['B2'].font = Font(size=16, bold=True, color=HEADER_COLOR, name="맑은 고딕")
        ws_month.merge_cells('B2:H2')
        ws_month['B2'].alignment = Alignment(horizontal='center', vertical='center')
        ws_month.row_dimensions[2].height = 30
        
        # Filter month data
        month_df = df[df['date'].dt.to_period('M') == pd.Timestamp(month_period).to_period('M')]
        
        # Summary
        summary_row = 4
        ws_month[f'B{summary_row}'] = "월별 요약"
        ws_month[f'B{summary_row}'].font = Font(size=13, bold=True, color=ACCENT_COLOR, name="맑은 고딕")
        ws_month.merge_cells(f'B{summary_row}:D{summary_row}')
        
        income_sum = month_df[month_df['type'] == '수입']['amount'].sum()
        expense_sum = month_df[month_df['type'] == '지출']['amount'].sum()
        
        # Income
        ws_month[f'B{summary_row+2}'] = "💰 수입"
        ws_month[f'B{summary_row+2}'].fill = PatternFill(start_color=SUCCESS_COLOR, end_color=SUCCESS_COLOR, fill_type="solid")
        ws_month[f'B{summary_row+2}'].font = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")
        ws_month[f'B{summary_row+2}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws_month[f'C{summary_row+2}'] = income_sum
        ws_month[f'C{summary_row+2}'].number_format = '₩#,##0'
        ws_month[f'C{summary_row+2}'].font = Font(bold=True, size=12, color=SUCCESS_COLOR, name="맑은 고딕")
        ws_month[f'C{summary_row+2}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Expense
        ws_month[f'B{summary_row+3}'] = "💸 지출"
        ws_month[f'B{summary_row+3}'].fill = PatternFill(start_color=DANGER_COLOR, end_color=DANGER_COLOR, fill_type="solid")
        ws_month[f'B{summary_row+3}'].font = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")
        ws_month[f'B{summary_row+3}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws_month[f'C{summary_row+3}'] = expense_sum
        ws_month[f'C{summary_row+3}'].number_format = '₩#,##0'
        ws_month[f'C{summary_row+3}'].font = Font(bold=True, size=12, color=DANGER_COLOR, name="맑은 고딕")
        ws_month[f'C{summary_row+3}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Total with SUM formula
        ws_month[f'B{summary_row+4}'] = "✅ 합계"
        ws_month[f'B{summary_row+4}'].fill = PatternFill(start_color=ACCENT_COLOR, end_color=ACCENT_COLOR, fill_type="solid")
        ws_month[f'B{summary_row+4}'].font = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")
        ws_month[f'B{summary_row+4}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws_month[f'C{summary_row+4}'] = f"=SUM(C{summary_row+2},-C{summary_row+3})"
        ws_month[f'C{summary_row+4}'].number_format = '₩#,##0'
        ws_month[f'C{summary_row+4}'].font = Font(bold=True, size=12, color=ACCENT_COLOR, name="맑은 고딕")
        ws_month[f'C{summary_row+4}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Borders for summary
        for r in range(summary_row+2, summary_row+5):
            for c in ['B', 'C']:
                ws_month[f'{c}{r}'].border = Border(
                    left=Side(style='medium'), right=Side(style='medium'),
                    top=Side(style='medium'), bottom=Side(style='medium')
                )
        
        # Transaction details
        detail_start = summary_row + 7
        ws_month[f'B{detail_start}'] = "거래 내역"
        ws_month[f'B{detail_start}'].font = Font(size=12, bold=True, color=HEADER_COLOR, name="맑은 고딕")
        
        # Headers
        detail_headers = ['날짜', '구분', '대분류', '소분류', '내용', '금액', '결제수단', '메모']
        detail_headers_row = detail_start + 1
        for col_idx, header in enumerate(detail_headers, 2):
            cell = ws_month.cell(row=detail_headers_row, column=col_idx, value=header)
            cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="맑은 고딕")
            cell.alignment = Alignment(horizontal='center')
        
        # Write monthly transactions
        for idx, row in enumerate(month_df.itertuples(), detail_headers_row + 1):
            ws_month.cell(row=idx, column=2, value=row.date.strftime('%Y-%m-%d'))
            ws_month.cell(row=idx, column=3, value=row.type)
            ws_month.cell(row=idx, column=4, value=row.main_category)
            ws_month.cell(row=idx, column=5, value=row.sub_category)
            ws_month.cell(row=idx, column=6, value=row.merchant)
            ws_month.cell(row=idx, column=7, value=row.amount).number_format = '₩#,##0'
            ws_month.cell(row=idx, column=8, value=row.payment_method)
            ws_month.cell(row=idx, column=9, value=row.memo)
            
            for col in range(2, 10):
                ws_month.cell(row=idx, column=col).font = Font(name="맑은 고딕", size=9)
                ws_month.cell(row=idx, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
        
        # Column widths
        ws_month.column_dimensions['B'].width = 12
        ws_month.column_dimensions['C'].width = 13
        ws_month.column_dimensions['D'].width = 12
        ws_month.column_dimensions['E'].width = 14
        ws_month.column_dimensions['F'].width = 25
        ws_month.column_dimensions['G'].width = 14
        ws_month.column_dimensions['H'].width = 14
        ws_month.column_dimensions['I'].width = 30
        
        print(f"  ✓ {month_name} 시트 생성 완료 ({len(month_df)}건)")
    
    # Save file
    output_file = f"c:\\Users\\JTC7\\Desktop\\01.Python Project\\01.Personal Expense Tracker\\01.Document\\{datetime.now().strftime('%Y%m%d')}_Dashboard_Report.xlsx"
    wb.save(output_file)
    
    print("\n" + "="*60)
    print(f"✅ Excel Dashboard Report 생성 완료!")
    print(f"📁 파일 위치: {output_file}")
    print(f"📊 대시보드 시트: 1개")
    print(f"📅 월별 상세 시트: {len(monthly_pivot)}개")
    print("="*60 + "\n")
    
    return output_file


if __name__ == "__main__":
    create_excel_dashboard()
