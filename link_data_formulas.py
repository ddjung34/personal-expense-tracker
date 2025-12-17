import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Use the ORIGINAL file as requested
SOURCE_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-07~2025-12-07.xlsx"
OUTPUT_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-07~2025-12-07_v5_LinkDB.xlsx"

HEADERS = [
    "date", "time", "type", "main_category", "sub_category", 
    "amount", "payment_method", "merchant", "memo"
]

# Source Sheet Name
SOURCE_SHEET_NAME = "Rawdata"

# Mapping: Target Header -> Source Column Letter (based on inspection)
# Source Headers: ['날짜', '시간', '수입/지출', '대분류', '소분류', '내용', '금액', '통화', '결제수단', '메모']
# Index 0->A, 1->B ...
# 날짜(0)=A, 시간(1)=B, 수입/지출(2)=C, 대분류(3)=D, 소분류(4)=E, 내용(5)=F, 금액(6)=G, 통화(7)=H, 결제수단(8)=I, 메모(9)=J
COL_MAP = {
    "date": "A",
    "time": "B",
    "type": "C",
    "main_category": "D",
    "sub_category": "E",
    "merchant": "F",
    "amount": "G",
    "payment_method": "I",
    "memo": "J"
}

def create_linked_db():
    print(f"Loading {SOURCE_FILE}...")
    wb = openpyxl.load_workbook(SOURCE_FILE)
    
    if SOURCE_SHEET_NAME not in wb.sheetnames:
        print(f"Error: {SOURCE_SHEET_NAME} not found.")
        return

    # Remove old DB_Raw if exists
    if "DB_Raw" in wb.sheetnames:
        del wb["DB_Raw"]
        
    ws_source = wb[SOURCE_SHEET_NAME]
    ws_db = wb.create_sheet("DB_Raw", 0)
    
    # max_row of source
    max_r = ws_source.max_row
    print(f"Source has {max_r} rows.")

    # 1. Setup Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    for idx, h in enumerate(HEADERS, 1):
        cell = ws_db.cell(row=1, column=idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 2. Write Formulas
    # Formula: ='가계부 내역'!A2
    print("Writing linking formulas...")
    
    # We'll go a bit beyond current max_row to allow for auto-update when new lines are added,
    # OR just stick to max_row as user requested user might drag it themselves.
    # User said: "Reflect intended content... based on Rawdata".
    # I will fill up to max_row for now.
    
    for row in range(2, max_r + 1):
        for col_idx, header in enumerate(HEADERS, 1):
            source_col_letter = COL_MAP.get(header)
            if source_col_letter:
                # Formula: ='가계부 내역'!A2
                # Note: Sheet name with spaces needs single quotes
                val = f"='{SOURCE_SHEET_NAME}'!{source_col_letter}{row}"
                ws_db.cell(row=row, column=col_idx, value=val)
                
    # Auto-adjust column widths
    ws_db.column_dimensions['A'].width = 12
    ws_db.column_dimensions['H'].width = 20

    print(f"Saving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print("Done.")

if __name__ == "__main__":
    create_linked_db()
