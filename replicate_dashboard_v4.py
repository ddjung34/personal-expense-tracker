"""
Dashboard 정밀 복제 v4 - 최종 해결 버전
1. 굵은 테두리 중복 제거 (전체 외곽만 적용)
2. 콤보 차트 (Bar + Line) 정석 구현
3. 하단 차트 직접 생성 (복제 대신)
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import Series

print("=" * 70)
print("Dashboard 정밀 복제 v4")
print("=" * 70)

wb = load_workbook(
    r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_수식연결_가계부엔진_최종.xlsx'
)

# Dashboard 시트 찾기
dashboard_sheet_name = None
for sheet_name in wb.sheetnames:
    if 'dashboard' in sheet_name.lower() and '복제' not in sheet_name and '분석' in sheet_name:
        dashboard_sheet_name = sheet_name
        break
if not dashboard_sheet_name:
    # Fallback
    for s in wb.sheetnames:
        if 'dashboard' in s.lower() and '최종' not in s and '복제' not in s:
            dashboard_sheet_name = s; break
    if not dashboard_sheet_name: dashboard_sheet_name = wb.sheetnames[1]

print(f"✅ 원본 시트: '{dashboard_sheet_name}'")
ws_original = wb[dashboard_sheet_name]

# [1] 기본 데이터 수집 (셀, 병합, 크기)
# ... (이전과 동일한 로직, 간소화)
min_row, max_row = 1, 60
min_col, max_col = 1, 25 # T열까지

cell_data = []
for row in range(min_row, max_row + 1):
    for col in range(min_col, max_col + 1):
        cell = ws_original.cell(row=row, column=col)
        # 테두리, 폰트, 채우기 정보 추출
        # (코드가 길어지므로 핵심 로직만 유지)
        b_data = {}
        if cell.border:
            for s in ['left', 'right', 'top', 'bottom']:
                side = getattr(cell.border, s)
                if side:
                    c_val = None
                    if side.color:
                        if hasattr(side.color, 'rgb') and isinstance(side.color.rgb, str):
                            tmp = side.color.rgb
                            # Validate Hex
                            if len(tmp) <= 8: # e.g. 'FF000000'
                                c_val = tmp
                    
                    b_data[s] = {'style': side.style, 'color': c_val}
        
        f_color = cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
        bg_color = cell.fill.start_color.index if cell.fill and cell.fill.start_color and cell.fill.fill_type else None
        
        if cell.value or bg_color or b_data:
            cell_data.append({
                'row': row, 'col': col, 'value': cell.value,
                'fmt': cell.number_format,
                'font': {'name': cell.font.name, 'sz': cell.font.size, 'b': cell.font.bold, 'color': f_color},
                'fill': {'c': bg_color, 't': cell.fill.fill_type},
                'align': {'h': cell.alignment.horizontal, 'v': cell.alignment.vertical, 'w': cell.alignment.wrap_text},
                'border': b_data
            })

merged_cells = list(ws_original.merged_cells.ranges)
row_heights = {r: ws_original.row_dimensions[r].height for r in range(min_row, max_row + 1) if ws_original.row_dimensions[r].height}
col_widths = {get_column_letter(c): ws_original.column_dimensions[get_column_letter(c)].width for c in range(min_col, max_col + 1) if ws_original.column_dimensions[get_column_letter(c)].width}

# [2] 새 시트 생성
if '📊 Dashboard_v4' in wb.sheetnames: del wb['📊 Dashboard_v4']
ws_new = wb.create_sheet('📊 Dashboard_v4')
ws_new.sheet_view.showGridLines = False

# [3] 복제 적용
for r, h in row_heights.items(): ws_new.row_dimensions[r].height = h
for c, w in col_widths.items(): ws_new.column_dimensions[c].width = w
# for m in merged_cells: ws_new.merge_cells(str(m)) # Moved to end

for d in cell_data:
    c = ws_new.cell(d['row'], d['col'])
    c.value = d['value']
    if d['fmt']: c.number_format = d['fmt']
    # Safe Font Color
    f_color = d['font']['color']
    if f_color and not isinstance(f_color, str): f_color = None # Ensure string or None
    
    c.font = Font(name=d['font']['name'], size=d['font']['sz'], bold=d['font']['b'], color=f_color)
    if d['fill']['c']: c.fill = PatternFill(start_color=d['fill']['c'], end_color=d['fill']['c'], fill_type=d['fill']['t'])
    c.alignment = Alignment(horizontal=d['align']['h'], vertical=d['align']['v'], wrap_text=d['align']['w'])
    
    borders = {}
    for k, v in d['border'].items():
        if v: borders[k] = Side(style=v['style'], color=v['color'])
    if borders: c.border = Border(**borders)

# [3.5] 병합 적용 (값 입력 후)
for m in merged_cells: ws_new.merge_cells(str(m))

# [4] Q48 테두리 보정
for r in range(38, 50): # 대략적 범위
    c = ws_new.cell(r, 17) # Q열
    if c.value:
        cur = c.border
        c.border = Border(left=cur.left, top=cur.top, bottom=cur.bottom, right=Side(style='thin'))

# [5] 굵은 테두리 (범위 재조정)
# 사용자 요청: "주요지표부터 소비지표 아래부분에만"
# 상단 섹션: C3 ~ T26 (여기서 T26은 차트 포함 전체)
# 하단 섹션: C27 ~ T52
thick = Side(style='medium')
sections = ['C3:T25', 'C27:T52'] # 행 번호 미세 조정

for rng_str in sections:
    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(rng_str)
    # Top
    for c in range(min_col, max_col+1):
        ws_new.cell(min_row, c).border = Border(top=thick, bottom=ws_new.cell(min_row, c).border.bottom, left=ws_new.cell(min_row, c).border.left, right=ws_new.cell(min_row, c).border.right)
    # Bottom
    for c in range(min_col, max_col+1):
        ws_new.cell(max_row, c).border = Border(bottom=thick, top=ws_new.cell(max_row, c).border.top, left=ws_new.cell(max_row, c).border.left, right=ws_new.cell(max_row, c).border.right)
    # Left
    for r in range(min_row, max_row+1):
        ws_new.cell(r, min_col).border = Border(left=thick, top=ws_new.cell(r, min_col).border.top, bottom=ws_new.cell(r, min_col).border.bottom, right=ws_new.cell(r, min_col).border.right)
    # Right
    for r in range(min_row, max_row+1):
        ws_new.cell(r, max_col).border = Border(right=thick, top=ws_new.cell(r, max_col).border.top, bottom=ws_new.cell(r, max_col).border.bottom, left=ws_new.cell(r, max_col).border.left)

# [6] 콤보 차트 생성 (BarChart + LineChart 결합 방식)
print("차트 생성 중...")

# 1. 막대 차트 (수입/지출)
c1 = BarChart()
c1.type = "col"
c1.grouping = "clustered"
c1.overlap = 100
c1.y_axis.title = '금액'

# 데이터 (수입, 지출) - D, E열 (9행 헤더 포함)
data = Reference(ws_new, min_col=4, min_row=9, max_col=5, max_row=23)
c1.add_data(data, titles_from_data=True)
cats = Reference(ws_new, min_col=3, min_row=10, max_row=23) # 날짜 데이터 (10행부터)
c1.set_categories(cats)

# 시리즈 이름 설정 삭제 (자동)
# c1.series[0].title = "수입" 
# c1.series[1].title = "지출"

# 2. 선 차트 (합계)
c2 = LineChart()
# 데이터 (합계) - F열 (9행 헤더 포함)
data2 = Reference(ws_new, min_col=6, min_row=9, max_col=6, max_row=23)
c2.add_data(data2, titles_from_data=True)
# c2.series[0].title = "합계"

# 선 스타일 (노란색)
# 복잡한 스타일 생략하고 기본 생성 후 결합
# c2.series[0].graphicalProperties.line.solidFill = "FFC000" (에러 위험 있어 생략)

# 3. 차트 결합
c1 += c2 # BarChart에 LineChart 추가

# 위치 설정
c1.anchor = "H6"
c1.height = 13
c1.width = 18 # 너비 조정 (T열까지 꽉 차게)
c1.title = "주요 지표 추이"

ws_new.add_chart(c1)
print("콤보 차트 생성 완료")

# [7] 하단 차트 직접 생성 (복제 X)
# 원본 차트를 분석하지 않고, 이미지 기반으로 '가로 막대형' 차트 생성
# 지출 구조 차트 (Top 10 등)

# 데이터 추정: 하단 Top 10 테이블 (B38:F48)
# 카테고리(C열), 비율(E열) 사용으로 추정됨

c3 = BarChart()
c3.type = "bar" # 가로 막대
c3.title = "지출 구조 차트"
c3.style = 10 # 적당한 스타일

# 데이터: 비율 (E열 37~47) (37행 헤더)
# 카테고리: (C열 38~47)
data3 = Reference(ws_new, min_col=5, min_row=37, max_col=5, max_row=47) 
cats3 = Reference(ws_new, min_col=3, min_row=38, max_row=47)

c3.add_data(data3, titles_from_data=True)
c3.set_categories(cats3)
# c3.series[0].title = "지출 비중"

c3.anchor = "K30"
c3.height = 15
c3.width = 16

ws_new.add_chart(c3)
print("하단 차트 생성 완료")

# [8] 저장
output_path = r'c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\20251214_Dashboard_v4.xlsx'
wb.save(output_path)
print(f"저장 완료: {output_path}")
