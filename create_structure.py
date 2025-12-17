import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Configuration
SOURCE_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-07~2025-12-07.xlsx"
TARGET_FILE = r"c:\Users\JTC7\Desktop\01.Python Project\01.Personal Expense Tracker\01.Document\2024-12-07~2025-12-07_v2.xlsx"

HEADERS = [
    "date", "time", "type", "main_category", "sub_category", 
    "amount", "payment_method", "merchant", "memo"
]

def create_db_sheet():
    print(f"Loading file: {SOURCE_FILE}")
    try:
        wb = openpyxl.load_workbook(SOURCE_FILE)
    except FileNotFoundError:
        print("Error: File not found.")
        return

    # Check if DB_Raw exists, if not create it
    if "DB_Raw" in wb.sheetnames:
        print("Sheet 'DB_Raw' already exists. Skipping creation.")
        ws = wb["DB_Raw"]
    else:
        print("Creating new sheet 'DB_Raw'...")
        ws = wb.create_sheet("DB_Raw", 0) # Create at first position

    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    # Write Headers
    for col_num, header_title in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Set column widths adjustments (optional)
    ws.column_dimensions['A'].width = 12 # Date
    ws.column_dimensions['D'].width = 15 # Main Category
    ws.column_dimensions['H'].width = 20 # Merchant

    print(f"Saving new file to: {TARGET_FILE}")
    wb.save(TARGET_FILE)
    print("Done! You can now open the _v2.xlsx file.")

if __name__ == "__main__":
    create_db_sheet()
